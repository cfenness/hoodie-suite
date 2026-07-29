#!/usr/bin/env python3
"""
server.py — local agent that embeds the pulls into the dashboard.

Run this next to hoodie_mdm.html and ttb_cola_scraper.py:

    pip install flask requests beautifulsoup4
    python server.py
    # open http://127.0.0.1:8765

The dashboard auto-detects the agent (GET /api/health). When present:
  - it loads real datasets (/api/datasets) and run history (/api/runs)
  - "Run now" / "Run all" / the scheduler POST /api/run and execute REAL pulls:
        fl-items / fl-outlets  -> live Florida DBPR/ABT CSV extracts
        ttb-cola               -> ttb_cola_scraper.scrape(...)
When the agent is absent the dashboard falls back to its built-in preview.

State is persisted to ./agent_state/ (datasets.json, runs.json, cola CSV).
"""
import csv, gzip, io, json, os, random, re, time, types, urllib.request, datetime, threading, logging
import subprocess, sys
from flask import Flask, request, jsonify, send_file, Response, redirect, session

import ttb_cola_scraper as cola   # the scraper you generated
import abc_fws_scraper as abc      # ABC FWS directional inventory tracker (BigCommerce)
import specs_scraper as specs      # Spec's directional tracker (Next.js, via Bright Data)
import binnys_scraper as binnys    # Binny's directional tracker (Algolia feed, no Bright Data)
# shopify_scraper retired → Shopify runs INSIDE the census sweep (off_premise.national_sweep("shopify")),
# conn 'shopify' routed through the registry. One Shopify code path, not a parallel app-side one.
# Instacart: the FREE Playwright connector (instacart.Instacart) is the active source — no Bright Data,
# no proxy (proven landing per instacart-free-verify CI). Imported lazily in instacart_pull so a slim
# image without the browser lib doesn't break server import. The old BD dataset scraper is archived.
import analyze                      # data-reader brain behind "Overlay your data"
import planogram                    # benchmark + shelf-vision + pitch behind the Planogram app
import hi_analyst                   # the real Claude analyst behind Hoodie Intelligence Q&A
import prism                        # data contract behind the Prism mobile app
import places                       # restaurant / on-premise-accounts connector (Orlando first)
import census                       # US Census ACS demographics — reference-data connector (by county)
import derive                       # field-derivation / dictionary compiler (rule → DuckDB SQL expression)
import enrich                       # join reference data (census) onto the outlet master by county
import tx_tabc                      # Texas TABC licenses (Socrata) — TX outlets + companies by county
import master                       # unify per-state outlet pulls into one normalized outlets_master
import chicago                      # Chicago liquor-licensed outlets (Socrata) — IL outlets + companies, geocoded
import ct_dcp                       # Connecticut liquor (Socrata) — CT premises + brand/supplier registry
import socrata_outlets              # GENERIC Socrata outlet connector — NY/CO/MO… as config, not modules
import total_wine                    # Total Wine & More direct catalog (mobile-UA + sitemap + microdata)
import vtinfo                        # brand → retailer distribution via the VTInfo/VIP "where to buy" finder
import ab_locator                    # Anheuser-Busch InBev retailer locator (beertech GraphQL) — universal-outlet spine
import warehouse                    # Parquet-on-Tigris (or local) queried by DuckDB
import upc                          # UPC/EAN QC + owned prefix->owner crosswalk (deterministic + inference)
import auth_gate                    # Google OIDC login gate (active only when configured)
import source_registry             # THE canonical source list — /api/run dispatch is derived from it (no drift)

APP_DIR   = os.path.dirname(os.path.abspath(__file__))
STATE_DIR = os.path.join(APP_DIR, "agent_state"); os.makedirs(STATE_DIR, exist_ok=True)
FULL_DIR  = os.path.join(STATE_DIR, "full");       os.makedirs(FULL_DIR, exist_ok=True)
HTML_PATH = os.path.join(APP_DIR, "hoodie_mdm.html")

# State store. Local disk by default (./agent_state/). Set STATE_BUCKET (+ optional
# STATE_PREFIX) to persist to S3 instead, so pulled data survives container redeploys.
# The container's local disk is ephemeral; S3 is the durable store. See unifyd/README.md.
STATE_BUCKET = os.environ.get("STATE_BUCKET", "").strip()
STATE_PREFIX = os.environ.get("STATE_PREFIX", "unifyd-state").strip("/")

app = Flask(__name__)
auth_gate.init(app)                # Google OIDC gate (whole origin) — a no-op unless configured

# ---------------- live run jobs (the on-theme console on Run) ----------------
# A run executes in a background thread; the front-end polls /api/run/progress to stream
# the ACTUAL progress. We capture the pulls' existing app.logger.info(...) output per-job
# via a logging handler keyed by the running thread — no scraper changes, real lines only.
JOBS = {}                       # jobId -> {id,connId,status,startedAt,finishedAt,log[],run,error}
_JOB_BY_THREAD = {}             # thread ident -> jobId  (so a log record finds its job)
JOB_LOG_CAP = 800

class _JobLogHandler(logging.Handler):
    def emit(self, record):
        jid = _JOB_BY_THREAD.get(threading.get_ident())
        job = JOBS.get(jid) if jid else None
        if job is None:
            return
        try: msg = record.getMessage()
        except Exception: return
        ln = job["log"]; ln.append({"t": int(time.time() * 1000), "lvl": record.levelname, "msg": msg})
        if len(ln) > JOB_LOG_CAP: del ln[:len(ln) - JOB_LOG_CAP]

_jh = _JobLogHandler(); _jh.setLevel(logging.INFO)
app.logger.addHandler(_jh); app.logger.setLevel(logging.INFO)   # INFO so progress lines flow

# Every source_registry id is a valid connId (the /api/run gate + UI 'runnable' flag accept it) — so a source
# the registry owns is runnable from the app the day it lands, without editing this set. This is what the
# universal registry check in _dispatch_pull needs to be reachable; dispatch_guard_test asserts every enabled
# registry id is here (check 6, COMPLETENESS).
_REGISTRY_IDS = {s["id"] for s in source_registry.SOURCES}
VALID_CONNS = {"ttb-cola", "abc-fws", "specs", "binnys", "shopify", "instacart", "orlando-accounts", "census-acs", "tx-tabc", "il-chicago", "ct-dcp", "total-wine", "vtinfo", "ab-inbev",
               "kroger", "walmart", "walmart-api", "target", "doordash", "google",
               "ubereats", "postmates", "naop", "meijer", "trader-joes",
               # bounded, manual-trigger-only registry entries (source_registry.py, all enabled=False) —
               # reachable via /api/run for the one-off runs they're built for, never the automatic scan.
               # (Also already covered by the `| _REGISTRY_IDS` union below — kept explicit for clarity/history.)
               "ubereats-full", "postmates-full", "doordash-full",
               "instacart-bevalc"} | set(socrata_outlets.VALID) | _REGISTRY_IDS
# Hosts served by an OWNED, dedicated scraper (search-form / bespoke) — not readable by the
# generalized Source Analyzer. If one is analyzed, we point the user to Pulls instead.
OWNED_HOSTS = {"ttbonline.gov": "ttb-cola", "abcfws.com": "abc-fws", "specsonline.com": "specs"}

# SINGLE SOURCE OF TRUTH: sources the registry owns run THROUGH the registry entrypoint (run_one), never a
# hand-maintained *_pull copy — so a fix in source_registry can never be silently bypassed by the app path
# (the drift that made "we fixed it 100 times" literally true — see _REGISTRY_CONN / dispatch_guard_test).
def _dispatch_pull(conn, body):
    # SINGLE SOURCE OF TRUTH: **any** source that exists in source_registry runs THROUGH the registry entrypoint
    # (run_one) — never a hand-maintained *_pull copy. A fix in source_registry therefore applies to the app path
    # automatically and can NEVER be silently bypassed by a stale bespoke copy. This universal rule (not just the
    # _REGISTRY_CONN allowlist) is the real cure for "we fixed it N times and it reverted": the app can no longer
    # run an old copy of a registry source — this is also what makes every newly enabled registry source
    # app-runnable the day it lands, with zero hand-wiring (VALID_CONNS' `| _REGISTRY_IDS` union is the other
    # half). Enforced by dispatch_guard_test.
    import source_registry as _reg
    if conn in _REGISTRY_CONN or _reg.by_id(conn):
        return _run_via_registry(conn, body)
    # NON-registry connectors only (state Socrata feeds, places, census-acs sub-feed, legacy doordash/google):
    if conn in _CONN_PULL:
        return _CONN_PULL[conn](body)
    return (instacart_pull(body) if conn == "instacart"
            else places_pull(body) if conn == "orlando-accounts"
            else tx_pull(body) if conn == "tx-tabc"
            else il_pull(body) if conn == "il-chicago"
            else ct_pull(body) if conn == "ct-dcp"
            else socrata_pull(conn, body) if conn in socrata_outlets.VALID
            else fl_pull(conn) if conn in FL_CONN else None)

def socrata_pull(conn, body):
    """Generic Socrata outlet pull (NY/CO/MO…) → <state>_outlets, normalised to one schema. NY/CO
    ship a Socrata point so they land pre-geocoded; MO-style feeds geocode later like FL/TX."""
    started = int(time.time() * 1000)
    ds, runs, _ = socrata_outlets.pull(conn, log=lambda m: app.logger.info("SOCRATA %s", m))
    DATASETS.update(_absorb(ds)); save()
    run = runs[0]; run["startedAt"] = started; run["finishedAt"] = int(time.time() * 1000)
    run["durationMs"] = run["finishedAt"] - started; run["trigger"] = (body or {}).get("trigger", "manual")
    return run

def tx_pull(body):
    """Texas TABC active retail licenses (Socrata) → tx_outlets + tx_companies, county-keyed for
    the census join. Works from a datacenter (data.texas.gov, unlike the FL/TTB gov hosts)."""
    started = int(time.time() * 1000)
    ds, runs, _ = tx_tabc.pull(status=(body or {}).get("status", "Active"),
                               tier=(body or {}).get("tier", "Retail"),
                               log=lambda m: app.logger.info("TX %s", m))
    DATASETS.update(_absorb(ds)); save()
    run = runs[0]; run["startedAt"] = started; run["finishedAt"] = int(time.time() * 1000)
    run["durationMs"] = run["finishedAt"] - started; run["trigger"] = (body or {}).get("trigger", "manual")
    return run

def il_pull(body):
    """Chicago active liquor-licensed outlets (Socrata) → il_outlets + il_companies. Outlets ship
    with latitude/longitude from the portal (no geocoding) and county='Cook' (census-joinable)."""
    started = int(time.time() * 1000)
    ds, runs, _ = chicago.pull(log=lambda m: app.logger.info("IL %s", m))
    DATASETS.update(_absorb(ds)); save()
    run = runs[0]; run["startedAt"] = started; run["finishedAt"] = int(time.time() * 1000)
    run["durationMs"] = run["finishedAt"] - started; run["trigger"] = (body or {}).get("trigger", "manual")
    return run

def ct_pull(body):
    """Connecticut liquor (Socrata data.ct.gov) → ct_outlets (CT premises, geocodable) +
    ct_brands (product/supplier registry) + ct_companies. Harvest of what CT publishes cleanly."""
    started = int(time.time() * 1000)
    ds, runs, _ = ct_dcp.pull(log=lambda m: app.logger.info("CT %s", m))
    DATASETS.update(_absorb(ds)); save()
    run = runs[0]; run["startedAt"] = started; run["finishedAt"] = int(time.time() * 1000)
    run["durationMs"] = run["finishedAt"] - started; run["trigger"] = (body or {}).get("trigger", "manual")
    return run

def places_pull(body):
    """Run the Orlando on-premise-accounts pull (FL ABT -> normalize -> filter -> Parquet)."""
    return places.pull(county=(body or {}).get("county", places.ORLANDO_COUNTY))

# census-acs now runs through the REGISTRY path (_REGISTRY_CONN -> run_sources.run_one -> census.build()),
# so it PERSISTS to the warehouse (census_demographic/economic/housing) with the registry's uniform
# creds-gating (requires=[CENSUS_API_KEY] -> honest no-creds skip). The old census_pull only refreshed the
# in-memory console preview and never landed — that gap is closed. NOTE: `census` (census_ref, business
# patterns) and `census-acs` (ACS demographics) are DISTINCT sources, not a drift to merge.

# ── Unified connector wrappers (kroger / walmart / target / doordash / google) ─────────────────────────────
# These make five formerly out-of-band scrapers FIRST-CLASS connectors on the same /api/run path as every
# other source: triggerable, schedulable, and visible in run history. Each wrapper runs the owned scraper
# (which lands its warehouse table + its own <x>_runs) and mirrors the result into ONE standard run record.
def _run_id():
    return "R-" + format(int(time.time()) % 100000, "05d")

def _wh_count(name):
    try:
        import warehouse
        return warehouse.query(name, "SELECT count(*) c FROM t")[0]["c"]
    except Exception:
        return 0

def _std_run(conn, started, total=0, extracts=None, status="success", warnings=None, note="", trigger="manual"):
    fin = int(time.time() * 1000)
    return {"id": _run_id(), "connId": conn, "startedAt": started, "finishedAt": fin,
            "durationMs": fin - started, "status": status, "trigger": trigger, "total": total,
            "degraded": status in ("degraded", "partial"), "warnings": warnings or [], "healed": [],
            "extracts": extracts or [], "note": note}


# ── ONE dispatch path: app connId -> source_registry id ──────────────────────────────────────────────────────
# For these, /api/run executes the REGISTRY entrypoint (run_sources.run_one) — the single source of truth the
# scheduled path already uses — instead of a parallel *_pull function that drifts. This is the fix for scrapers
# "regressing": a fix in source_registry now applies to the app path automatically and can't be bypassed.
# ubereats/postmates/naop were previously not runnable through the app at all. Enforced by dispatch_guard_test.
_REGISTRY_CONN = {"abc-fws": "abc-fws", "specs": "specs", "binnys": "binnys", "walmart": "walmart",
                  "kroger": "kroger", "total-wine": "total-wine", "ubereats": "ubereats",
                  "postmates": "postmates", "naop": "naop", "meijer": "meijer", "trader-joes": "trader-joes",
                  "shopify": "shopify", "vtinfo": "vtinfo", "ab-inbev": "ab-inbev",
                  "census-acs": "census-acs"}


def _run_via_registry(conn, body):
    """Run `conn` through its source_registry entrypoint via run_sources.run_one — the exact code the scheduled
    path runs (isolated subprocess, creds-gated via `requires`, timeout-bounded, OOM-reported, warehouse-count
    delta). Adapts run_one's record into the app run-record shape. This is what keeps the app from drifting."""
    import source_registry as _reg
    import run_sources as _rs
    body = body or {}
    rid = _REGISTRY_CONN.get(conn, conn)       # conn IS the registry id for the universal path; allowlist is legacy
    src = _reg.by_id(rid) or next((s for s in _reg.SOURCES if s["id"] == rid), None)
    if not src:
        return _std_run(conn, int(time.time() * 1000), status="failed",
                        warnings=["no source_registry entry '%s'" % rid], trigger=body.get("trigger", "manual"))
    rec = _rs.run_one(src, log=lambda m: app.logger.info("RUN[%s] %s", conn, m))
    started = int(rec.get("ts_start", time.time()) * 1000)
    app_status = {"ok": "success", "no-change": "success"}.get(rec.get("status"), rec.get("status", "failed"))
    tables = [t.strip() for t in (rec.get("tables") or "").split(",") if t.strip()]
    rows_after, delta = rec.get("rows_after", 0), rec.get("delta", 0)
    return _std_run(conn, started, total=rows_after, status=app_status, trigger=body.get("trigger", "manual"),
                    warnings=([rec["error"]] if rec.get("error") else []),
                    extracts=[{"id": t, "rows": rows_after, "delta": delta, "status": app_status} for t in tables])

def doordash_pull(body):
    """DoorDash geo merchant sweep → <market>_merchants (market intelligence). HEAVY (BD Browser). Toggleable."""
    started = int(time.time() * 1000); body = body or {}
    import doordash_geo as dd
    market = body.get("market", "orlando")
    dd.run(market=market, log=lambda m: app.logger.info("DOORDASH %s", m))
    n = _wh_count("%s_merchants" % market)
    return _std_run("doordash", started, total=n, trigger=body.get("trigger", "manual"),
                    extracts=[{"id": "%s_merchants" % market, "rows": n, "delta": 0, "status": "success"}])

def google_pull(body):
    """Google Maps coverage/hours (Bright Data SERP+Maps) → <market>_outlet_hours. HEAVY. Toggleable."""
    started = int(time.time() * 1000); body = body or {}
    import place_coverage as pcov
    market = body.get("market", "orlando"); near = body.get("near", "Orlando FL")
    pcov.enrich_hours(market=market, near=near, log=lambda m: app.logger.info("GOOGLE %s", m))
    n = _wh_count("%s_outlet_hours" % market)
    return _std_run("google", started, total=n, trigger=body.get("trigger", "manual"),
                    extracts=[{"id": "%s_outlet_hours" % market, "rows": n, "delta": 0, "status": "success"}])

def walmart_api_pull(body):
    """Walmart bev-alc CATALOG via the OFFICIAL Walmart I/O Affiliate API (direct, no Bright Data). Needs
    WALMART_CONSUMER_ID + WALMART_PRIVATE_KEY(_FILE) (enroll at walmart.io). Lands walmart_products (accumulates)."""
    started = int(time.time() * 1000); body = body or {}
    import walmart_api
    terms = [t.strip() for t in (body.get("terms") or "").split(",") if t.strip()] or None
    res = walmart_api.run(terms, max_per_term=int(body.get("max_per_term", 1000)),
                          log=lambda m: app.logger.info("WMAPI %s", m))
    if not res:
        return _std_run("walmart-api", started, status="degraded", trigger=body.get("trigger", "manual"),
                        warnings=["Walmart I/O creds missing — set WALMART_CONSUMER_ID + WALMART_PRIVATE_KEY (enroll at walmart.io)"])
    n = _wh_count("walmart_products")
    return _std_run("walmart-api", started, total=n, trigger=body.get("trigger", "manual"),
                    extracts=[{"id": "walmart_products", "rows": n, "delta": 0, "status": "success"}])

# _CONN_PULL holds ONLY non-registry conns. Registry sources (kroger/walmart/target/ttb-cola/…) run via the
# registry path — their old *_pull copies were DELETED, not just unwired. dispatch_guard_test blocks re-adding
# one. instacart/orlando-accounts/census-acs/tx-tabc/il-chicago/ct-dcp are NON-registry app-only pulls dispatched
# directly by the ternary chain in _dispatch_pull (not a table here) — see APP_ONLY_CONN in dispatch_guard_test.py
# for why each is exempt from the registry route.
_CONN_PULL = {"walmart-api": walmart_api_pull,   # NON-registry conns only (target now routes via the registry)
              "doordash": doordash_pull, "google": google_pull}

# The connector registry — ONE source of truth the Pulls console reads (/api/connectors): what sources exist,
# how they run, their warehouse data + <x>_runs table (for last-run), and whether they're enabled (on/off).
CONNECTORS_META = [
    {"id": "ttb-cola", "label": "TTB COLA", "group": "Federal", "runs": None, "data": "ttb_cola", "heavy": True},
    {"id": "abc-fws", "label": "ABC Fine Wine & Spirits (FL)", "group": "Retail chain", "runs": "abc_runs", "data": "abc_products"},
    {"id": "specs", "label": "Spec's (TX)", "group": "Retail chain", "runs": "specs_runs", "data": "specs_products"},
    {"id": "binnys", "label": "Binny's (IL)", "group": "Retail chain", "runs": "binnys_runs", "data": "binnys_products"},
    {"id": "shopify", "label": "Shopify (census sweep)", "group": "Off-premise", "runs": "shopify_runs", "data": "national_shopify_products"},
    {"id": "instacart", "label": "Instacart", "group": "Aggregator", "runs": None, "data": None},
    {"id": "kroger", "label": "Kroger", "group": "Grocery chain", "runs": "kroger_runs", "data": "kroger_products", "needs_creds": True},
    {"id": "walmart", "label": "Walmart (BD sample)", "group": "Grocery chain", "runs": "walmart_runs", "data": "walmart_products", "heavy": True},
    {"id": "walmart-api", "label": "Walmart catalog (Walmart I/O API)", "group": "Grocery chain", "runs": "walmart_runs", "data": "walmart_products", "needs_creds": True},
    {"id": "target", "label": "Target", "group": "Grocery chain", "runs": None, "data": "target_products", "heavy": True},
    {"id": "doordash", "label": "DoorDash (market sweep)", "group": "Aggregator", "runs": None, "data": "orlando_merchants", "heavy": True, "toggle": True},
    {"id": "google", "label": "Google Maps (coverage)", "group": "Reference", "runs": None, "data": "orlando_outlet_hours", "heavy": True, "toggle": True},
    {"id": "orlando-accounts", "label": "Orlando on-premise", "group": "Market", "runs": None, "data": "orlando_accounts"},
    {"id": "census-acs", "label": "US Census — ACS demographics", "group": "Reference", "runs": None, "data": "census_demographic"},
    {"id": "tx-tabc", "label": "Texas TABC", "group": "State", "runs": None, "data": "tx_outlets"},
    {"id": "il-chicago", "label": "Chicago", "group": "State", "runs": None, "data": "il_outlets"},
    {"id": "ct-dcp", "label": "Connecticut", "group": "State", "runs": None, "data": "ct_outlets"},
    {"id": "total-wine", "label": "Total Wine", "group": "Retail chain", "runs": None, "data": "total_wine_products"},
    {"id": "vtinfo", "label": "VTInfo locator", "group": "Reference", "runs": None, "data": "vtinfo_titos"},
    {"id": "ab-inbev", "label": "AB InBev locator", "group": "Reference", "runs": None, "data": None},
]

def _conn_enabled():
    return (load("connectors.json", {}) or {}).get("enabled", {})

def _conn_is_enabled(cid):
    return _conn_enabled().get(cid, True)          # default ON; only an explicit false disables

def _new_job(conn):
    jid = "J-%d-%d" % (int(time.time()), len(JOBS) + 1)
    JOBS[jid] = {"id": jid, "connId": conn, "status": "running",
                 "startedAt": int(time.time() * 1000), "finishedAt": None,
                 "log": [], "run": None, "error": None}
    if len(JOBS) > 40:          # keep the last ~40 jobs
        for k in sorted(JOBS, key=lambda k: JOBS[k]["startedAt"])[:len(JOBS) - 40]:
            JOBS.pop(k, None)
    return jid

def _emit_pull_highlights(conn, rec):
    """An owned pull emits a highlight per landed extract — so the estate model pulses that dataset
    node (it matches on `dataset`) and the highlights feed shows owned pulls, not just recipe scrapes."""
    try:
        ts = int(time.time() * 1000)
        for e in (rec.get("extracts") or []):
            n = e.get("rows") or 0
            if n <= 0 or e.get("status") == "failed":
                continue
            HIGHLIGHTS.insert(0, {"host": conn, "dataset": e.get("id"), "engine": "owned pull",
                                  "count": n, "ts": ts,
                                  "headline": "%s · %s rows" % (_SRC_LABEL.get(conn, conn), format(n, ","))})
        del HIGHLIGHTS[30:]
        _save_json("highlights.json", HIGHLIGHTS)
    except Exception as ex:
        app.logger.warning("highlight emit failed for %s: %s", conn, ex)

def _run_job(jid, conn, body):
    _JOB_BY_THREAD[threading.get_ident()] = jid
    job = JOBS[jid]
    try:
        app.logger.info("%s: run started", conn)
        rec = _dispatch_pull(conn, body)
        if rec is None:
            job["status"] = "error"; job["error"] = "unknown connId: %s" % conn
            app.logger.warning("%s: unknown connId", conn); return
        RUNS.insert(0, rec); del RUNS[200:]; save()
        _emit_pull_highlights(conn, rec)   # so the estate model pulses + the feed shows owned pulls
        job["run"] = rec; job["status"] = rec.get("status", "success")
        app.logger.info("%s: done — %s rows, status %s", conn, rec.get("total"), rec.get("status"))
    except Exception as e:
        app.logger.exception("run failed")
        job["status"] = "error"; job["error"] = str(e)
    finally:
        job["finishedAt"] = int(time.time() * 1000)
        _JOB_BY_THREAD.pop(threading.get_ident(), None)

# ---------------- hardening: optional auth + JSON errors ----------------
# AGENT_TOKEN gates /api/* for non-browser callers (off by default — local dev and
# browser apps behind the CloudFront password function are unaffected). /api/health
# stays open so load-balancer / uptime probes work.
AGENT_TOKEN = os.environ.get("AGENT_TOKEN", "").strip()

@app.before_request
def _auth():
    if not AGENT_TOKEN:
        return
    p = request.path
    # NOTE: there are TWO gates. auth_gate._PUBLIC is not enough on its own — this AGENT_TOKEN
    # check runs as well, and exempting a route from one but not the other still 401s it. That is
    # exactly how /api/version shipped unreachable: _PUBLIC listed it, this did not.
    if p in ("/api/health", "/api/version") or not p.startswith("/api/"):
        return
    # A signed-in browser is already authorized by auth_gate (which runs first and only
    # falls through to here when it ALLOWED the request). Without this, setting
    # AGENT_TOKEN 401s every /api/* call from a logged-in browser — and session-guard.js
    # reads that 401 as an expired session and bounces the user back to Google.
    if session.get("email"):
        return
    if request.headers.get("Authorization", "") == "Bearer " + AGENT_TOKEN \
       or request.headers.get("X-Agent-Token", "") == AGENT_TOKEN:
        return
    return jsonify(ok=False, error="unauthorized"), 401

# Opt-in CORS for cross-origin web clients (the production web app). Off by default;
# set API_CORS to an allowed origin (e.g. http://localhost:8082) or "*" to enable.
API_CORS = os.environ.get("API_CORS", "").strip()

@app.after_request
def _cors(resp):
    if API_CORS and request.path.startswith("/api/"):
        resp.headers["Access-Control-Allow-Origin"] = API_CORS
        if API_CORS != "*":
            resp.headers["Access-Control-Allow-Credentials"] = "true"
        resp.headers["Access-Control-Allow-Headers"] = "Authorization, Content-Type, Accept"
        resp.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    # Anti-scraping / no-index defense-in-depth on top of the OIDC gate. The suite iframes
    # its own apps (launcher, sources.html, mdm.html) so framing is SAMEORIGIN, not DENY.
    resp.headers.setdefault("X-Robots-Tag", "noindex, nofollow, noarchive")
    resp.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
    resp.headers.setdefault("X-Content-Type-Options", "nosniff")
    resp.headers.setdefault("Referrer-Policy", "no-referrer")
    # Always revalidate the CODE (html/js/css/json) so a deploy is picked up immediately —
    # browser heuristic caching was serving a stale launcher/app after a redeploy (a mix of
    # old shell + new app). ETag/Last-Modified still make it a cheap 304 when unchanged.
    ct = (resp.content_type or "")
    if ct.startswith("text/html") or "javascript" in ct or ct.startswith("text/css") or ct.startswith("application/json"):
        resp.headers["Cache-Control"] = "no-cache"
    return resp


@app.after_request
def _gzip(resp):
    """Gzip large text/JSON responses (stdlib, no dep). The coverage map's /api/outlets/geo ships
    100k+ light points — very repetitive JSON that compresses ~5:1 — so the whole national footprint
    transfers in a few MB instead of tens. Guarded so it never touches streamed/binary/already-encoded
    responses."""
    try:
        if resp.direct_passthrough or resp.status_code != 200:
            return resp
        if "gzip" not in request.headers.get("Accept-Encoding", "") or resp.headers.get("Content-Encoding"):
            return resp
        ct = resp.content_type or ""
        if not (ct.startswith("application/json") or ct.startswith("text/") or "javascript" in ct):
            return resp
        data = resp.get_data()
        if len(data) < 1024:
            return resp
        resp.set_data(gzip.compress(data, 5))
        resp.headers["Content-Encoding"] = "gzip"
        resp.headers["Content-Length"] = str(len(resp.get_data()))
        resp.headers["Vary"] = "Accept-Encoding"
    except Exception:
        pass
    return resp


# Light per-IP rate limit — the OIDC gate already restricts to allowlisted humans; this
# just stops an authenticated session (or the login endpoint) from being hammered to pull
# the whole book/catalog fast. In-memory sliding window (single Fly machine). Generous for
# a human, restrictive for a scraper. Tune with RATE_MAX / RATE_WINDOW; 0 disables.
_RATE = {}
RATE_MAX    = int(os.environ.get("RATE_MAX", "600"))     # requests per window per IP on /api
RATE_WINDOW = int(os.environ.get("RATE_WINDOW", "60"))   # seconds

def _client_ip():
    return (request.headers.get("Fly-Client-IP")
            or (request.headers.get("X-Forwarded-For", "").split(",")[0].strip())
            or request.remote_addr or "?")

@app.before_request
def _ratelimit():
    if RATE_MAX <= 0 or not request.path.startswith("/api/") or request.path == "/api/health":
        return
    now = time.time(); ip = _client_ip()
    hits = [t for t in _RATE.get(ip, []) if now - t < RATE_WINDOW]
    hits.append(now); _RATE[ip] = hits
    if len(_RATE) > 5000:                                   # cap memory: drop the coldest IPs
        for k in [k for k, v in list(_RATE.items()) if not v or now - v[-1] > RATE_WINDOW]:
            _RATE.pop(k, None)
    if len(hits) > RATE_MAX:
        return jsonify(ok=False, error="rate limited — slow down"), 429


@app.get("/robots.txt")
def robots():
    # Belt-and-suspenders: the OIDC gate already blocks anonymous crawlers; this tells the
    # polite ones to stay out entirely and keeps the site out of search indexes.
    return Response("User-agent: *\nDisallow: /\n", mimetype="text/plain")

@app.errorhandler(404)
def _e404(e):
    # A browser NAVIGATION that misses (e.g. a mistyped /prism.html, or a post-login
    # redirect to a stale path) should land on the launcher, not a raw JSON blob.
    # API calls and asset requests (image/*, etc.) still get JSON.
    if not request.path.startswith("/api/") and "text/html" in request.headers.get("Accept", ""):
        return redirect("/")
    return jsonify(ok=False, error="not found"), 404
@app.errorhandler(405)
def _e405(e): return jsonify(ok=False, error="method not allowed"), 405
@app.errorhandler(Exception)
def _e500(e):
    from werkzeug.exceptions import HTTPException
    if isinstance(e, HTTPException):
        return jsonify(ok=False, error=e.description), e.code
    app.logger.exception("unhandled error")
    return jsonify(ok=False, error="internal error"), 500

# ---------------- persisted state (disk | S3) ----------------
def _s3():
    import boto3                      # lazy — only imported when STATE_BUCKET is set
    return boto3.client("s3")
def _key(name):
    return (STATE_PREFIX + "/" + name) if STATE_PREFIX else name

def load(name, default):
    if STATE_BUCKET:
        try:
            obj = _s3().get_object(Bucket=STATE_BUCKET, Key=_key(name))
            return json.loads(obj["Body"].read())
        except Exception:
            return default            # missing object / first boot -> default
    try: return json.load(open(os.path.join(STATE_DIR, name)))
    except Exception: return default

DATASETS = load("datasets.json", {})
RUNS     = load("runs.json", [])
# Service desk — data-quality reports raised from the suite (e.g. clicking a "degraded"
# pill in Hoodie Pulls). Logged here, surfaced in the CRM's Service tab, tracked to
# resolution. Same disk-or-S3 persistence as the rest of the agent state.
SERVICE  = load("service_reports.json", [])
# Self-reinforcing scrape recipes — host -> proven config + validation baseline.
# Each analyze/extract folds into this book (see recipes.py); persisted like the rest.
RECIPES  = load("recipes.json", {})
# Scrape highlights — a rolling feed of 'what the last pulls yielded' (counts, ranges,
# geo, top categories), surfaced as a card on completion and in Hoodie Intelligence.
HIGHLIGHTS = load("highlights.json", [])
# Per-account planograms — shelf facings for an account (entered by hand or derived from
# a photo in the Planogram tool). Keyed by account id; persisted so the numbers stick.
PLANOGRAM = load("planograms.json", {})
# Admin console: usage events (app opens, per user) + feature flags (app visibility /
# status overrides, feature toggles). Both persisted with the rest of the agent state.
USAGE = load("usage.json", [])
FLAGS = load("admin_flags.json", {"apps": {}, "features": {}})
# UPC prefix->owner crosswalk (built by upc_crosswalk.py from enriched COLA; empty until uploaded).
# Deterministic UPC checks work with or without it; owner-agreement activates once it's present.
UPC_XWALK = load("upc_crosswalk.json", {})
# Hoodie Relations extras — per-account delivery day + goals (goals matched to the
# master via relations.match_goal). Keyed by account id; persisted so it's reportable.
RELATIONS = load("relations.json", {})

# Canonical hierarchy served at /api/hierarchy. Curated copy from the state store if
# present, else the bundled seed (unifyd/hierarchy.json). Eventually derived from the
# owned master data; for now it mirrors spine/hierarchy.sample.json.
HIER_SEED = os.path.join(APP_DIR, "hierarchy.json")
def load_hierarchy():
    h = load("hierarchy.json", None)
    if h is not None: return h
    try: return json.load(open(HIER_SEED))
    except Exception: return {"id": "root", "level": "root", "name": "All Portfolios", "children": []}
HIERARCHY = load_hierarchy()

def save():
    blobs = {"datasets.json": DATASETS, "runs.json": RUNS}
    if STATE_BUCKET:
        c = _s3()
        for name, data in blobs.items():
            try:
                c.put_object(Bucket=STATE_BUCKET, Key=_key(name),
                             Body=json.dumps(data).encode("utf-8"),
                             ContentType="application/json")
            except Exception as e:
                app.logger.warning("S3 save %s failed: %s", name, e)
    else:
        for name, data in blobs.items():
            json.dump(data, open(os.path.join(STATE_DIR, name), "w"))

# Full pulled rows live OUTSIDE the in-memory DATASETS sample (kept small for the UI):
# written per-dataset at pull time, read on demand by /api/datasets/download. Same
# disk-or-S3 abstraction as save()/load(), so the complete pull is always extractable.
def save_full(did, header, rows):
    blob = json.dumps({"header": header, "rows": rows, "total": len(rows)})
    if STATE_BUCKET:
        try: _s3().put_object(Bucket=STATE_BUCKET, Key=_key("full/" + did + ".json"),
                              Body=blob.encode("utf-8"), ContentType="application/json")
        except Exception as e: app.logger.warning("S3 full save %s failed: %s", did, e)
    else:
        try: open(os.path.join(FULL_DIR, did + ".json"), "w").write(blob)
        except Exception as e: app.logger.warning("full save %s failed: %s", did, e)
    # keep a tiny index (did → header+total) so /api/catalog can list the full store without reading
    # every (possibly 30MB) blob — this is what lets the estate model see save_full'd datasets.
    try:
        idx = _full_index(); idx[did] = {"header": header, "total": len(rows)}
        blob = json.dumps(idx)
        if STATE_BUCKET:
            _s3().put_object(Bucket=STATE_BUCKET, Key=_key("full/_index.json"),
                             Body=blob.encode("utf-8"), ContentType="application/json")
        else:
            open(os.path.join(FULL_DIR, "_index.json"), "w").write(blob)
    except Exception as e:
        app.logger.warning("full index update %s failed: %s", did, e)

def _full_index():
    """did → {header, total} for everything in the full store (from full/_index.json)."""
    if STATE_BUCKET:
        try:
            obj = _s3().get_object(Bucket=STATE_BUCKET, Key=_key("full/_index.json"))
            return json.loads(obj["Body"].read())
        except Exception:
            return {}
    try:
        return json.load(open(os.path.join(FULL_DIR, "_index.json")))
    except Exception:
        return {}

def load_full(did):
    if STATE_BUCKET:
        try:
            obj = _s3().get_object(Bucket=STATE_BUCKET, Key=_key("full/" + did + ".json"))
            return json.loads(obj["Body"].read())
        except Exception: return None
    try: return json.load(open(os.path.join(FULL_DIR, did + ".json")))
    except Exception: return None

def _to_warehouse(ds):
    """Write a pull's full rows to the warehouse as Parquet so the apply engine, the data dictionary,
    and all DuckDB queries can reach them (memory DATASETS alone can't be read by read_parquet, and
    won't scale to the national AB census). Call BEFORE _absorb (which pops _rows_full). Best-effort."""
    for name, d in (ds or {}).items():
        header = d.get("header") or []
        rows = d.get("_rows_full") or d.get("rows") or []
        if not (header and rows):
            continue
        try:
            recs = [dict(zip(header, r)) for r in rows]
            warehouse.write_parquet(name, recs, header)
            app.logger.info("warehouse ← %s: %d rows", name, len(recs))
        except Exception as e:
            app.logger.warning("warehouse write %s failed: %s", name, e)

def _absorb(ds):
    """Lift the full rows a scraper attached as `_rows_full` out to the on-demand full store
    (for complete CSV/JSON export), then leave the in-memory dataset capped at its UI sample
    so DATASETS / agent_state stay small. No `_rows_full` (e.g. FL/COLA) → unchanged no-op."""
    for did, d in (ds or {}).items():
        full = d.pop("_rows_full", None)
        if full is not None:
            save_full(did, d.get("header") or [], full)
    return ds

# ---------------- Florida pull (live, no extra deps) ----------------
FL_BASE = "https://www2.myfloridalicense.com/sto/file_download/extracts"
FL_HEADER = ["Board","Profession","Owner Name","Series","Modifier","Mail Address 1","Mail Address 2",
    "Mail Address 3","Mail City","Mail State","Mail ZIP","Mail County","DBA","Location Address 1",
    "Location Address 2","Location Address 3","Location City","Location State","Location ZIP",
    "Location County","License Number","Primary Status","Secondary Status","Original Licensure Date",
    "Effective Date","Expiration Date","Tax Stamp Designation","Smoking Designation","Retail Tobacco Indicator"]
FL_CONN = {
  "fl-items":   [("bd4008lic", True, 600), ("bd4011lic", True, 500), ("abtbrands", True, 600)],
  "fl-outlets": [("bd4006lic", True, 600), ("bd4005lic", True, 500),
                 ("bd400revok", False, 500), ("bd4002lic", True, 500)],
}
def _fl_looks_blocked(t):
    head = (t or "")[:800].lower()
    return ("just a moment" in head) or ("cloudflare" in head) or ("<html" in head and "cf-" in head)

def _fl_fetch(url):
    """Download an FL DBPR extract. FL's host (myfloridalicense.com) is now behind Cloudflare, so:
    try direct first; if it 403s / returns a 'Just a moment' challenge, fall back to Bright Data
    Web Unlocker (which solves Cloudflare — needs BRIGHTDATA_API_KEY). Returns (csv_text, via)."""
    try:
        req = urllib.request.Request(url, headers={
            "User-Agent": "Mozilla/5.0 (compatible; UnifydMDM/1.0; +data-pipeline)"})
        txt = urllib.request.urlopen(req, timeout=180).read().decode("utf-8", "replace")
        if txt and not _fl_looks_blocked(txt):
            return txt, "direct"
    except Exception:
        pass                                     # 403 / cert / timeout → try the unlocker
    try:
        import brightdata
    except Exception:
        brightdata = None
    if brightdata and brightdata.enabled():
        txt = brightdata.fetch(url, data_format="html", timeout=180)
        if txt and not _fl_looks_blocked(txt):
            return txt, "bright-data"
        raise RuntimeError("Bright Data returned a blocked/empty response for FL")
    raise RuntimeError("Cloudflare-blocked — set BRIGHTDATA_API_KEY to unlock FL DBPR (Web Unlocker)")

def fl_pull(conn_id):
    started = int(time.time() * 1000); exs = []; warns = []
    extracts = FL_CONN[conn_id]
    for i, (eid, hashdr, n) in enumerate(extracts):
        try:
            app.logger.info("downloading %s (%d/%d) from Florida…", eid, i + 1, len(extracts))
            txt, via = _fl_fetch(f"{FL_BASE}/{eid}.csv")
            app.logger.info("%s: %s KB downloaded via %s, parsing…", eid, len(txt) // 1024, via)
            rows = list(csv.reader(io.StringIO(txt)))
            header = [h.strip() for h in rows[0]] if hashdr else FL_HEADER
            data = [r for r in (rows[1:] if hashdr else rows) if any((c or "").strip() for c in r)]
            DATASETS[eid] = {"header": header, "rows": cola.sample(data, header, n),
                             "total": len(data), "profile": cola.profile(header, data)}
            save_full(eid, header, data)   # keep ALL rows extractable, not just the sample
            prev = next((e for r in RUNS if r["connId"] == conn_id for e in r["extracts"] if e["id"] == eid), None)
            delta = len(data) - (prev["rows"] if prev else len(data))
            app.logger.info("%s: %s rows (%s records) ✓", eid, len(data), len(data))
            exs.append({"id": eid, "rows": len(data), "delta": delta, "status": "success"})
        except Exception as e:
            app.logger.warning("FL %s failed: %s", eid, e)
            msg = str(e)
            if "CERTIFICATE_VERIFY_FAILED" in msg or "local issuer" in msg:
                msg = "TLS certificate verify failed for the FL data host (CA chain incomplete on this host)"
            elif "403" in msg or "Forbidden" in msg:
                msg = "Cloudflare-blocked — set BRIGHTDATA_API_KEY to unlock FL DBPR via Web Unlocker"
            warns.append("%s: %s" % (eid, msg[:140]))
            exs.append({"id": eid, "rows": 0, "delta": 0, "status": "failed"})
    fin = int(time.time() * 1000)
    status = "failed" if all(e["status"] == "failed" for e in exs) else \
             "partial" if any(e["status"] == "failed" for e in exs) else "success"
    return {"id": "R-" + format(int(time.time()) % 100000, "05d"), "connId": conn_id,
            "startedAt": started, "finishedAt": fin, "durationMs": fin - started,
            "status": status, "trigger": "manual", "total": sum(e["rows"] for e in exs),
            "degraded": status == "partial", "warnings": warns, "extracts": exs}

# NOTE: cola_pull / abc_pull / specs_pull / binnys_pull were removed — ttb-cola/abc-fws/specs/binnys all run
# through the registry (run_one) now; those old DATASETS-preview copies were the drift that caused reversions.
def instacart_pull(params):
    """FREE Instacart pull — self-hosted Playwright browser drives Instacart's own SearchResultsPlacements
    GraphQL (no Bright Data, no proxy). Lands per-store product+price into instacart_products +
    retail_observations via the aggregator base. Instacart is a browser source: this needs Chromium in the
    image (it ships one) — if the browser can't launch it reports an HONEST failed run, never a 0-row pass."""
    started = int(time.time() * 1000)
    zip_ = str(params.get("zip") or params.get("address") or "10001")
    queries = params.get("queries") or ["milk", "vodka", "chips", "coffee", "eggs", "water",
                                        "soda", "bread", "cheese", "beer"]
    if isinstance(queries, str):
        queries = [q.strip() for q in queries.split(",") if q.strip()]
    pages = int(params.get("pages", 2))
    rows, warnings, status = [], [], "success"
    try:
        from instacart import Instacart
        rows = Instacart().pull(address=zip_, retailers=["grocery"], queries=queries,
                                per_query_pages=pages, log=lambda m: app.logger.info("INSTACART %s", m))
        if not rows:
            status = "degraded"
            warnings.append("Free browser reached Instacart but landed 0 products (membership wall / zone "
                            "capture / anti-bot at this IP). It's a browser source — needs Chromium here or "
                            "the residential executor.")
    except ImportError as e:
        status = "failed"
        warnings.append("Instacart needs Playwright + Chromium (free, no bd) — this image should ship it: %s"
                        % str(e)[:160])
    except Exception as e:
        status = "failed"
        warnings.append("Instacart free pull failed: %s" % str(e)[:200])
    if rows:                                          # keep the console preview populated (real data is in the warehouse)
        header = ["Store", "Product", "Price", "In Stock"]
        drows = [[r.get("store"), r.get("name"), r.get("price"), r.get("in_stock")] for r in rows]
        DATASETS.update(_absorb({"instacart_products": {"header": header, "rows": drows[:800],
                                 "total": len(rows), "stores": len({r.get("store_id") for r in rows})}}))
        save()
    finished = int(time.time() * 1000)
    return {"id": "R-IC%05d" % (started % 100000), "connId": "instacart", "startedAt": started,
            "finishedAt": finished, "durationMs": finished - started, "status": status,
            "total": len(rows), "degraded": status == "degraded", "warnings": warnings, "healed": [],
            "extracts": [{"id": "instacart_products", "rows": len(rows), "delta": len(rows),
                          "status": status}], "trigger": params.get("trigger", "manual")}

# EVERY registry-owned source runs through the REGISTRY path (server._dispatch_pull -> run_sources.run_one),
# so the app runs EXACTLY what the scheduler runs. All the old hand-wired *_pull copies (kroger, walmart, target,
# cola/ttb-cola, abc, specs, binnys, total_wine, vtinfo, ab-inbev) were DELETED — they were the drift that made
# "we fixed it N times and it reverted" literally true (e.g. ab-inbev ran ab_locator.pull() in the app but
# ab_fill.run() in the registry). Re-adding a *_pull for a registry-owned source is blocked by dispatch_guard_test.

# ---------------- API ----------------
@app.get("/api/version")
def version():
    """What code is this container actually running?

    Deliberately public and deliberately boring: a content fingerprint and a file count, nothing
    identifying. It exists because `flyctl releases` reports that a deploy SUCCEEDED, not what it
    deployed — and twice in one day production silently stopped matching main while every release
    read `complete`. deploy_drift.py compares this against the build we last deployed on purpose.
    """
    import deploy_drift
    fp = deploy_drift.fingerprint()
    return jsonify(fingerprint=fp["sha256"], files=fp["files"], computed_at=int(time.time()))

@app.get("/api/health")
def health():
    return jsonify(ok=True, agent="unifyd-local", sources=list(FL_CONN) + ["ttb-cola", "abc-fws", "specs", "binnys", "shopify", "instacart", "census-acs"],
                   datasets=len(DATASETS), runs=len(RUNS),
                   state=("s3:" + STATE_BUCKET) if STATE_BUCKET else "disk",
                   warehouse=("tigris:" + os.environ.get("BUCKET_NAME", "")) if warehouse.remote() else "local")

# ---- Product Locator: brand → accounts carrying it near a location (Grey-Goose-style UI) ----
def _titlecase(s):
    return " ".join(w if (w.isupper() and len(w) <= 3) else w.capitalize() for w in str(s).split())

@app.get("/api/locator/brands")
def locator_brands():
    return jsonify(brands=[{"id": k, "name": v.get("name", k)} for k, v in vtinfo.BRANDS.items()])

@app.get("/api/locator/offers")
def locator_offers():
    """The "why go here" layer: priced, ranked offers for an item near a point.

    `/api/locator` answers WHO CARRIES IT from the distributor feed. This answers WHY GO HERE
    from our own observations — verified stock, the everyday-vs-promo price, a Google-Flights
    style "is this a good price" verdict against the local trailing distribution, and a
    wait-or-buy read off the store's promo cadence. All of it in unifyd/locator_signal.py.

    Params: q (brand / product / UPC), lat+lng or city+state, radius (mi), mode, size.
    Results are scoped to ONE bottle format (`size`, default = deepest pool) because a
    within-format percentile can't be ranked against another format's — `formats` lists the
    alternatives so the UI can tab between them.
    `mode=brand` strips every claim a brand-embedded widget may not make (see
    locator_signal.for_render_mode) — the filtering is server-side on purpose.
    """
    import locator_signal
    q = (request.args.get("q") or "").strip()
    if not q:
        return jsonify(offers=[], error="missing q"), 200
    center = None
    try:
        if request.args.get("lat") and request.args.get("lng"):
            center = (float(request.args["lat"]), float(request.args["lng"]))
        elif request.args.get("city"):
            import city_centroid
            c = city_centroid.centroid(request.args["city"], request.args.get("state", ""))
            if c:
                center = (c[0], c[1]) if isinstance(c, (list, tuple)) else (c["lat"], c["lng"])
    except Exception as e:                     # noqa: BLE001 — a bad center is not a 500
        app.logger.info("LOCATOR offers center error %s", e)
    try:
        radius = float(request.args.get("radius", 15))
    except (TypeError, ValueError):
        radius = 15.0
    mode = "brand" if request.args.get("mode") == "brand" else "consumer"
    try:
        out = locator_signal.offers(q, center=center, radius_mi=radius, mode=mode,
                                    size=request.args.get("size"),
                                    variant=request.args.get("variant"),
                                    log=lambda m: app.logger.info("LOCATOR %s", m))
    except Exception as e:                     # noqa: BLE001
        app.logger.info("LOCATOR offers error %s", e)
        return jsonify(offers=[], error=str(e)[:160], degraded="offers-failed"), 200
    out["center"] = list(center) if center else None
    out["radius_mi"] = radius
    return jsonify(**out)

@app.get("/api/locator/accounts")
def locator_accounts():
    """Accounts near a point, from OUR outlet spine (`src_outlets`, ~1.76M rows).

    The locator's account list used to be a live passthrough to VIP's brand locator — a third
    party's ZIP-keyed page walk capped at 150 rows — while the national outlet universe we acquire
    sat unused. This is that universe.

    Params: lat+lng or `where` (ZIP / "City, ST"), radius (mi), category
    (spirits|beer|wine|hemp|cannabis|rtd_spirits), chains (1/0), q, limit.

    `total` is how many accounts match in the warehouse; `shown` is how many pins were shipped.
    They are reported separately on purpose — conflating them is how "150" came to look like an
    answer. An account row says WHERE a store is and what categories it carries, never that it
    stocks a specific bottle; `/api/locator/offers` is the product-level layer.
    """
    import outlet_locator
    center = None
    resolved = None
    try:
        if request.args.get("lat") and request.args.get("lng"):
            center = (float(request.args["lat"]), float(request.args["lng"]))
    except (TypeError, ValueError):
        center = None
    if center is None and request.args.get("where"):
        import zcta
        loc = zcta.resolve(request.args["where"])
        if loc.get("error"):
            return jsonify(accounts=[], total=0, shown=0, error=loc["error"],
                           degraded="location-unresolved"), 200
        # A city resolves straight to a centroid; a typed ZIP only gives us the ZIP string, so look
        # up that ZCTA's own centre. If neither yields a point we say so — we never fall back to a
        # default location, which is the bug that made a Houston search render Tampa.
        if loc.get("lat") is not None:
            center, resolved = (loc["lat"], loc["lng"]), loc
        else:
            z = zcta.center_of(loc.get("zip"))
            if z:
                center, resolved = z, loc
            else:
                return jsonify(accounts=[], total=0, shown=0,
                               error="couldn't place ZIP %s — ZIP reference may not be built"
                                     % loc.get("zip"),
                               degraded="location-unresolved"), 200
    if center is None:
        return jsonify(accounts=[], total=0, shown=0, error="need lat+lng or where",
                       degraded="location-unresolved"), 200
    try:
        radius = float(request.args.get("radius", 15))
    except (TypeError, ValueError):
        radius = 15.0
    chains = request.args.get("chains")
    try:
        limit = int(request.args.get("limit", outlet_locator.MAX_POINTS))
    except (TypeError, ValueError):
        limit = outlet_locator.MAX_POINTS
    out = outlet_locator.query(center, radius_mi=radius,
                               category=(request.args.get("category") or "").strip() or None,
                               chains_only=(None if chains in (None, "") else chains == "1"),
                               q=(request.args.get("q") or "").strip() or None, limit=limit,
                               log=lambda m: app.logger.info("ACCOUNTS %s", m))
    out["center"] = list(center)
    out["radius_mi"] = radius
    if resolved:
        out["resolved"] = {"label": resolved.get("label"), "how": resolved.get("how")}
    return jsonify(**out)

@app.get("/api/locator/serves")
def locator_serves():
    """The ON-PREMISE half: where can I DRINK this near here — and which one should I go to?

    `/api/locator/offers` finds the bottle; this finds the pour. Backed by `menu_beverages`
    (cocktail_taxonomy's root/sub/base_spirit + the BUILD in `description`) joined to the outlet
    spine by name. See unifyd/serve_locator.py — a venue we cannot pin to one branch is returned
    as a COUNT, never as a pin, because sending someone to the wrong Beef O'Brady's is worse than
    admitting we don't know which one.

    Params: q (drink / family / base spirit), lat+lng, radius (mi), state, recommend=1.

    `recommend=1` adds a Claude pass (unifyd/serve_advisor.py) that picks the most interesting
    builds and says how each departs from the classic — grounded ONLY in the menu description and
    price, and re-verified against these same rows before it is returned. It carries no review or
    popularity claim because we hold no review text; when it is unavailable the reason is stated
    rather than the field quietly going missing.
    """
    import serve_locator
    q = (request.args.get("q") or "").strip()
    if not q:
        return jsonify(serves=[], error="missing q"), 200
    center = None
    try:
        if request.args.get("lat") and request.args.get("lng"):
            center = (float(request.args["lat"]), float(request.args["lng"]))
    except (TypeError, ValueError) as e:
        app.logger.info("SERVES center error %s", e)
    try:
        radius = float(request.args.get("radius", 25))
    except (TypeError, ValueError):
        radius = 25.0
    try:
        out = serve_locator.serves(q, center=center, radius_mi=radius,
                                   state=(request.args.get("state") or "").strip().upper() or None,
                                   log=lambda m: app.logger.info("SERVES %s", m))
    except Exception as e:                     # noqa: BLE001
        app.logger.info("SERVES error %s", e)
        return jsonify(serves=[], error=str(e)[:160], degraded="serves-failed"), 200
    if request.args.get("recommend") == "1":
        import serve_advisor
        out["recommendations"] = serve_advisor.recommend(
            q, out["serves"], log=lambda m: app.logger.info("ADVISOR %s", m))
    out["center"] = list(center) if center else None
    out["radius_mi"] = radius
    return jsonify(**out)

@app.get("/api/locator")
def locator():
    """Live account list for a brand near a ZIP, shaped for the locator UI. Bounded (max_pages)
    so it stays responsive; the map centers on the mean of returned coordinates."""
    brand = request.args.get("brand", "titos")
    b = vtinfo.BRANDS.get(brand)
    if not b:
        return jsonify(accounts=[], error="unknown brand"), 200
    # Resolve the typed location HONESTLY. This previously read
    #     zc = re.sub(r"[^0-9]", "", request.args.get("zip","")) or "33602"
    # so "Houston, TX" — which has no digits — silently became 33602, i.e. TAMPA. The user searched
    # one city and got a confident map of another, with nothing on screen saying so. There is now no
    # default at all: unresolvable input is an error the UI shows, never a substituted city.
    import zcta
    loc = zcta.resolve(request.args.get("zip", ""), log=lambda m: app.logger.info("LOCATOR %s", m))
    if loc.get("error"):
        return jsonify(accounts=[], error=loc["error"], degraded="location-unresolved"), 200
    zc = loc["zip"]
    # `pages` bounded the VIP page walk at 3 (=150 accounts) while the UI rendered the result as
    # "150 accounts near you" — a truncation presented as a census. Walk the whole result set by
    # default and report `truncated` when a cap is actually applied, per the no-silent-caps rule.
    try:
        pages = max(1, int(request.args.get("pages", 40)))
    except (TypeError, ValueError):
        pages = 40
    walk = {}
    try:
        rows = vtinfo.search_zip(b["custID"], b.get("uuid", ""), zc, delay=0.15,
                                 max_pages=pages, stats=walk,
                                 m=b.get("m", "5"), theme=b.get("theme", "1"),
                                 log=lambda m: app.logger.info("LOCATOR %s", m))
    except Exception as e:
        app.logger.info("LOCATOR error %s", e)
        return jsonify(accounts=[], error=str(e)[:120]), 200
    accts, lats, lngs = [], [], []
    for r in rows:
        if not (r.get("lat") and r.get("lng")):
            continue
        lat, lng = float(r["lat"]), float(r["lng"])
        lats.append(lat); lngs.append(lng)
        accts.append({"name": _titlecase(r["account"]), "street": _titlecase(r["street"]),
                      "city": _titlecase(r["city"]), "state": r["state"], "lat": lat, "lng": lng,
                      "source": r.get("source", ""), "type": r.get("store_type", "")})
    center = [sum(lats) / len(lats), sum(lngs) / len(lngs)] if lats else None
    # `resolved` tells the UI WHICH place it actually searched and how it got there, so a ZIP derived
    # from a typed city is visible rather than implied. `truncated` is true only when the page cap
    # actually bit — an unbounded walk that finished must not look like a capped one.
    return jsonify(label=b.get("name", brand), accounts=accts, center=center, zip=zc,
                   resolved={"label": loc.get("label", zc), "how": loc.get("how"),
                             "zip_distance_mi": loc.get("zip_distance_mi")},
                   truncated=bool(walk.get("truncated")), pages=walk.get("walked_pages", pages),
                   available_pages=walk.get("available_pages"))

# Source labels + a first-cut scope tree derived from the pulled data.
_SRC_LABEL = {"fl-items": "Florida — Items", "fl-outlets": "Florida — Outlets",
              "ttb-cola": "TTB — COLA Labels", "abc-fws": "ABC FWS — Inventory",
              "specs": "Spec's — Inventory", "binnys": "Binny's — Inventory",
              "shopify": "Shopify — census sweep", "instacart": "Instacart — Store-level",
              "census-acs": "US Census — ACS demographics", "tx-tabc": "Texas TABC — licenses",
              "il-chicago": "Chicago — Liquor Licenses", "ct-dcp": "Connecticut — Liquor (DCP)",
              "ny-sla": "New York — SLA licenses", "co-led": "Colorado — Liquor licenses",
              "mo-atc": "Missouri — Alcohol licenses",
              "total-wine": "Total Wine — Catalog (direct)", "vtinfo": "VTInfo — Brand distribution (VIP)",
              "ab-inbev": "AB InBev — Retailer locator (universal outlets)"}
_EXTRACT_SRC = {eid: src for src, exs in FL_CONN.items() for (eid, _h, _n) in exs}
_NAME_COLS = ("Owner Name", "Registrant Name", "Applicant", "Brand Name", "DBA")
def _name_idx(header):
    for n in _NAME_COLS:
        if n in header: return header.index(n)
    return 2 if len(header) > 2 else 0
def derive_hierarchy(datasets, top=50):
    """First-cut scope tree from real data: source -> entity (top-N by row count).
    Returns None when there's nothing to derive, so the caller falls back to the seed."""
    from collections import Counter
    by_src = {}
    for dsid, ds in datasets.items():
        rows = ds.get("rows") or []
        if not rows: continue
        ci = _name_idx(ds.get("header") or [])
        src = _EXTRACT_SRC.get(dsid) or ("ttb-cola" if "cola" in dsid.lower() else "other")
        counts = by_src.setdefault(src, Counter())
        for r in rows:
            v = str(r[ci]).strip() if ci < len(r) else ""
            if v: counts[v] += 1
    if not by_src: return None
    pfs = []
    for src, counts in by_src.items():
        brands = [{"id": f"sc:{src}:{i}", "level": "brand", "name": nm, "count": n, "children": []}
                  for i, (nm, n) in enumerate(counts.most_common(top))]
        pfs.append({"id": f"sc:{src}", "level": "portfolio", "name": _SRC_LABEL.get(src, src.title()),
                    "children": brands})
    pfs.sort(key=lambda p: p["name"])
    return {"id": "root", "level": "root", "name": "All Sources", "children": pfs}

def _compute_live_datasets():
    out = {k: dict(v) for k, v in DATASETS.items()}
    try:
        import warehouse
        for d in warehouse.list_datasets():
            k = d.get("name", "")
            if not k or k.startswith("_"):
                continue
            total = d.get("rows") or 0
            if k in out:
                out[k]["total"] = total                       # refresh the cached count to live truth
                out[k].setdefault("store", "warehouse")
            else:
                out[k] = {"header": d.get("fields") or [], "rows": [], "total": total,
                          "store": "warehouse", "live": True}
    except Exception as e:
        app.logger.warning("live datasets merge failed: %s", e)
    return out


import threading as _threading
_DS_CACHE = {"t": 0.0, "v": None, "refreshing": False}


def _live_datasets():
    """DATASETS (in-memory samples) MERGED with the LIVE warehouse catalog. The warehouse footer scan
    (list_datasets over ~170 Tigris files) can take many seconds under crawl write-contention, so this is cached
    with STALE-WHILE-REVALIDATE: once warm it returns instantly and refreshes in the background — the endpoint
    never hangs (which read as 'site down' while a national crawl hammered Tigris)."""
    import time as _t
    c = _DS_CACHE
    now = _t.time()
    if c["v"] is not None and now - c["t"] < 45:
        return c["v"]                                         # fresh
    # cold OR stale → kick a background refresh and NEVER block the request thread on the 170-file Tigris footer
    # scan. That synchronous cold scan could exceed gunicorn's 120s timeout under crawl write-contention → the
    # worker was killed and the machine crash-looped (site "down"). Cold now returns the in-memory base instantly;
    # live warehouse counts populate a few seconds later once the daemon thread finishes.
    if not c["refreshing"]:
        c["refreshing"] = True

        def _bg():
            try:
                v = _compute_live_datasets(); c["v"] = v; c["t"] = _t.time()
            except Exception as e:
                app.logger.warning("datasets refresh failed: %s", e)
            finally:
                c["refreshing"] = False
        _threading.Thread(target=_bg, daemon=True).start()
    if c["v"] is not None:
        return c["v"]                                         # stale — serve it
    return {k: dict(v) for k, v in DATASETS.items()}          # cold — instant base; live counts fill in async

@app.get("/api/datasets")
def datasets():
    q = (request.args.get("q") or "").strip().lower()
    only = request.args.get("dataset")
    allds = _live_datasets()
    src = {only: allds[only]} if only in allds else allds
    if not q:
        return jsonify(src)
    out = {}
    for k, ds in src.items():
        rows = [r for r in (ds.get("rows") or []) if any(q in str(c).lower() for c in r)]
        out[k] = dict(ds, rows=rows, matched=len(rows))
    return jsonify(out)


@app.get("/api/source")
def api_source():
    """Live raw-data viewer feed. No ?name → the source list with live row counts (for the picker). With
    ?name=<table> → that source's newest rows straight from the warehouse, so a UI can poll and watch it build."""
    import warehouse
    name = request.args.get("name")
    if not name:
        def _srclist():                                       # cached + background-refreshed so the 170-file
            return sorted(({"name": d.get("name"), "rows": d.get("rows", 0) or 0, "modified": d.get("modified")}
                           for d in warehouse.list_datasets() if d.get("name") and not d["name"].startswith("_")),
                          key=lambda x: x["name"])              # Tigris scan never blocks the worker (crash-loop)
        return jsonify({"sources": _ttl("source_list", 60, _srclist, cold=[])})
    limit = min(int(request.args.get("limit", 100) or 100), 500)
    try:
        cols = list(warehouse.query(name, "SELECT * FROM t LIMIT 1")[0].keys())
        # newest first if the table has a time-ish column (quoted — 'at' is a DuckDB reserved word)
        order = next((c for c in ("captured_at", "pulled_at", "at", "modified", "date", "run_id") if c in cols), None)
        sql = "SELECT * FROM t" + (' ORDER BY "%s" DESC' % order if order else "") + " LIMIT %d" % limit
        rows = warehouse.query(name, sql)
        cnt = warehouse.query(name, "SELECT count(*) c FROM t")[0]["c"]
        # drop bulky raw_json from the wire payload (keep a marker) so the live table stays snappy
        for r in rows:
            if r.get("raw_json") is not None:
                r["raw_json"] = "…"
        return jsonify({"name": name, "count": cnt, "columns": [c for c in cols], "rows": rows, "order_by": order})
    except Exception as e:
        return jsonify({"name": name, "error": str(e)[:200], "rows": [], "count": 0, "columns": []}), 200


# ── Data Console: the one trustworthy monitoring surface (pull health · counts · errors · live data) ──────────
@app.get("/api/monitor")
def api_monitor():
    """The manifest: every source/scrape/account/master with real counts (parquet footers), freshness, run-health,
    deltas, plus connector rollup + totals. Stamped (as_of/live/cached/warehouse) so the console proves it's real.
    Cheap + cached — safe to poll. ?fresh=1 forces a synchronous rebuild."""
    import monitor
    try:
        if request.args.get("fresh") == "1":
            monitor._refresh_async()          # kick a background rebuild; never block the request on a ~130s build
        return jsonify(monitor.snapshot())
    except Exception as e:
        return jsonify({"live": False, "error": str(e)[:200], "sources": [], "totals": {}}), 200


@app.get("/api/monitor/health")
def api_monitor_health():
    """The daily data-health VERDICT (health_digest.py): deterministic findings w/ evidence + first_seen. Reads the
    local digest when this host produced one (the Mac), else the copy the digest pushes to the warehouse bucket —
    so the Fly-served console shows the SAME verdict. Always stamped; honest 'no digest yet' when neither exists.
    The optional Claude triage rides along only where it exists locally (judgment layer — never part of the verdict)."""
    base = os.path.join(os.path.dirname(os.path.abspath(__file__)), "agent_state", "health")
    digest, origin = None, None
    try:
        with open(os.path.join(base, "latest.json")) as f:
            digest, origin = json.load(f), "local"
    except Exception:
        try:
            import warehouse
            raw = warehouse.get_bytes("_health_digest.json")
            if raw:
                digest, origin = json.loads(raw), "warehouse"
        except Exception:
            pass
    if not digest:
        return jsonify({"available": False, "note": "no digest yet — health_digest.py has not run"}), 200
    triage = None
    try:
        with open(os.path.join(base, "latest_triage.md")) as f:
            triage = f.read()
    except Exception:
        pass
    return jsonify(dict(digest, available=True, origin=origin, triage=triage,
                        age_s=round(time.time() - (digest.get("as_of") or 0))))


_DATA_SAMPLE_CACHE = {}                       # name -> (built_at, payload); makes re-opening a source instant
_MON_CON = {"con": None}                       # ONE warm DuckDB connection reused for drawer reads
_MON_LOCK = threading.Lock()


def _mon_con():
    """A persistent DuckDB connection for the console's data reads. `INSTALL/LOAD httpfs` + S3 config costs ~6s and
    was being paid on EVERY warehouse.query() — reusing one warm connection drops a small-table read from ~18s to
    ~2s against object storage. Pre-warmed at boot so the first user click isn't the one that pays the 6s."""
    if _MON_CON["con"] is None:
        import warehouse
        _MON_CON["con"] = warehouse.connect()
    return _MON_CON["con"]


def _query_timeout(name, sql, timeout_s):
    """Run a drawer read on the warm connection with a hard timeout via DuckDB interrupt() — object storage from the
    cloud host is high-latency and a big-table read can take >60s, so this bounds it (the connection stays usable
    after an interrupt). Serialized by a lock (one shared connection; drawer reads are infrequent). Returns
    (rows|None, timed_out)."""
    import warehouse
    with _MON_LOCK:
        try:
            import monitor
            con = _mon_con()
            src = monitor.read_expr(name)                 # partitioned-aware (glob for retail_observations etc.)
            timed = {"v": False}

            def kill():
                timed["v"] = True
                try:
                    con.interrupt()
                except Exception:
                    pass
            timer = threading.Timer(timeout_s, kill)
            timer.start()
            try:
                con.execute("CREATE OR REPLACE VIEW t AS SELECT * FROM %s" % src)
                cur = con.execute(sql)
                cols = [d[0] for d in cur.description]
                rows = [dict(zip(cols, row)) for row in cur.fetchall()]
                return rows, False
            except Exception:
                return None, timed["v"]          # interrupted (timeout) or read error
            finally:
                timer.cancel()
        except Exception:
            _MON_CON["con"] = None                # drop a broken connection so the next read reconnects clean
            return None, False


@app.get("/api/monitor/data")
def api_monitor_data():
    """Real rows for ONE source — the 'view the actual data' pane. SCALES to 100M+ row datasets because it does NOT
    read the lake at request time: it serves a PRECOMPUTED preview sample held in memory (built + refreshed in the
    background, bounded to N rows per dataset). Exact count comes from the manifest footer. A dataset not yet sampled
    falls back to one bounded, timeout-guarded live read (and gets picked up by the next background pass)."""
    import monitor
    name = request.args.get("name", "")
    if not name:
        return jsonify({"error": "name required"}), 400
    try:
        total = next((s["rows"] for s in monitor.snapshot().get("sources", []) if s["name"] == name), None)
    except Exception:
        total = None

    smp = monitor.sample(name)
    if smp and not request.args.get("fresh"):
        return jsonify({"name": name, "count": total, "columns": smp.get("columns", []),
                        "rows": smp.get("rows", []), "order_by": None, "cached": True,
                        "sampled_at": smp.get("at"),
                        "note": "preview sample — precomputed, refreshed in the background (instant at any scale)"})

    # no precomputed sample yet → one bounded, timeout-guarded live read; the background pass will cache it next
    limit = min(int(request.args.get("limit", 60) or 60), 200)
    rows, timed_out = _query_timeout(name, "SELECT * FROM t LIMIT %d" % limit, timeout_s=20)
    if timed_out or rows is None:
        return jsonify({"name": name, "count": total, "columns": [], "rows": [], "order_by": None,
                        "note": "building this source's preview… (exact row count above is from file metadata)"}), 200
    cols = list(rows[0].keys()) if rows else []
    for r in rows:
        if r.get("raw_json") is not None:
            r["raw_json"] = "…"
    return jsonify({"name": name, "count": total, "columns": cols, "rows": rows, "order_by": None,
                    "note": "live sample (preview being precomputed in the background)"})


@app.get("/api/monitor/source")
def api_monitor_source():
    """One SOURCE's drill-in: its datasets + its ACCOUNTS-WITH-INVENTORY (per-store skus / in-stock / units on
    the source's latest observation day). Reads only that source's latest-date part files (filename-pruned),
    cached until a new part lands — so the console can filter any scrape down to real accounts instantly."""
    import monitor
    key = request.args.get("key", "")
    if not key:
        return jsonify({"error": "key required"}), 400
    try:
        return jsonify(monitor.source_detail(key))
    except Exception as e:
        return jsonify({"key": key, "accounts": [], "note": "error: %s" % str(e)[:150]}), 200


@app.get("/api/monitor/search")
def api_monitor_search():
    """Search everything — one query across all datasets / all rows / all columns. Returns instant hits (samples +
    names) immediately, then a progressive deep-sweep snapshot; poll the same URL until done=true."""
    import monitor
    q = (request.args.get("q") or "").strip()
    try:
        return jsonify(monitor.search(q))
    except Exception as e:
        return jsonify({"q": q, "done": True, "error": str(e)[:200], "instant": [], "hits": []}), 200


@app.get("/api/monitor/history")
def api_monitor_history():
    """The saved row-count time-series for one source → [[ts, rows], …] (rows-over-time sparkline)."""
    import monitor
    name = request.args.get("name", "")
    return jsonify({"name": name, "series": monitor.history(name) if name else []})


@app.post("/api/monitor/rerun")
def api_monitor_rerun():
    """Trigger a pull for a connector/source from the console — reuses the existing run path so behavior matches
    the Pulls surface exactly."""
    conn = (request.get_json(silent=True) or {}).get("conn") or request.args.get("conn", "")
    if not conn:
        return jsonify({"error": "conn required"}), 400
    try:
        import monitor
        monitor._CACHE["at"] = 0                       # invalidate so the next snapshot reflects the rerun
    except Exception:
        pass
    with app.test_request_context(json={"connId": conn}):
        return run()                                   # delegate to the canonical /api/run handler


# TTL memoize for read-only rollups the estate/coverage UIs POLL. Each of these fans out into many remote-parquet
# scans (a /api/catalog poll was ~25s, /api/coverage ~8s); polling stacked them into timeouts. These only change
# on a pull/rebuild, so a short TTL is safe and turns the poll into a served-from-RAM hit.
_TTL_CACHE = {}


_TTL_REFRESHING = {}


def _ttl(key, ttl, fn, cold=None):
    """TTL cache with STALE-WHILE-REVALIDATE that NEVER blocks a request thread — even on the cold first call.
    A cold fn() over the now-huge warehouse (170+ Tigris files) could exceed gunicorn's 120s worker timeout →
    the worker is killed and the machine crash-loops (site 'down'). So cold kicks a background compute and returns
    `cold` (None/[] — the polling UI fills in on its next tick); stale serves the last value while refreshing."""
    now = time.time()
    hit = _TTL_CACHE.get(key)
    if hit and (now - hit[0]) < ttl:
        return hit[1]                                   # fresh
    if not _TTL_REFRESHING.get(key):                    # cold OR stale → refresh in a daemon thread, never block
        _TTL_REFRESHING[key] = True

        def _bg():
            try:
                _TTL_CACHE[key] = (time.time(), fn())
            except Exception as e:
                app.logger.warning("ttl refresh %s failed: %s", key, e)
            finally:
                _TTL_REFRESHING[key] = False
        _threading.Thread(target=_bg, daemon=True).start()
    return hit[1] if hit is not None else cold          # stale → last value; cold → placeholder until first land


def _catalog_map():
    """Every dataset we hold, across all three stores: {key: {header, total, store}}. memory = in-RAM
    sample; full = save_full'd on Tigris; warehouse = Parquet. Precedence memory → full → warehouse."""
    out = {}
    for k, ds in DATASETS.items():
        out[k] = {"header": ds.get("header") or [],
                  "total": ds.get("total") or len(ds.get("rows") or []), "store": "memory"}
    for k, v in (_full_index() or {}).items():
        if k not in out or not out[k].get("total"):
            out[k] = {"header": v.get("header") or [], "total": v.get("total") or 0, "store": "full"}
    try:
        import warehouse
        for d in warehouse.list_datasets():
            k = d.get("name", "")
            if k and not k.startswith("_"):
                out.setdefault(k, {"header": d.get("fields") or [], "total": d.get("rows") or 0,
                                   "store": "warehouse"})
    except Exception as e:
        app.logger.warning("warehouse catalog failed: %s", e)
    return out

@app.get("/api/velocity")
def api_velocity():
    """The velocity SURFACE payload (MOAT-PLAN V6). MARTS ONLY — signal_voids / signal_movers /
    mart_velocity_brand_week + the calibration trust stamp — never the raw fact grain. Every section
    carries confidence + as_of so nothing renders without proof of how trustworthy it is. Honest by
    construction: movers split into CONFIRMED (both weeks full) vs early-read (partial); if no full
    week exists yet the confirmed list is simply empty. Empty-safe on any missing table."""
    import time as _t
    import warehouse

    def q(name, sql):
        try:
            return warehouse.query(name, sql)
        except Exception:
            return []

    voids = q("signal_voids",
              "SELECT brand, oos_stores, stores_seen, est_recoverable_units_wk, pct_stores_out, confidence "
              "FROM t ORDER BY est_recoverable_units_wk DESC NULLS LAST LIMIT 100")
    confw = q("signal_movers", "SELECT MAX(week) w FROM t WHERE NOT partial")
    cw = confw[0]["w"] if confw and confw[0].get("w") else None
    movers_up, movers_down = [], []
    if cw:
        movers_up = q("signal_movers",
                      "SELECT brand, week, units, prev_units, pct_change, matched_cells, confidence FROM t "
                      "WHERE NOT partial AND week = (SELECT MAX(week) FROM t WHERE NOT partial) AND pct_change > 0 "
                      "ORDER BY pct_change DESC LIMIT 25")
        movers_down = q("signal_movers",
                        "SELECT brand, week, units, prev_units, pct_change, matched_cells, confidence FROM t "
                        "WHERE NOT partial AND week = (SELECT MAX(week) FROM t WHERE NOT partial) AND pct_change < 0 "
                        "ORDER BY pct_change ASC LIMIT 25")
    n_partial = (q("signal_movers", "SELECT COUNT(*) n FROM t WHERE partial") or [{"n": 0}])[0].get("n", 0)
    leaders = q("mart_velocity_brand_week",
                "SELECT brand, week, implied_units, stores, avg_confidence FROM t "
                "WHERE week = (SELECT MAX(week) FROM t) ORDER BY implied_units DESC NULLS LAST LIMIT 50")
    cons = q("velocity_calibration",
             "SELECT source, value FROM t WHERE kind='conservation' AND metric='sales_restock_ratio'")
    mape = q("velocity_calibration",
             "SELECT anchor, coverage, value FROM t WHERE kind='external_mape'")
    return jsonify({
        "as_of": _t.time(), "live": True,
        "voids": voids,
        "movers": {"confirmed_week": str(cw) if cw else None, "up": movers_up, "down": movers_down,
                   "partial_flagged": n_partial,
                   "note": None if cw else "No confirmed movers yet — needs two consecutive full "
                                           "observation weeks. Early-read (partial-week) movers are staged."},
        "leaders": leaders,
        "calibration": {"conservation": cons,
                        "external_mape": mape,
                        "note": "conservation ratio → 1 is well-calibrated; external MAPE pending an "
                                "overlapping ground-truth footprint"},
    })


@app.get("/api/catalog")
def catalog_ep():
    """The estate model's 'whole thing' view (it polls this so new datasets appear on their own)."""
    return jsonify(_ttl("catalog", 90, _catalog_map))


# ── Source Spec — per source, the fields we capture (% populated) + sample rows. LIVE off _stage_product so the
# field-capture picture updates itself every rebuild (the page reads this instead of a baked-in snapshot). ──
_SPEC_FIELDS = ["brand", "product_name", "class_type", "varietal", "flavor", "category", "style", "abv", "size_ml",
                "container", "pack", "upc", "vintage", "region", "sub_region", "appellation", "country", "state",
                "origin", "taste", "body", "food_pairing", "expert_rating", "finish", "image"]


def _source_spec_data():
    import warehouse
    sel = ", ".join("count(*) FILTER(WHERE nullif(trim(CAST(%s AS VARCHAR)),'') IS NOT NULL) AS \"%s\"" % (f, f)
                    for f in _SPEC_FIELDS)
    try:
        rows = warehouse.query("_stage_product", "SELECT _source, count(*) n, %s FROM t GROUP BY _source "
                               "ORDER BY n DESC" % sel)
    except Exception as e:
        return {"ok": False, "error": str(e)[:140], "sources": []}
    spec = []
    for r in rows:
        n = r["n"] or 1
        fields = {f: round(100.0 * (r.get(f) or 0) / n) for f in _SPEC_FIELDS if r.get(f)}
        try:
            ex = [(e.get("product_name") or "")[:52] for e in warehouse.query(
                "_stage_product", "SELECT product_name FROM t WHERE _source = ? AND product_name IS NOT NULL "
                "LIMIT 3", [r["_source"]])]
        except Exception:
            ex = []
        spec.append({"source": r["_source"], "rows": n, "fields_pct": fields, "examples": ex})
    return {"ok": True, "sources": spec, "all_fields": _SPEC_FIELDS}


@app.get("/api/master/source-spec")
def source_spec_ep():
    """Live field-capture spec per source (cached 10 min — only changes on a rebuild)."""
    return jsonify(_ttl("source_spec", 600, _source_spec_data))

# ── /api/sources — the SINGLE source of truth for what we hold + what each dataset FEEDS. Driven by the
# mappings/seeds themselves (authoritative) + name heuristics, so labels/tags stay accurate as the list
# grows. Powers the MDM Sources view and keeps the estate tags current (no more hardcoded name map). ──
_FEED_ORDER = ["master", "outlet", "product", "inventory", "pricing", "reference", "other"]
_FEED_GROUP = {"master": "Owned masters", "outlet": "Outlet sources", "product": "Product & label sources",
               "inventory": "Inventory feeds", "pricing": "Pricing feeds", "reference": "Reference data",
               "other": "Unclassified"}
_DS_LABELS = {"outlets_master": "Outlet master", "ab_outlets": "AB InBev · Retailer locator",
              "total_wine_products": "Total Wine · Catalog", "ttb_cola": "TTB · COLA labels",
              "cola_cluster": "COLA · Product clusters", "bc_liquor": "BC Liquor · Price list",
              "or_pricing": "Oregon OLCC · Pricing", "orlando_accounts": "Orlando · On-premise accounts",
              "census_acs": "Census · ACS demographics", "census_reference": "Census · CBP/NES/PEP",
              "outlets_census": "Outlets × census"}

def _ds_label(name):
    if name in _DS_LABELS:
        return _DS_LABELS[name]
    if name.startswith("dim_"):
        return "Master · " + name[4:].replace("_", " ").title()
    if name.startswith("fact_"):
        return "Fact · " + name[5:].replace("_", " ").title()
    if name.startswith("vtinfo_"):
        return "VTInfo · " + name[7:].replace("-", " ").title()
    return name.replace("_", " ").replace("-", " ").title()

def _ds_feeds(name, mp):
    """Which master(s) a dataset feeds — mappings/seeds are authoritative; heuristics fill unmapped ones."""
    if name.startswith("dim_") or name.startswith("fact_"):
        return ["master"]
    f = set()
    if _seed_for(name) or name in mp["outlet"]:
        f.add("outlet")
    if name in mp["product"]:
        f.add("product")
    for fact in FACTS:
        if _fact_seed_for(fact, name) or name in mp.get(fact, {}):
            f.add(fact)
    if not f:
        n = name.lower()
        if re.search(r"census|_census", n): f.add("reference")
        elif re.search(r"\bcola\b|ttb", n): f.add("product")
        elif re.search(r"outlet|licens|premise|account|_sla|tabc|dcp|\bny_|\bco_|\bmo_|\bct_|\btx_|\bil_", n): f.add("outlet")
        elif re.search(r"pric|bc_liquor|or_pricing|total_wine|iowa_prod|catalog|mont_", n): f.add("product")
        elif re.search(r"binny|spec|abc_|instacart|shopify|iowa_sales", n): f.add("inventory")
    return [x for x in _FEED_ORDER if x in f] or ["other"]

@app.get("/api/sources")
def sources_ep():
    cat = _catalog_map()
    mp = {"outlet": load(ENTITIES["outlet"]["maps"], {}), "product": load(ENTITIES["product"]["maps"], {})}
    for fact in FACTS:
        mp[fact] = load(FACTS[fact]["maps"], {})
    items = []
    for k, v in cat.items():
        feeds = _ds_feeds(k, mp)
        is_master = feeds[0] == "master"
        mapped = (any(k in mp[e] for e in mp) or bool(_seed_for(k))
                  or any(_fact_seed_for(fc, k) for fc in FACTS))
        items.append({"id": k, "label": _ds_label(k), "rows": v.get("total", 0),
                      "store": v.get("store"), "cols": len(v.get("header") or []), "feeds": feeds,
                      "group": _FEED_GROUP[feeds[0]],
                      "status": "master" if is_master else ("mapped" if mapped else "unmapped")})
    gi = {_FEED_GROUP[fk]: i for i, fk in enumerate(_FEED_ORDER)}
    items.sort(key=lambda x: (gi.get(x["group"], 9), -(x["rows"] or 0)))
    present = [_FEED_GROUP[fk] for fk in _FEED_ORDER if any(it["group"] == _FEED_GROUP[fk] for it in items)]
    return jsonify(ok=True, count=len(items), groups=present, sources=items)

# ── Product registrations (TTB COLA) — label-approval filings + distinct-product clusters ──
def _cola_q(name, sql, params=None):
    import warehouse
    return warehouse.query(name, sql, params or [])

@app.get("/api/cola/stats")
def cola_stats_ep():
    """Overview for the product-registration page: total filings, distinct product clusters, UPC
    coverage, filings-by-year, top class/types + applicants. Reads the warehouse Parquet via DuckDB."""
    try:
        total = _cola_q("ttb_cola", "SELECT count(*) c FROM t")[0]["c"]
    except Exception:
        return jsonify(ok=True, landed=False, total=0, note="ttb_cola not landed yet")
    def q(sql):
        try: return _cola_q("ttb_cola", sql)
        except Exception: return []
    by_year = q("SELECT substr(\"Completed Date\",7,4) yr, count(*) n FROM t "
                "WHERE length(\"Completed Date\")>=10 GROUP BY 1 ORDER BY yr")
    top_class = q("SELECT \"Class/Type\" k, count(*) n FROM t WHERE \"Class/Type\"<>'' GROUP BY 1 ORDER BY n DESC LIMIT 12")
    top_appl = q("SELECT \"Applicant\" k, count(*) n FROM t WHERE \"Applicant\"<>'' GROUP BY 1 ORDER BY n DESC LIMIT 12")
    upc = q("SELECT count(*) c FROM t WHERE \"UPC\"<>''")
    clusters = None
    try: clusters = _cola_q("cola_cluster", "SELECT count(*) c FROM t")[0]["c"]
    except Exception: pass
    return jsonify(ok=True, landed=True, total=total, clusters=clusters,
                   with_upc=(upc[0]["c"] if upc else 0), by_year=by_year,
                   top_class=top_class, top_applicant=top_appl)

@app.get("/api/cola/registrations")
def cola_regs_ep():
    """Paged + searchable COLA filings (brand / applicant / TTB ID / UPC), server-side over ~1M+ rows."""
    q = (request.args.get("q") or "").strip()
    try:
        off = max(0, int(request.args.get("offset", "0") or 0)); lim = min(200, max(1, int(request.args.get("limit", "50") or 50)))
    except ValueError:
        off, lim = 0, 50
    where, params = "1=1", []
    if q:
        where = "(\"Brand Name\" ILIKE ? OR \"Applicant\" ILIKE ? OR \"TTB ID\"=? OR \"UPC\"=?)"
        params = ["%" + q + "%", "%" + q + "%", q, q]
    cols = '"TTB ID","Brand Name","Fanciful Name","Class/Type","Origin","Applicant","Completed Date","Net Contents","UPC"'
    try:
        total = _cola_q("ttb_cola", "SELECT count(*) c FROM t WHERE " + where, params)[0]["c"]
        rows = _cola_q("ttb_cola", "SELECT %s FROM t WHERE %s ORDER BY \"TTB ID\" DESC LIMIT %d OFFSET %d"
                       % (cols, where, lim, off), params)
    except Exception as e:
        return jsonify(ok=False, landed=False, error=str(e)[:140], rows=[]), 200
    return jsonify(ok=True, total=total, offset=off, limit=lim, rows=rows)

@app.get("/api/cola/clusters")
def cola_clusters_ep():
    """Paged distinct-PRODUCT clusters (collapsed filing noise) — brand/fanciful/class/size/supplier,
    member count, confidence, flagged. From cola_cluster (the scale dedup layer)."""
    q = (request.args.get("q") or "").strip()
    try:
        off = max(0, int(request.args.get("offset", "0") or 0)); lim = min(200, max(1, int(request.args.get("limit", "50") or 50)))
    except ValueError:
        off, lim = 0, 50
    where, params = "1=1", []
    if q:
        where = "(brand ILIKE ? OR fanciful ILIKE ? OR supplier ILIKE ?)"
        params = ["%" + q + "%"] * 3
    try:
        total = _cola_q("cola_cluster", "SELECT count(*) c FROM t WHERE " + where, params)[0]["c"]
        rows = _cola_q("cola_cluster", "SELECT cluster_id, brand, fanciful, class_type, size_ml, supplier, "
                       "member_count, confidence, flagged FROM t WHERE %s ORDER BY member_count DESC LIMIT %d OFFSET %d"
                       % (where, lim, off), params)
    except Exception as e:
        return jsonify(ok=False, landed=False, error=str(e)[:140], rows=[]), 200
    return jsonify(ok=True, total=total, offset=off, limit=lim, rows=rows)

COLA_FIELDS = ["Brand Name", "Fanciful Name", "Class/Type", "Origin", "Applicant",
               "Status", "Net Contents", "UPC"]

@app.get("/api/cola/profile")
def cola_profile_ep():
    """Field-level DATA DICTIONARY for ttb_cola — per column: fill-rate + distinct cardinality + a few
    samples. The basis for string extraction: which registration fields carry structured signal worth
    parsing into controlled vocabularies + derived fields."""
    try:
        total = _cola_q("ttb_cola", "SELECT count(*) c FROM t")[0]["c"]
    except Exception:
        return jsonify(ok=True, landed=False, fields=[])
    out = []
    for col in COLA_FIELDS:
        try:
            r = _cola_q("ttb_cola", 'SELECT count(*) FILTER (WHERE "%s"<>\'\') filled, '
                        'count(DISTINCT "%s") dct FROM t' % (col, col))[0]
            samp = _cola_q("ttb_cola", 'SELECT "%s" v FROM t WHERE "%s"<>\'\' LIMIT 3' % (col, col))
        except Exception:
            continue
        out.append({"field": col, "filled": r["filled"], "distinct": r["dct"],
                    "fill_pct": round(100 * r["filled"] / max(1, total), 1),
                    "samples": [s["v"] for s in samp]})
    return jsonify(ok=True, landed=True, total=total, fields=out)

@app.get("/api/cola/dictionary")
def cola_dictionary_ep():
    """The value DICTIONARY (controlled vocabulary) for one field: distinct values + counts, desc —
    the raw material for a data dictionary and for deriving a normalized field from free text."""
    field = (request.args.get("field") or "Class/Type").strip()
    if field not in COLA_FIELDS:
        return jsonify(ok=False, error="unknown field"), 400
    try:
        lim = min(500, max(1, int(request.args.get("limit", "150") or 150)))
    except ValueError:
        lim = 150
    q = (request.args.get("q") or "").strip()
    where, params = '"%s"<>\'\'' % field, []
    if q:
        where += ' AND "%s" ILIKE ?' % field; params = ["%" + q + "%"]
    try:
        distinct = _cola_q("ttb_cola", 'SELECT count(DISTINCT "%s") c FROM t WHERE %s' % (field, where), params)[0]["c"]
        rows = _cola_q("ttb_cola", 'SELECT "%s" v, count(*) n FROM t WHERE %s GROUP BY 1 ORDER BY n DESC LIMIT %d'
                       % (field, where, lim), params)
    except Exception as e:
        return jsonify(ok=False, error=str(e)[:140], values=[]), 200
    return jsonify(ok=True, field=field, distinct=distinct, values=rows)

# ── Generic field dictionary — profile + value vocabulary for ANY item/product dataset ──
def _mem_dataset(ds):
    """In-RAM datasets (pull output: outlets, vtinfo, ab_outlets…) as (header, rows) — so the
    data dictionary works on ANY dataset in the catalog, not only warehouse Parquet ones."""
    d = DATASETS.get(ds)
    if not d:
        return None
    header = d.get("header") or []
    rows = d.get("_rows_full") or d.get("rows") or []
    return (header, rows) if header else None

def _ds_columns(ds):
    try:
        s = _cola_q(ds, "SELECT * FROM t LIMIT 1")
        if s:
            return list(s[0].keys())
    except Exception:
        pass
    mem = _mem_dataset(ds)
    return list(mem[0]) if mem else []

@app.get("/api/item/profile")
def item_profile_ep():
    """Field-level DATA DICTIONARY for ANY warehouse item/product dataset (?dataset=bc_liquor,
    or_pricing, iowa_products, ttb_cola…). Introspects columns, then per column: fill-rate + distinct
    cardinality + samples — the basis for string extraction into derived fields on any dataset."""
    ds = (request.args.get("dataset") or "ttb_cola").strip()
    cols = _ds_columns(ds)[:24]
    if not cols:
        return jsonify(ok=True, landed=False, dataset=ds, fields=[])
    try:
        total = _cola_q(ds, "SELECT count(*) c FROM t")[0]["c"]
    except Exception:
        mem = _mem_dataset(ds)                                    # in-RAM dataset (e.g. outlets): profile in Python
        if not mem:
            return jsonify(ok=True, landed=False, dataset=ds, fields=[])
        header, rows = mem
        out = []
        for ci, col in enumerate(header[:24]):
            if col.startswith(":@") or col.startswith("_"):
                continue
            vals = [str(r[ci]) for r in rows if ci < len(r) and str(r[ci]).strip() != ""]
            out.append({"field": col, "filled": len(vals), "distinct": len(set(vals)),
                        "fill_pct": round(100 * len(vals) / max(1, len(rows)), 1), "samples": vals[:3]})
        return jsonify(ok=True, landed=True, dataset=ds, total=len(rows), fields=out)
    out = []
    for col in cols:
        if col.startswith(":@") or col.startswith("_"):
            continue
        try:
            r = _cola_q(ds, 'SELECT count(*) FILTER (WHERE CAST("%s" AS VARCHAR)<>\'\') filled, '
                        'count(DISTINCT "%s") dct FROM t' % (col, col))[0]
            samp = _cola_q(ds, 'SELECT "%s" v FROM t WHERE CAST("%s" AS VARCHAR)<>\'\' LIMIT 3' % (col, col))
        except Exception:
            continue
        out.append({"field": col, "filled": r["filled"], "distinct": r["dct"],
                    "fill_pct": round(100 * r["filled"] / max(1, total), 1),
                    "samples": [str(s["v"]) for s in samp]})
    return jsonify(ok=True, landed=True, dataset=ds, total=total, fields=out)

@app.get("/api/item/dictionary")
def item_dictionary_ep():
    """Value DICTIONARY (controlled vocabulary) for one field of ANY item dataset: distinct values +
    counts, desc. ?dataset=&field=&q=. The raw material for deriving a normalized field from text."""
    ds = (request.args.get("dataset") or "ttb_cola").strip()
    field = (request.args.get("field") or "").strip()
    if field not in _ds_columns(ds):
        return jsonify(ok=False, error="unknown field for dataset"), 400
    try:
        lim = min(500, max(1, int(request.args.get("limit", "150") or 150)))
    except ValueError:
        lim = 150
    q = (request.args.get("q") or "").strip()
    where, params = 'CAST("%s" AS VARCHAR)<>\'\'' % field, []
    if q:
        where += ' AND CAST("%s" AS VARCHAR) ILIKE ?' % field; params = ["%" + q + "%"]
    try:
        distinct = _cola_q(ds, 'SELECT count(DISTINCT "%s") c FROM t WHERE %s' % (field, where), params)[0]["c"]
        rows = _cola_q(ds, 'SELECT CAST("%s" AS VARCHAR) v, count(*) n FROM t WHERE %s GROUP BY 1 ORDER BY n DESC LIMIT %d'
                       % (field, where, lim), params)
    except Exception:
        mem = _mem_dataset(ds)                                    # in-RAM dataset: value counts in Python
        if not mem:
            return jsonify(ok=False, error="dataset not queryable", values=[]), 200
        header, rows_all = mem
        ci = header.index(field)
        from collections import Counter
        c = Counter(str(r[ci]) for r in rows_all if ci < len(r) and str(r[ci]).strip() != ""
                    and (not q or q.lower() in str(r[ci]).lower()))
        return jsonify(ok=True, dataset=ds, field=field, distinct=len(c),
                       values=[{"v": v, "n": n} for v, n in c.most_common(lim)])
    return jsonify(ok=True, dataset=ds, field=field, distinct=distinct, values=rows)

# ── DICTIONARY STORE — named, reusable value dictionaries (synonym-normalize or string-extract) ──
# A dictionary = ordered [{match,value}] + mode (exact|contains|regex) + fallback (original|null). It
# compiles (derive.dict_case) to ONE DuckDB CASE, so 'cab sauv → Cabernet Sauvignon' runs at scale. A
# field-mapping row can reference one by id ({mode:'dict', dict:<id>}); the build resolves it inline.
DICT_MODES = ["exact", "contains", "regex"]
def _dicts():
    return load("dictionaries.json", [])
def _dict_by_id(did):
    return next((d for d in _dicts() if d.get("id") == did), None)

@app.get("/api/dict/store")
def dict_store_list():
    """All saved dictionaries (light): id, name, source field, mode, fallback, entry count."""
    out = [{"id": d.get("id"), "name": d.get("name"), "source_dataset": d.get("source_dataset"),
            "source_field": d.get("source_field"), "target_field": d.get("target_field"),
            "mode": d.get("mode"), "fallback": d.get("fallback"), "entries": len(d.get("entries") or []),
            "updated": d.get("updated"), "by": d.get("by")} for d in _dicts()]
    return jsonify(ok=True, dictionaries=out, modes=DICT_MODES)

@app.get("/api/dict/store/get")
def dict_store_get():
    d = _dict_by_id((request.args.get("id") or "").strip())
    return jsonify(ok=bool(d), dictionary=d) if d else (jsonify(ok=False, error="not found"), 404)

@app.post("/api/dict/store/save")
def dict_store_save():
    """Upsert a dictionary. Body: {id?, name, source_dataset, source_field, target_field, mode,
    fallback, entries:[{match,value}]}. Returns the saved record (with a minted id if new)."""
    b = request.get_json(silent=True) or {}
    name = (b.get("name") or "").strip()
    if not name:
        return jsonify(ok=False, error="name required"), 400
    entries = [{"match": str(e.get("match", "")).strip(), "value": str(e.get("value", "")).strip()}
               for e in (b.get("entries") or []) if str(e.get("match", "")).strip() != ""]
    dicts = _dicts()
    did = (b.get("id") or "").strip()
    rec = _dict_by_id(did) if did else None
    if not rec:
        did = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")[:40] or "dict"
        base, i = did, 2
        while _dict_by_id(did):
            did = "%s-%d" % (base, i); i += 1
        rec = {"id": did}; dicts.append(rec)
    rec.update({"name": name, "source_dataset": (b.get("source_dataset") or "").strip(),
                "source_field": (b.get("source_field") or "").strip(),
                "target_field": (b.get("target_field") or "").strip(),
                "mode": (b.get("mode") if b.get("mode") in DICT_MODES else "contains"),
                "fallback": ("null" if b.get("fallback") == "null" else "original"),
                "entries": entries, "updated": int(time.time()), "by": _user_rec().get("code", "SYS")})
    _save_json("dictionaries.json", dicts)
    return jsonify(ok=True, dictionary=rec)

@app.post("/api/dict/store/delete")
def dict_store_delete():
    did = ((request.get_json(silent=True) or {}).get("id") or "").strip()
    dicts = [d for d in _dicts() if d.get("id") != did]
    _save_json("dictionaries.json", dicts)
    return jsonify(ok=True, count=len(dicts))

@app.post("/api/dict/coverage")
def dict_coverage():
    """Score a dictionary against real source data — the feedback loop for authoring. Body: {dataset,
    field, mode, entries}. Returns rows total/covered (+%) plus what each entry captured (by_value) and
    the biggest still-uncovered raw values (uncovered) so you know what to add next, and raw→derived
    samples. Coverage is computed with a NULL fallback (covered = an entry matched), independent of the
    dictionary's own fallback choice."""
    b = request.get_json(silent=True) or {}
    ds = (b.get("dataset") or "ttb_cola").strip()
    field = (b.get("field") or "").strip()
    if field not in _ds_columns(ds):
        return jsonify(ok=False, error="unknown field for dataset"), 400
    mode = b.get("mode") if b.get("mode") in DICT_MODES else "contains"
    entries = b.get("entries") or []
    try:
        derived = derive.dict_case(derive.col(field), entries, mode, "null")
    except Exception as e:
        return jsonify(ok=False, error="bad entries: %s" % str(e)[:120]), 200
    nonempty = 'CAST("%s" AS VARCHAR)<>\'\'' % field
    try:
        tot = _cola_q(ds, "SELECT count(*) c FROM t WHERE %s" % nonempty)[0]["c"]
        cov = _cola_q(ds, "SELECT count(*) c FROM t WHERE %s AND (%s) IS NOT NULL" % (nonempty, derived))[0]["c"]
        by_value = _cola_q(ds, "SELECT (%s) v, count(*) n FROM t WHERE %s AND (%s) IS NOT NULL "
                           "GROUP BY 1 ORDER BY n DESC LIMIT 40" % (derived, nonempty, derived))
        uncovered = _cola_q(ds, 'SELECT CAST("%s" AS VARCHAR) v, count(*) n FROM t WHERE %s AND (%s) IS NULL '
                            "GROUP BY 1 ORDER BY n DESC LIMIT 30" % (field, nonempty, derived))
        samples = _cola_q(ds, 'SELECT CAST("%s" AS VARCHAR) raw, (%s) derived FROM t WHERE %s LIMIT 12'
                          % (field, derived, nonempty))
    except Exception as e:
        return jsonify(ok=False, error=str(e)[:160]), 200
    return jsonify(ok=True, dataset=ds, field=field, mode=mode, total=tot, covered=cov,
                   pct=round(100 * cov / max(1, tot), 1), by_value=by_value,
                   uncovered=uncovered, samples=samples)

def _resolve_dict_refs(maps):
    """Expand any field-mapping row that references a saved dictionary ({mode:'dict', dict:<id>}) into
    inline entries the compiler understands, so the master build applies it. A row that already carries
    inline dict_entries is left as-is. Non-destructive (returns a resolved copy)."""
    if not isinstance(maps, dict):
        return maps
    out = {}
    for ds, rules in maps.items():
        rr = []
        for r in (rules or []):
            if isinstance(r, dict) and r.get("mode") == "dict" and r.get("dict") and not r.get("dict_entries"):
                d = _dict_by_id(r.get("dict"))
                if d:
                    r = dict(r, dict_entries=d.get("entries") or [], dict_mode=d.get("mode") or "contains",
                             dict_fallback=d.get("fallback") or "original")
            rr.append(r)
        out[ds] = rr
    return out

# ── Walmart bev-alc tracker — Bright Data walmart_product → warehouse (walmart_products / walmart_runs).
# The tracking surface reads live counts + data-quality concerns computed over the current snapshot. ──
_WM_CONCERNS = [
    ("high", "not_alcohol",  "is_alcohol = false",
     "not classified under Alcohol/Spirits/Wine/Beer — possible non-beverage contamination"),
    ("high", "missing_price", "price IS NULL",
     "no price captured (often an availability/parse failure)"),
    ("med", "missing_abv",  "abv IS NULL",              "ABV missing from Walmart specs"),
    ("med", "missing_size", "size_ml IS NULL",          "size (mL) not parseable"),
    ("med", "missing_upc",  "CAST(upc AS VARCHAR)=''",  "no UPC / GTIN on the listing"),
    ("med", "out_of_stock", "in_stock = false",         "out of stock at the resolved store"),
    ("low", "on_promo",     "on_promo = true",          "final price below list (promo)"),
]

# ── /api/tax/* — beverage-alcohol tax reference (tax_rates.py / tax_revenue.py / landed_cost.py).
# rates   = per-unit schedule (federal CBMA, encoded + state excise seed, effective-dated)
# revenue = Census STC collections (T10 alc sales tax, T20 alc license) per state
# landed  = the itemized cost stack ordering/keystone calls; also the pre-tax normalizer for the price signal
@app.get("/api/tax/rates")
def tax_rates_ep():
    """Effective-dated bev-alc tax RATE schedule. ?jurisdiction=TX&class=spirits&type=excise&level=state.
    Serves the assembled seed (federal + state CSV) if the warehouse table isn't landed yet."""
    import tax_rates
    j = request.args.get("jurisdiction") or None; c = request.args.get("class") or None
    t = request.args.get("type") or None; lvl = request.args.get("level") or None
    try:
        rows = tax_rates.query(jurisdiction=j, beverage_class=c, rate_type=t, level=lvl)
        return jsonify(ok=True, landed=True, count=len(rows), rows=rows)
    except Exception:
        rows, _ = tax_rates.assemble(log=lambda *a: None)
        rows = [r for r in rows if (not j or r["jurisdiction"] == j) and (not c or r["beverage_class"] == c)
                and (not t or r["rate_type"] == t) and (not lvl or r["jurisdiction_level"] == lvl)]
        return jsonify(ok=True, landed=False, count=len(rows), rows=rows, note="assembled from seed (tax_rates not landed)")

@app.get("/api/tax/revenue")
def tax_revenue_ep():
    """Bev-alc tax REVENUE (collections). ?source=census_stc&jurisdiction=TX&kind=sales&year=2021."""
    import tax_revenue
    try:
        yr = int(request.args["year"]) if request.args.get("year") else None
        rows = tax_revenue.query(source=request.args.get("source") or None,
                                 jurisdiction=request.args.get("jurisdiction") or None,
                                 tax_kind=request.args.get("kind") or None, year=yr)
        return jsonify(ok=True, landed=True, count=len(rows), rows=rows)
    except Exception:
        return jsonify(ok=True, landed=False, count=0, rows=[],
                       note="tax_revenue not landed yet (needs a CENSUS_API_KEY run)")

@app.get("/api/tax/landed")
def tax_landed_ep():
    """Itemized landed-cost / tax stack for one item. ?base=12&state=TX&class=spirits&abv=40&size_ml=750&sales=0.
    The pre-tax basis (landed_cost - base) is what price_signal.py subtracts to compare shelf prices across states."""
    import landed_cost
    try:
        base = float(request.args.get("base", "0") or 0)
        size_ml = float(request.args.get("size_ml", "750") or 750)
        abv = request.args.get("abv"); abv = float(abv) if abv not in (None, "") else None
    except ValueError as e:
        return jsonify(ok=False, error="bad numeric param: %s" % e), 400
    state = (request.args.get("state") or "").upper()
    if not state:
        return jsonify(ok=False, error="state required (2-letter)"), 400
    r = landed_cost.landed_cost(base, state, request.args.get("class") or "spirits", abv=abv, size_ml=size_ml,
                                include_sales_tax=(request.args.get("sales") in ("1", "true", "yes")))
    return jsonify(ok=True, **r)

@app.get("/api/walmart/products")
def walmart_products_ep():
    import warehouse
    try:
        rows = warehouse.query("walmart_products",
            "SELECT product_name, brand, category, is_alcohol, size_ml, abv, "
            "CAST(upc AS VARCHAR) upc, price, list_price, on_promo, availability, in_stock, "
            "store, store_loc, rating, reviews, url FROM t ORDER BY brand, product_name")
        return jsonify(ok=True, landed=True, count=len(rows), products=rows)
    except Exception as e:
        return jsonify(ok=True, landed=False, error=str(e)[:140], products=[]), 200

@app.get("/api/walmart/runs")
def walmart_runs_ep():
    """Run history + data-quality concerns + field fill-rates over the current snapshot."""
    import warehouse
    try:
        total = warehouse.query("walmart_products", "SELECT count(*) c FROM t")[0]["c"]
    except Exception as e:
        return jsonify(ok=True, landed=False, error=str(e)[:140], runs=[], concerns=[], total=0), 200
    concerns = []
    for sev, key, cond, msg in _WM_CONCERNS:
        try:
            n = warehouse.query("walmart_products", "SELECT count(*) c FROM t WHERE %s" % cond)[0]["c"]
        except Exception:
            n = 0
        if n:
            concerns.append(dict(severity=sev, key=key, count=n, message=msg))
    try:
        d = warehouse.query("walmart_products",
            "SELECT count(*) c FROM (SELECT upc FROM t WHERE CAST(upc AS VARCHAR)<>'' "
            "GROUP BY upc HAVING count(*)>1)")[0]["c"]
        if d:
            concerns.append(dict(severity="med", key="duplicate_upc", count=d,
                                 message="same UPC on more than one product"))
    except Exception:
        pass
    fills = {}
    for f in ("brand", "category", "size_ml", "abv", "upc", "price"):
        try:
            filled = warehouse.query("walmart_products",
                "SELECT count(*) c FROM t WHERE %s IS NOT NULL AND CAST(%s AS VARCHAR)<>''" % (f, f))[0]["c"]
            fills[f] = round(100 * filled / max(1, total), 1)
        except Exception:
            fills[f] = None
    runs = []
    try:
        runs = warehouse.query("walmart_runs", 'SELECT * FROM t ORDER BY "at" DESC')   # "at" is a DuckDB reserved word
    except Exception:
        pass
    return jsonify(ok=True, landed=True, total=total, runs=runs, concerns=concerns, fills=fills,
                   sev_order={"high": 0, "med": 1, "low": 2})

def _walmart_concern_count():
    import warehouse
    try:
        warehouse.query("walmart_products", "SELECT 1 FROM t LIMIT 1")
    except Exception:
        return 0
    n = 0
    for sev, key, cond, msg in _WM_CONCERNS:
        try:
            n += warehouse.query("walmart_products", "SELECT count(*) c FROM t WHERE %s" % cond)[0]["c"]
        except Exception:
            pass
    return n

# ── Scrape/crawl tracker — every active pull that lands in the warehouse, with freshness + concerns. ──
_SCRAPES = [
    {"id": "ttb-crawl",  "name": "TTB COLA · crawl (index)",  "kind": "crawl",  "source": "ttbonline.gov",
     "table": "ttb_cola_index",  "feeds": "product master", "detail": None},
    {"id": "ttb-detail", "name": "TTB COLA · label detail",   "kind": "detail", "source": "ttbonline.gov",
     "table": "ttb_cola_detail", "feeds": "product master", "detail": None},
    {"id": "ttb-labels", "name": "TTB COLA · label OCR (ABV/UPC)", "kind": "enrich", "source": "ttbonline.gov (label images)",
     "table": "ttb_cola_labels", "feeds": "ABV · net contents · UPC · claims", "detail": None},
    {"id": "walmart",    "name": "Walmart · bev-alc products", "kind": "product", "source": "walmart.com (Bright Data)",
     "table": "walmart_products", "feeds": "pricing + inventory", "detail": "walmart"},
    {"id": "kroger",     "name": "Kroger · store-level (API)",  "kind": "product", "source": "developer.kroger.com",
     "table": "kroger_products", "feeds": "pricing + inventory", "detail": None},
    {"id": "target",     "name": "Target · store-level",        "kind": "product", "source": "target.com (RedSky)",
     "table": "target_products", "feeds": "pricing + inventory", "detail": None},
    {"id": "total-wine", "name": "Total Wine · store-level (getProduct API)", "kind": "product",
     "source": "totalwine.com (BD Browser)", "table": "total_wine_products", "feeds": "pricing + inventory", "detail": None},
    {"id": "abc",        "name": "ABC Fine Wine · store in/out", "kind": "product", "source": "abcfws.com",
     "table": "abc_products", "feeds": "pricing + inventory", "detail": None},
    {"id": "specs",      "name": "Spec's · full PDP + store in/out", "kind": "product", "source": "specsonline.com",
     "table": "specs_products", "feeds": "product catalog + pricing + inventory",
     "detail": "type, ABV, region/origin, varietal, tasting notes, brand, UPC (from image), description"},
    {"id": "binnys",     "name": "Binny's · store qty",         "kind": "product", "source": "binnys.com",
     "table": "binnys_products", "feeds": "pricing + inventory", "detail": None},
    {"id": "shopify",    "name": "Shopify DTC brands",          "kind": "product", "source": "shopify sites",
     "table": "shopify_products", "feeds": "pricing + inventory", "detail": None},
    {"id": "publix",     "name": "Publix · weekly ad (BOGO)",   "kind": "promo",   "source": "publix.com (FL/GA/SC)",
     "table": "publix_products", "feeds": "pricing + promotions", "detail": None},
    # NOTE: the bev-alc chains registry is a source *catalog*, not a scrape — it lives on its own
    # Bev-Alc Chains page, not here.
]

@app.get("/api/scrapes")
def scrapes_ep():
    """Every tracked scrape/crawl: rows, freshness (last warehouse landing), status, concerns."""
    import warehouse, time
    ds = {d["name"]: d for d in warehouse.list_datasets()}
    now = time.time(); out = []
    for s in _SCRAPES:
        d = ds.get(s["table"], {})
        rows = d.get("rows", 0); mod = d.get("modified")
        fresh = (now - mod) if mod else None
        status = "active" if (fresh is not None and fresh < 1800) else ("idle" if rows else "empty")
        out.append(dict(s, rows=rows, modified=mod,
                        fresh_secs=(int(fresh) if fresh is not None else None), status=status,
                        concerns=(_walmart_concern_count() if s["id"] == "walmart" else 0)))
    return jsonify(ok=True, now=int(now), scrapes=out)

# ── Retail Intelligence — the per-store chain catalogs (DoorDash pulls) as an analyst surface.
# Three lenses over <chain>_products[_full] + doordash_stores: chains overview, chain assortment,
# and cross-chain product/brand price comparison. apps/retail.html renders these.
_RETAIL_CHAINS = {"totalwine": "Total Wine & More", "albertsons": "Albertsons",
                  "cvs": "CVS", "circlek": "Circle K"}


def _retail_dataset(chain):
    import warehouse
    for name in (chain + "_products_full", chain + "_products"):
        try:
            if warehouse.query(name, "SELECT 1 FROM t LIMIT 1"):
                return name
        except Exception:
            pass
    return None


@app.get("/api/retail/chains")
def retail_chains_ep():
    import warehouse
    out = []
    for ch, label in _RETAIL_CHAINS.items():
        ds = _retail_dataset(ch)
        if not ds:
            continue
        try:
            r = warehouse.query(ds, "SELECT count(DISTINCT name) skus, count(DISTINCT store) stores, "
                                    "min(price_value) pmin, median(price_value) pmed, max(price_value) pmax "
                                    "FROM t WHERE price_value IS NOT NULL")[0]
        except Exception:
            r = {}
        out.append(dict(chain=ch, name=label, dataset=ds, skus=r.get("skus") or 0, stores=r.get("stores") or 0,
                        price_min=r.get("pmin"), price_median=r.get("pmed"), price_max=r.get("pmax")))
    return jsonify(ok=True, chains=out)


@app.get("/api/retail/chain/<chain>")
def retail_chain_ep(chain):
    import warehouse
    ds = _retail_dataset(chain)
    if not ds:
        return jsonify(ok=False, error="no data for %s" % chain), 404
    q = (request.args.get("q") or "").strip().lower()
    lim = min(int(request.args.get("limit", "600")), 3000)
    where, params = "WHERE price_value IS NOT NULL", []
    if q:
        where += " AND lower(name) LIKE ?"; params.append("%" + q + "%")
    rows = warehouse.query(ds, "SELECT DISTINCT name, price_value, container, unit_size, size_uom, "
                               "pack_count, image_url, is_hemp FROM t %s ORDER BY price_value DESC LIMIT %d"
                               % (where, lim), params)
    return jsonify(ok=True, chain=chain, name=_RETAIL_CHAINS.get(chain, chain), count=len(rows), products=rows)


@app.get("/api/retail/search")
def retail_search_ep():
    """Cross-chain price comparison for a product/brand query — one row per (chain, product) with price range."""
    import warehouse
    q = (request.args.get("q") or "").strip().lower()
    if len(q) < 2:
        return jsonify(ok=True, q=q, results=[])
    out = []
    for ch, label in _RETAIL_CHAINS.items():
        ds = _retail_dataset(ch)
        if not ds:
            continue
        try:
            rows = warehouse.query(ds, "SELECT name, min(price_value) pmin, max(price_value) pmax, "
                                       "count(DISTINCT store) stores, any_value(image_url) img FROM t "
                                       "WHERE price_value IS NOT NULL AND lower(name) LIKE ? "
                                       "GROUP BY name ORDER BY pmin LIMIT 50", ["%" + q + "%"])
        except Exception:
            rows = []
        for r in rows:
            out.append(dict(r, chain=ch, chain_name=label))
    out.sort(key=lambda r: (r.get("pmin") if r.get("pmin") is not None else 1e9))
    return jsonify(ok=True, q=q, results=out[:150])


@app.get("/api/retail/stores")
def retail_stores_ep():
    import warehouse
    try:
        rows = warehouse.query("doordash_stores", "SELECT store_id, chain, store_name, city, state, lat, lon FROM t")
    except Exception:
        rows = []
    return jsonify(ok=True, count=len(rows), stores=rows)


# ── Market Intelligence: the geographic OUTLET landscape of a metro (chains + independents, retail +
# on-premise), how complete our sweep is vs Google, and the on-premise (NAOP) beverage menus. Distinct from
# Retail Intelligence (product/assortment across a few chains) — this is the market's whole outlet universe.
def _market_merchants(market):
    import warehouse
    try:
        return warehouse.query(market + "_merchants", "SELECT * FROM t")
    except Exception:
        return []


def _market_coverage(market):
    import warehouse
    try:
        c = warehouse.query(market + "_coverage", "SELECT * FROM t ORDER BY run_id DESC LIMIT 1")
        return c[0] if c else None
    except Exception:
        return None


@app.get("/api/market/list")
def market_list_ep():
    import warehouse
    try:
        ms = sorted({d["name"][:-len("_merchants")] for d in warehouse.list_datasets()
                     if d["name"].endswith("_merchants")})
    except Exception:
        ms = []
    return jsonify(ok=True, markets=ms or ["orlando"])


@app.get("/api/market/summary")
def market_summary_ep():
    from collections import Counter
    market = (request.args.get("market") or "orlando").strip().lower()
    ms = _market_merchants(market)

    def cnt(f):
        return sum(1 for m in ms if f(m))
    cui = Counter((m.get("cuisine") or "—") for m in ms if m.get("type") == "restaurant")
    return jsonify(ok=True, market=market, total=len(ms),
                   retail=cnt(lambda m: m.get("type") == "retail"),
                   restaurant=cnt(lambda m: m.get("type") == "restaurant"),
                   chain=cnt(lambda m: m.get("is_chain")),
                   independent=cnt(lambda m: not m.get("is_chain")),
                   google_confirmed=cnt(lambda m: m.get("google_confirmed")),
                   cuisines=[{"cuisine": k, "n": v} for k, v in cui.most_common(14)],
                   coverage=_market_coverage(market))


@app.get("/api/market/merchants")
def market_merchants_ep():
    market = (request.args.get("market") or "orlando").strip().lower()
    typ = (request.args.get("type") or "").strip()          # retail | restaurant
    kind = (request.args.get("kind") or "").strip()         # chain | independent
    q = (request.args.get("q") or "").strip().lower()
    ms = _market_merchants(market)

    def keep(m):
        if typ and m.get("type") != typ:
            return False
        if kind == "chain" and not m.get("is_chain"):
            return False
        if kind == "independent" and m.get("is_chain"):
            return False
        if q and q not in (m.get("name") or "").lower() and q not in (m.get("cuisine") or "").lower():
            return False
        return True
    rows = [m for m in ms if keep(m)]
    rows.sort(key=lambda m: (m.get("type") or "", 1 if m.get("is_chain") else 0, (m.get("name") or "").lower()))
    return jsonify(ok=True, market=market, count=len(rows), merchants=rows[:1500])


@app.get("/api/market/coverage")
def market_coverage_ep():
    market = (request.args.get("market") or "orlando").strip().lower()
    return jsonify(ok=True, market=market, coverage=_market_coverage(market))


@app.get("/api/market/onpremise")
def market_onpremise_ep():
    import warehouse
    q = (request.args.get("q") or "").strip().lower()
    acct = (request.args.get("account") or "").strip()
    try:
        accts = warehouse.query("naop_accounts", "SELECT * FROM t WHERE serves_alcohol = true")
    except Exception:
        accts = []
    # Fold in FIRST-PARTY menus (menu_site.py: Google-discovered site/PDF menus, read by Claude). Most bars pour
    # DINE-IN only, so DoorDash misses them — these are the real drink lists (price_basis='menu_list', not
    # delivery-inflated). Prefer them; add any venue DoorDash didn't surface.
    try:
        mcounts = warehouse.query("menu_beverages", "SELECT account, count(*) n FROM t GROUP BY account")
    except Exception:
        mcounts = []
    by_name = {(a.get("account") or "").lower(): a for a in accts}
    for m in mcounts:
        nm = m.get("account") or ""
        a = by_name.get(nm.lower())
        if a:
            a["first_party"] = True
            a["n_beverages"] = max(a.get("n_beverages") or 0, m["n"])
        else:
            accts.append({"account": nm, "cuisine": "", "serves_alcohol": True,
                          "n_beverages": m["n"], "first_party": True})
    if q:
        accts = [a for a in accts if q in (a.get("account") or "").lower() or q in (a.get("cuisine") or "").lower()]
    accts.sort(key=lambda a: -(a.get("n_beverages") or 0))
    bevs = []
    if acct:
        cols = "name, description, price, category, base_spirit, beer_style, sub, root, price_basis"
        rows = []
        for tbl, src in (("menu_beverages", "menu"), ("naop_beverages", "doordash")):
            try:
                for r in warehouse.query(tbl, "SELECT %s FROM t WHERE account = ?" % cols, [acct]):
                    r["source"] = src
                    rows.append(r)
            except Exception:
                pass
        seen = {}                                # dedup by name — the first-party dine-in menu wins over delivery
        for r in rows:
            k = (r.get("name") or "").strip().lower()
            if k and (k not in seen or r["source"] == "menu"):
                seen[k] = r

        def _pr(r):
            try:
                return float(r.get("price") or 0)
            except (TypeError, ValueError):
                return 0.0
        bevs = sorted(seen.values(), key=lambda r: (r.get("category") or "", -_pr(r)))
    return jsonify(ok=True, count=len(accts), accounts=accts, account=acct, beverages=bevs)


# ── Master Matching Workbench: the canonical master = TTB items CONFIRMED to exist elsewhere. Drill from a
# master record down to the COLA filings underneath it + the source that corroborated it (see: master_ttb.py).
def _wq(ds, sql, params=None):
    import warehouse
    try:
        return warehouse.query(ds, sql, params)
    except Exception:
        return []


@app.get("/api/master/workbench/summary")
def mwb_summary_ep():
    # Whole-response TTL cache (60s): the funnel numbers are rebuild-stable and human decisions lag <=60s, so
    # serve from RAM and let the warmer keep it hot -> no user eats the per-request Tigris round-trips.
    return jsonify(_ttl("wb_summary_resp", 90, _wb_summary_data))


def _wb_summary_data():
    # The workbench reflects the REAL product master (dim_product), not the TTB-COLA tiering slice.
    # ONE GRAIN — every funnel stage counts distinct PRODUCTS so the funnel reconciles: each candidate sits in
    # exactly one process state, and Auto + Claude + Analyst + Sr + Steward = Candidates. The 196,535 raw source
    # rows collapse INTO the 76,618 candidates (shown as source_rows/collapsed, not a stage). Corroboration is
    # the confidence signal: 2+ sources = auto-trusted; single-source = the analyst queue. Confirmed to master =
    # the trusted subset (auto-corroborated + Claude-adjudicated + human-confirmed), NOT every built row.
    # SKU is the matching TARGET (product + size + pack, UPC-identified) — the only grain where price/inventory/
    # distribution compare across sources. Count SKUs; corroboration = the same SKU seen in 2+ sources.
    # Rebuild-stable numbers from the pre-computed wb_summary view (avoids the dim_sku unnest/GROUP BY scan on
    # the cold load); fall back to the live aggregates if the view isn't built yet.
    sv = _wb_view("wb_summary")
    if sv:
        s0 = sv[0]
        total = s0.get("total") or 0
        multi = s0.get("multi") or 0
        source_rows = s0.get("source_rows") or total
        by = [{"s": x.get("source"), "n": x.get("n")} for x in json.loads(s0.get("by_source") or "[]")]
    else:
        agg = _wq_cached("dim_sku", "SELECT count(*) total, "
                         "sum(CASE WHEN sources >= 2 THEN 1 ELSE 0 END) multi FROM t")
        total = (agg[0]["total"] if agg else 0) or 0
        multi = (agg[0]["multi"] if agg else 0) or 0
        stg = _wq_cached("_stage_product", "SELECT count(*) n FROM t")
        source_rows = (stg[0]["n"] if stg and stg[0].get("n") else total)
        by = _wq_cached("dim_sku", "SELECT s, count(*) n FROM (SELECT unnest(source_list) s FROM t) "
                        "GROUP BY 1 ORDER BY 2 DESC")
    single = total - multi                              # single-source → analyst queue
    decisions = _wq("master_decisions", "SELECT tier, action FROM t")
    conf_h = sum(1 for d in decisions if d.get("action") == "confirm")       # human-confirmed single-source
    rej_h = sum(1 for d in decisions if d.get("action") == "reject")
    claude = sum(1 for d in decisions if d.get("tier") == "claude")
    sr = sum(1 for d in decisions if d.get("action") == "escalate" and d.get("tier") == "senior")
    st = sum(1 for d in decisions if d.get("action") == "escalate" and d.get("tier") == "steward")
    analyst = max(0, single - claude - sr - st - conf_h - rej_h)             # single-source still pending
    # process states partition the candidates; Confirmed is the promoted-to-master roll-up (a subset)
    stages = {"candidates": total,
              "auto": multi, "claude": claude, "analyst": analyst, "senior": sr, "steward": st,
              "confirmed": multi + claude + conf_h}
    # product-level tiering (TTB's real grain — it can't form SKUs): products corroborated across sources,
    # and specifically the TTB-registered products ALSO seen in retail (Tier-1 promotions).
    s0 = sv[0] if sv else {}
    product_tier = {"products": s0.get("prod_total") or 0, "corroborated": s0.get("prod_corr") or 0,
                    "ttb_corroborated": s0.get("ttb_corr") or 0, "ttb_registered": s0.get("ttb_total") or 0}
    return dict(ok=True, master=total, review=analyst, quarantine=0,
                source_rows=source_rows, collapsed=max(0, source_rows - total),
                avg_confidence=round((multi * 0.88 + single * 0.5) / max(1, total), 3),
                stages=stages, product_tier=product_tier,
                by_source=[{"source": r.get("s"), "n": r.get("n")} for r in by])


def _wb_brandmap():
    return {b["brand_key"]: (b.get("brand") or "") for b in _wq("dim_brand", "SELECT brand_key, brand FROM t")}


def _wb_productmap():
    """product_key -> (product_name, brand_key) — to name an item (product + size) and find same-brand items."""
    return {p["product_key"]: (p.get("product_name"), p.get("brand_key"))
            for p in _wq("dim_product", "SELECT product_key, product_name, brand_key FROM t")}


def _wb_size_label(size_ml):
    if not size_ml:
        return ""
    try:
        v = float(size_ml)
    except Exception:
        return ""
    return ("%gL" % (v / 1000.0)) if v >= 1000 else ("%gml" % v)


def _wb_item_stub(it, prodmap, brands):
    """Reshape a dim_item row into the fields the workbench frontend consumes. The ITEM is the key grain:
    product + size. Name = product name + size; brand + corroboration resolve via product_key. Corroboration
    (`sources`) is the confidence signal: 2+ independent sources = trusted, 1 = needs review."""
    pname, bkey = prodmap.get(it.get("product_key"), (None, None))
    src = it.get("source_list") or []
    if isinstance(src, str):
        src = [src]
    n = it.get("sources") or len(src) or 1
    conf = 0.9 if n >= 3 else (0.82 if n == 2 else 0.5)
    sz = _wb_size_label(it.get("size_ml"))
    name = " · ".join([x for x in [(pname or "").strip(), sz, (it.get("container") or "")] if x]) or "(unnamed item)"
    return {"cluster_id": it.get("item_key"), "candidate_name": name, "product_key": it.get("product_key"),
            "brand": brands.get(bkey) or "", "class_type": None,
            "confidence": conf, "corroborated_by": ", ".join(src), "sources": n,
            "match_kind": ("multi-source" if n >= 2 else "single-source"),
            "member_count": it.get("source_rows") or 1, "tier": (1 if n >= 2 else 0),
            "upc": None, "size_ml": it.get("size_ml")}


def _wb_itemmap():
    """item_key -> (product_key, size_ml, container) — to name a SKU (product + size + pack) via its item."""
    return {i["item_key"]: (i.get("product_key"), i.get("size_ml"), i.get("container"))
            for i in _wq("dim_item", "SELECT item_key, product_key, size_ml, container FROM t")}


# The workbench's dominant cost was re-reading big master tables from S3 on EVERY request — the full-table
# name maps (dim_product ~830k + dim_item ~871k + dim_brand) AND re-scanning dim_sku (~874k) per query. All of
# it only changes on a master REBUILD, so cache it, invalidated by a cheap row-count "version" (count = parquet
# metadata, no full scan) revalidated at most every 30s. First request after a rebuild pays the load; the rest
# are instant. `_wq_cached` memoizes read-query RESULTS the same way (safe — workbench reads are rebuild-stable).
_WB = {"ts": 0.0, "ver": None, "maps": None, "mupc": None, "detail": None, "pitems": None, "q": {}}


def _wb_ver():
    now = time.time()
    if _WB["ver"] is not None and (now - _WB["ts"]) < 30:
        return _WB["ver"]
    try:
        ver = tuple((_wq(t, "SELECT count(*) c FROM t") or [{}])[0].get("c")
                    for t in ("dim_brand", "dim_product", "dim_item", "dim_sku"))
    except Exception:
        ver = _WB["ver"]
    # A transient count-query failure returns None for that table — do NOT let a flaky read look like a rebuild
    # and nuke every workbench cache (which would make the whole console intermittently cold). Only accept a
    # version made of clean, non-None counts; otherwise keep the current one and try again next tick.
    if _WB["ver"] is not None and any(c is None for c in ver):
        _WB["ts"] = now
        return _WB["ver"]
    if ver != _WB["ver"]:                 # a rebuild changed the tables → drop all caches
        _WB["maps"] = None
        _WB["mupc"] = None
        _WB["detail"] = None
        _WB["pitems"] = None
        _WB["q"] = {}
        _WB["ver"] = ver
    _WB["ts"] = now
    return _WB["ver"]


def _wb_maps():
    _wb_ver()
    if _WB["maps"] is None:
        _WB["maps"] = (_wb_brandmap(), _wb_productmap(), _wb_itemmap())
    return _WB["maps"]


def _wb_prod_items():
    """product_key -> [item_key] reverse index, so 'other SKUs of the same product' is an O(1) dict lookup
    instead of a full 1M-row Python scan of itemmap on EVERY candidates request."""
    _wb_ver()
    if _WB["pitems"] is None:
        _, _, itemmap = _wb_maps()
        idx = {}
        for ik, (pk, _sz, _c) in itemmap.items():
            if pk:
                idx.setdefault(pk, []).append(ik)
        _WB["pitems"] = idx
    return _WB["pitems"]


def _wb_detail():
    """Index the warmed wb_queue + wb_master view rows by sku_key. Those views are pre-joined AND now carry the
    detail-modal columns (image / varietal / region / abv / pack / gtin …), so the item-detail modal serves
    straight from RAM with NO dim_sku / dim_product scan (each open used to full-scan the 1M-row tables, ~26s
    cold). Tiny (~8k rows) → instant warm, negligible memory. A non-queue cid misses → per-cid scan fallback."""
    _wb_ver()
    if _WB["detail"] is None:
        idx = {}
        for v in ("wb_queue", "wb_master"):
            for r in (_wb_view(v) or []):
                k = r.get("sku_key")
                if k is not None:
                    idx[str(k)] = r
        _WB["detail"] = idx
    return _WB["detail"]


def _wb_merge_upc_attrs():
    """upc -> per-member attributes (image/name/size/abv/…) for the merge-review clusters, built ONCE and
    cached until rebuild. Opening a cluster used to run a fresh _stage_product scan per click (~1-3s each);
    but the member UPCs across ALL merge groups are a small bounded set (~thousands), so we scan them once
    up front and every cluster-open becomes an instant dict lookup."""
    _wb_ver()
    if _WB["mupc"] is None:
        view = _wb_view("wb_merges") or []
        upcs = set()
        for g in view:
            try:
                for m in json.loads(g.get("members") or "[]"):
                    if m.get("upc"):
                        upcs.add(str(m["upc"]))
            except Exception:
                pass
        upcs = sorted(upcs)
        attrs = {}
        for i in range(0, len(upcs), 2000):                 # chunk the IN(...) list so no single query is huge
            chunk = upcs[i:i + 2000]
            ph = ",".join("?" * len(chunk))
            try:
                for r in _wq("_stage_product",
                             "SELECT CAST(upc AS VARCHAR) u, any_value(image) img, any_value(product_name) nm, "
                             "any_value(size_ml) size_ml, any_value(abv) abv, any_value(container) container, "
                             "any_value(pack) pack, any_value(origin) origin, any_value(varietal) varietal, "
                             "any_value(_source_id) source_id, list(DISTINCT _source) srcs "
                             "FROM t WHERE CAST(upc AS VARCHAR) IN (%s) GROUP BY 1" % ph, chunk):
                    attrs[r.get("u")] = r
            except Exception:
                pass
        _WB["mupc"] = attrs
    return _WB["mupc"]


def _wq_cached(ds, sql, params=None):
    """A workbench read query, memoized until the next master rebuild. Turns a per-request S3 parquet scan
    into a one-time cost. Use ONLY for read-only, rebuild-stable queries (never for decisions/writes)."""
    _wb_ver()
    key = (ds, sql, tuple(params or []))
    if key not in _WB["q"]:
        _WB["q"][key] = _wq(ds, sql, params)
    return _WB["q"][key]


def _wb_view(name):
    """Read a pre-computed workbench view (wb_master / wb_queue / wb_summary from wb_views.py), cached until the
    next rebuild. Returns None if the table isn't present yet -> callers fall back to the live computation."""
    try:
        return _wq_cached(name, "SELECT * FROM t")
    except Exception:
        return None


def _wb_sku_stub(sk, itemmap, prodmap, brands):
    """Reshape a dim_sku row into the fields the workbench consumes. The SKU is the MATCHING TARGET — the
    fully-specified sellable unit (product + size + pack), identified by UPC where present. Name + brand resolve
    up through item_key -> product_key. Corroboration (`sources`) = the same SKU seen in N sources."""
    pk, size_ml, container = itemmap.get(sk.get("item_key"), (None, None, None))
    pname, bkey = prodmap.get(pk, (None, None))
    src = sk.get("source_list") or []
    if isinstance(src, str):
        src = [src]
    n = sk.get("sources") or len(src) or 1
    conf = 0.9 if n >= 3 else (0.82 if n == 2 else 0.5)
    pack = sk.get("pack")
    parts = [(pname or "").strip(), _wb_size_label(size_ml), (container or ""),
             ("%s-pk" % pack if pack else "")]
    name = " · ".join([x for x in parts if x]) or "(unnamed sku)"
    return {"cluster_id": sk.get("sku_key"), "candidate_name": name, "item_key": sk.get("item_key"),
            "product_key": pk, "brand": brands.get(bkey) or "", "class_type": None,
            "confidence": conf, "corroborated_by": ", ".join(src), "sources": n,
            "match_kind": ("multi-source" if n >= 2 else "single-source"),
            "member_count": sk.get("source_rows") or 1, "tier": (1 if n >= 2 else 0),
            "upc": (sk.get("upc") or None), "size_ml": size_ml}


def _wb_view_stub(r):
    """Same frontend shape as _wb_sku_stub, but from a pre-joined wb_master/wb_queue row (brand + product_name
    already resolved) — so the list endpoints need no name maps or dim_sku scan at request time."""
    src = r.get("source_list") or []
    if isinstance(src, str):
        src = [src]
    n = r.get("sources") or len(src) or 1
    conf = 0.9 if n >= 3 else (0.82 if n == 2 else 0.5)
    sz = _wb_size_label(r.get("size_ml"))
    name = " · ".join([x for x in [(r.get("product_name") or "").strip(), sz, (r.get("container") or "")] if x]) \
        or "(unnamed item)"
    return {"cluster_id": r.get("sku_key"), "candidate_name": name, "product_key": r.get("product_key"),
            "brand": r.get("brand") or "", "class_type": None, "confidence": conf,
            "corroborated_by": ", ".join(src), "sources": n,
            "match_kind": ("multi-source" if n >= 2 else "single-source"),
            "member_count": r.get("source_rows") or 1, "tier": (1 if n >= 2 else 0),
            "upc": (r.get("upc") or None), "size_ml": r.get("size_ml")}


_JOIN_SERVER = ("SELECT s.sku_key, s.item_key, i.product_key, b.brand, p.product_name, i.size_ml, i.container, "
                "s.pack, s.upc, s.sources, s.source_list, s.source_rows "
                "FROM read_parquet('%s') s LEFT JOIN read_parquet('%s') i ON s.item_key = i.item_key "
                "LEFT JOIN read_parquet('%s') p ON i.product_key = p.product_key "
                "LEFT JOIN read_parquet('%s') b ON p.brand_key = b.brand_key ")


def _wb_search(q, single_source, limit=400):
    """Search the FULL master (dim_sku join, ~1M) by name / brand / source — so a query finds items that sort
    past the capped wb_ views (e.g. a low-volume Total Wine SKU). single_source True → analyst queue else master."""
    import warehouse
    con = warehouse.connect()
    p = {t: warehouse.uri(t).strip("'") for t in ("dim_sku", "dim_item", "dim_product", "dim_brand")}
    like = "%" + q.replace(" ", "%") + "%"
    scond = "s.sources = 1" if single_source else "s.sources >= 2"
    sql = (_JOIN_SERVER % (p["dim_sku"], p["dim_item"], p["dim_product"], p["dim_brand"])
           + "WHERE %s AND (lower(p.product_name) LIKE ? OR lower(b.brand) LIKE ? OR "
             "replace(lower(array_to_string(s.source_list, ' ')), '_', ' ') LIKE ?) "
             "ORDER BY s.source_rows DESC LIMIT %d" % (scond, limit))
    try:
        cur = con.execute(sql, [like, like, like])
        cols = [d[0] for d in cur.description]
        return [dict(zip(cols, row)) for row in cur.fetchall()]
    except Exception as e:
        app.logger.warning("wb_search: %s", e)
        return []


@app.get("/api/master/workbench/master")
def mwb_master_ep():
    q = (request.args.get("q") or "").strip().lower()
    src = (request.args.get("source") or "").strip()
    if not q and not src:               # the unfiltered master is the landing view — TTL-cache + warm it
        return jsonify(_ttl("wb_master_resp", 90, lambda: _wb_master_data("", "")))
    return jsonify(_wb_master_data(q, src))


def _wb_master_data(q, src):
    # the trustworthy master = SKUs corroborated by 2+ independent sources (the same SKU seen across sources).
    # Prefer the pre-computed wb_master view (tiny, names pre-joined); fall back to the live scan if not built yet.
    view = _wb_view("wb_master")
    if q and view is not None:              # search the FULL master (past the cap), not just the loaded view
        rows, stub = _wb_search(q, single_source=False, limit=5000), _wb_view_stub
        if src:
            rows = [r for r in rows if src in (r.get("source_list") or [])]
    elif view is not None:
        rows, stub = view[:5000], _wb_view_stub
    else:
        brands, prodmap, itemmap = _wb_maps()
        rows = _wq_cached("dim_sku", "SELECT sku_key, item_key, pack, upc, sources, source_list, "
                          "source_rows FROM t WHERE sources >= 2 ORDER BY sources DESC, source_rows DESC LIMIT 600")
        stub = lambda r: _wb_sku_stub(r, itemmap, prodmap, brands)
    items = []
    for r in rows:
        st = stub(r)
        st["matched_by"] = "auto"
        if src and src not in (r.get("source_list") or []):
            continue
        if q and q not in (st["candidate_name"] or "").lower() and q not in (st["brand"] or "").lower() \
                and q not in (st["upc"] or ""):
            continue
        items.append(st)
    return dict(ok=True, count=len(items), items=items)


def _sku_full(cid):
    """Everything attached to a SKU — the sku stub + its product-grain attributes (image/geo/varietal/abv) +
    its vintages (dim_vintage). Shared by the item detail + the expand-modal + the vision run."""
    row = _wb_detail().get(str(cid))                     # warmed, widened view row for queue/master items
    if row is not None:                                  # the hot path — everything from RAM, no dim scans
        st = _wb_view_stub(row)
        st["image"] = row.get("image") or ""
        for f in ("origin", "bottled_in", "region", "sub_region", "appellation", "varietal", "category", "abv", "style"):
            st[f] = row.get(f) or ""
        st["pack"] = row.get("pack") or ""
        st["gtin"] = row.get("gtin") or ""
    else:                                                # not a queue/master item → per-cid scan fallback
        got = _wq_cached("dim_sku", "SELECT * FROM t WHERE sku_key = ?", [cid])
        if not got:
            return None
        sk = got[0]
        brands, prodmap, itemmap = _wb_maps()
        st = _wb_sku_stub(sk, itemmap, prodmap, brands)
        pk = (itemmap.get(sk.get("item_key")) or (None, None, None))[0]
        attrs = (_wq_cached("dim_product", "SELECT * FROM t WHERE product_key = ?", [pk]) or [{}])[0] if pk else {}
        st["image"] = attrs.get("image") or ""
        for f in ("origin", "bottled_in", "region", "sub_region", "appellation", "varietal", "category", "abv", "style"):
            st[f] = attrs.get(f) or st.get(f) or ""
        st["pack"] = sk.get("pack") or ""
        st["gtin"] = sk.get("gtin") or ""
    st["vintages"] = [v["vintage"] for v in _wq_cached("dim_vintage",   # small table; cached per cid
                      "SELECT DISTINCT vintage FROM t WHERE sku_key = ? ORDER BY vintage", [cid])]
    return st


_COLA_URL = "https://www.ttbonline.gov/colasonline/viewColaDetails.do?action=publicDisplaySearchBasic&ttbid=%s"


def _item_cola_filings(cid):
    """The real TTB COLA label filings that collapsed into this master SKU: the ttb_ids (via the source xwalk,
    where TTB is source 'ttb') joined to ttb_cola_detail — the DETAIL table keyed by the same 14-digit ttb_id (the
    ttb_cola search table uses a different 9-digit id, so it doesn't join) — for fanciful / brand / class-type /
    status / approval-date / varietal. Each row links to its public COLA detail page. Cached per item."""
    try:
        ids = [str(r["product_id"]) for r in _wq_cached(
            "xwalk_source_sku", "SELECT product_id FROM t WHERE source='ttb' AND sku_key=? LIMIT 60", [cid])]
    except Exception:
        ids = []
    if not ids:
        return []
    ph = ",".join("?" * len(ids))
    try:
        rows = _wq_cached("ttb_cola_detail",
                          "SELECT CAST(ttb_id AS VARCHAR) ttb_id, fanciful_name fanciful, brand_name applicant, "
                          "status, approval_date completed, class_type_desc class_type, grape_varietal varietal, "
                          "alcohol_content abv FROM t WHERE CAST(ttb_id AS VARCHAR) IN (%s)" % ph, ids)
    except Exception:
        rows = []
    seen = {str(r.get("ttb_id")): r for r in rows}       # fill in ttb_ids the detail table didn't have, link still works
    out = []
    for i in ids:
        r = dict(seen.get(i, {"ttb_id": i}))
        r["ttb_id"] = i
        r["cola_url"] = _COLA_URL % i
        out.append(r)
    return out


@app.get("/api/master/workbench/item/<cid>")
def mwb_item_ep(cid):
    st = _sku_full(cid)
    if st is None:
        return jsonify(ok=False, error="not found"), 404
    src = st.get("corroborated_by", "")
    src = [s for s in (src.split(", ") if src else []) if s]
    filings = [{"source": s, "brand": st["brand"], "name": st["candidate_name"], "category": (st["upc"] or "no UPC")}
               for s in src]
    return jsonify(ok=True, item=st, filings=filings, cola_filings=_item_cola_filings(cid))


@app.post("/api/master/workbench/vision/<cid>")
def mwb_vision_ep(cid):
    """On-demand: run label-vision on THIS SKU's image (the button in the queue/modal). Returns the fields Claude
    read off the label so a steward can eyeball + accept them. Only runs one image — cheap, human-triggered."""
    st = _sku_full(cid)
    if st is None:
        return jsonify(ok=False, error="not found"), 404
    img = st.get("image") or ""
    if not img:
        return jsonify(ok=False, error="no image on this SKU to read"), 400
    try:
        import label_vision
        fields = label_vision.extract(img)
    except Exception as e:
        return jsonify(ok=False, error="vision failed: %s" % str(e)[:200]), 500
    return jsonify(ok=True, sku_key=cid, image=img, fields=fields)


def _wb_norm_brand(s):
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def _wb_upc_prefix(u):
    d = re.sub(r"\D", "", u or "")
    return d[-12:][:7] if len(d) >= 11 else ""       # GS1 company-prefix zone (normalize to 12-digit UPC-A)


def _wb_tokens(s):
    return set(re.findall(r"[a-z0-9]+", (s or "").lower()))


@app.get("/api/master/workbench/candidates/<cid>")
def mwb_candidates_ep(cid):
    """The RANKED possible SKUs this one could really be — the disambiguation the workbench is built on. #1 is
    the SKU's own proposed home; the rest are OTHER SKUs of the SAME PRODUCT (across sizes/packs/UPCs) scored by
    name + UPC overlap, so a genuine duplicate SKU (same sellable unit, different source string / stray UPC)
    surfaces. A big gap to #2 (or a high #1) = low ambiguity; a near-tie = a human decides whether to merge."""
    base = _wb_detail().get(str(cid))                    # warmed view row (queue/master); else per-cid scan
    if base is None:
        got = _wq_cached("dim_sku", "SELECT * FROM t WHERE sku_key = ?", [cid])
        if not got:
            return jsonify(ok=False, error="not found"), 404
        base = got[0]
    brands, prodmap, itemmap = _wb_maps()
    bstub = _wb_sku_stub(base, itemmap, prodmap, brands)
    btok = _wb_tokens(bstub["candidate_name"])
    bpk = (itemmap.get(base.get("item_key")) or (None, None, None))[0]
    same_prod_items = _wb_prod_items().get(bpk, [])       # O(1) reverse index, not a 1M-row scan of itemmap
    pool = []
    if same_prod_items:
        iks = same_prod_items[:400]
        ph = ",".join("?" * len(iks))
        pool = _wq_cached("dim_sku", "SELECT sku_key, item_key, pack, upc, sources, source_list "
                   "FROM t WHERE item_key IN (%s) AND sku_key <> ? LIMIT 400" % ph, iks + [cid])
    cands = [{"cluster_id": cid, "name": bstub["candidate_name"], "grain": "sku",
              "table": "dim_sku", "conf": round(bstub["confidence"], 2), "win": True,
              "why": "Proposed master SKU — %s, corroborated by %s (%d source row%s)." % (
                  bstub["match_kind"], bstub["corroborated_by"] or "one source",
                  bstub["member_count"], "" if bstub["member_count"] == 1 else "s"),
              "keys": ([("upc:" + bstub["upc"]) if bstub["upc"] else "no-upc",
                        "brand:" + (bstub["brand"] or "").lower().replace(" ", "-")] +
                       ["src:" + s for s in (bstub["corroborated_by"].split(", ") if bstub["corroborated_by"] else [])[:2]])[:5]}]
    alts = []
    for c in pool:
        cst = _wb_sku_stub(c, itemmap, prodmap, brands)
        ctok = _wb_tokens(cst["candidate_name"])
        nj = (len(btok & ctok) / len(btok | ctok)) if (btok and ctok) else 0.0
        upc_same = bool(bstub["upc"]) and bstub["upc"] == cst["upc"]
        if nj < 0.34 and not upc_same:
            continue
        alts.append({"cluster_id": cst["cluster_id"], "name": cst["candidate_name"], "grain": "sku",
                     "table": "dim_sku", "conf": round(1.0 if upc_same else nj, 2),
                     "why": ("Same UPC — definitively the same SKU, merge." if upc_same else
                             "%d%% name overlap, same product — a potential duplicate SKU." % int(nj * 100)),
                     "keys": ([("upc:" + cst["upc"]) if cst["upc"] else "no-upc", "%d src" % cst["sources"]])})
    alts.sort(key=lambda x: -x["conf"])
    return jsonify(ok=True, cluster_id=cid, candidates=cands + alts[:4])


@app.get("/api/master/workbench/review")
def mwb_review_ep():
    # The default (unfiltered) queue is the landing view — TTL-cache it (90s) and warm it so it's never cold;
    # a search (q=...) runs live against the full queue.
    q = (request.args.get("q") or "").strip().lower()
    if not q:
        return jsonify(_ttl("wb_review_resp", 90, lambda: _wb_review_data("")))
    return jsonify(_wb_review_data(q))


def _wb_review_data(q):
    # analyst queue = single-source SKUs (in only one catalog) — need an independent source to corroborate
    # before they're fully trusted. Highest source_rows first (most-seen singletons first). Prefer wb_queue.
    view = _wb_view("wb_queue")
    if q and view is not None:              # search the FULL queue (past the cap) so "total wine" finds items
        rows, stub = _wb_search(q, single_source=True, limit=5000), _wb_view_stub
    elif view is not None:
        rows, stub = view[:5000], _wb_view_stub
    else:
        brands, prodmap, itemmap = _wb_maps()
        rows = _wq_cached("dim_sku", "SELECT sku_key, item_key, pack, upc, sources, source_list, "
                          "source_rows FROM t WHERE sources = 1 ORDER BY source_rows DESC LIMIT 400")
        stub = lambda r: _wb_sku_stub(r, itemmap, prodmap, brands)
    decs = {d.get("cluster_id"): d for d in _wq("master_decisions", "SELECT cluster_id, action, tier FROM t")}
    terminal = {"confirm", "reject", "merge"}
    out = []
    for r in rows:
        st = stub(r)
        dd = decs.get(st["cluster_id"])
        if dd and dd.get("action") in terminal:               # confirmed/rejected/merged leave the queue
            continue
        st["review_tier"] = dd.get("tier") if (dd and dd.get("action") == "escalate") else None  # analyst→senior→steward
        out.append(st)
    return dict(ok=True, count=len(out), items=out)


@app.get("/api/master/workbench/merges")
def mwb_merges_ep():
    q = (request.args.get("q") or "").strip().lower()
    if not q:                           # the unfiltered merge queue is the landing view — TTL-cache + warm it
        return jsonify(_ttl("wb_merges_resp", 90, lambda: _wb_merges_data("")))
    return jsonify(_wb_merges_data(q))


def _wb_merges_data(q):
    """Ambiguous MERGE groups for the cluster-review page — SKUs that are the same product+size+pack split
    across different/missing UPCs (the 'one unit fractured by UPC noise' duplicates). From the pre-computed
    wb_merges view; groups already decided (merge / keep-separate) drop out. Prioritized by impact (rows)."""
    view = _wb_view("wb_merges")
    if view is None:
        return dict(ok=True, count=0, merges=[], note="wb_merges not built yet — runs on the next master rebuild")
    decided = {d.get("cluster_id") for d in _wq("master_decisions", "SELECT cluster_id, action FROM t")
               if d.get("action") in ("merge", "keep_separate")}
    out = []
    for r in view:
        if r["merge_id"] in decided:
            continue
        if q and q not in (r.get("name") or "").lower() and q not in (r.get("brand") or "").lower():
            continue
        m = dict(r)
        try:
            m["members"] = json.loads(r.get("members") or "[]")
        except Exception:
            m["members"] = []
        out.append(m)
    out.sort(key=lambda x: -(x.get("total_rows") or 0))        # highest-impact duplicates first
    return dict(ok=True, count=len(out), merges=out[:400])


@app.get("/api/src")
def src_feeds_ep():
    """The inbound normalization spine — the 7 src_<grain> feeds every source shreds into, with per-grain
    records / distinct entities / how many are corroborated by 2+ sources (tagged match, no key join). Plus the
    sources that emit each grain. Powers the estate's clean feed view (7 feeds, not 50 raw tables)."""
    import warehouse
    try:
        summ = warehouse.query("src_summary", "SELECT grain, \"table\", records, entities, corroborated FROM t")
    except Exception:
        return jsonify(ok=True, feeds=[], note="src_summary not built yet — runs on the next master rebuild")
    order = {"brand": 0, "product": 1, "item": 2, "sku": 3, "outlet": 4, "pricing": 5, "inventory": 6}
    for f in summ:
        try:
            f["sources"] = sorted({r["source"] for r in warehouse.query(f["table"], "SELECT DISTINCT source FROM t")})
        except Exception:
            f["sources"] = []
    for g, tbl in (("pricing", "src_pricing"), ("inventory", "src_inventory")):
        try:
            n = warehouse.query(tbl, "SELECT count(*) c FROM t")[0]["c"]
            summ.append({"grain": g, "table": tbl, "records": n, "entities": None, "corroborated": None,
                         "sources": sorted({r["source"] for r in warehouse.query(tbl, "SELECT DISTINCT source FROM t")})})
        except Exception:
            pass
    summ.sort(key=lambda f: order.get(f["grain"], 9))
    return jsonify(ok=True, feeds=summ)


@app.get("/api/coverage")
def coverage_ep():
    """Coverage map — the FULL book of accounts (src_outlets, every store-bearing source) as by-source and
    by-state rollups plus the geocoded points to pin. The map opens on ALL sources. Accounts without precise
    lat/lng roll up to their state (bubble); geocoded ones drop a pin."""
    return jsonify(_ttl("coverage", 90, _coverage_data))


def _coverage_data():
    import warehouse
    try:
        by_src = warehouse.query("src_outlets", "SELECT source, count(*) n, "
                                 "count(*) FILTER (WHERE lat IS NOT NULL) geo FROM t GROUP BY source ORDER BY n DESC")
    except Exception:
        return dict(ok=True, total=0, by_source=[], by_state=[], points=[],
                    note="src_outlets not built yet — runs on the next master rebuild")
    by_state = warehouse.query("src_outlets", "SELECT upper(trim(state)) state, count(*) n, "
                               "count(*) FILTER (WHERE lat IS NOT NULL) geo FROM t "
                               "WHERE nullif(trim(state),'') IS NOT NULL GROUP BY state ORDER BY n DESC")
    total = sum(r["n"] for r in by_src)
    xs = warehouse.query("src_outlets", "SELECT upper(trim(state)) state, source, count(*) n FROM t "
                         "WHERE nullif(trim(state),'') IS NOT NULL GROUP BY state, source")
    try:
        by_chain = warehouse.query("src_outlets", "SELECT chain, count(*) n FROM t WHERE nullif(chain,'') "
                                   "IS NOT NULL GROUP BY chain ORDER BY n DESC LIMIT 60")
        chained = warehouse.query("src_outlets", "SELECT count(*) n FROM t WHERE is_chain")[0]["n"]
    except Exception:
        by_chain, chained = [], 0
    try:                                    # sell-flag rollup (beer/wine/spirits/hemp/cannabis + rtd-spirits class)
        flags = warehouse.query("src_outlets", "SELECT count(*) FILTER (WHERE f_beer) beer, "
                                "count(*) FILTER (WHERE f_wine) wine, count(*) FILTER (WHERE f_spirits) spirits, "
                                "count(*) FILTER (WHERE f_hemp) hemp, count(*) FILTER (WHERE f_cannabis) cannabis, "
                                "count(*) FILTER (WHERE f_rtd_spirits) rtd_spirits FROM t")[0]
    except Exception:
        flags = {}
    try:                                    # geocoded total (pins are loaded per-viewport, not shipped here)
        geocoded = warehouse.query("src_outlets", "SELECT count(*) n FROM t WHERE lat IS NOT NULL")[0]["n"]
    except Exception:
        geocoded = 0
    return dict(ok=True, total=total, geocoded=geocoded, chained=chained, flags=flags, by_source=by_src,
                by_state=by_state, by_state_source=xs, by_chain=by_chain)


@app.get("/api/coverage/accounts")
def coverage_accounts_ep():
    """Browse / search the accounts behind the map — filter by source, state, or text. Returns full account
    rows (name, address, city, state, zip, geo, phone, hoodie_outlet) so a click opens the account's info."""
    import warehouse
    src = (request.args.get("source") or "").strip()
    state = (request.args.get("state") or "").strip().upper()
    chain = (request.args.get("chain") or "").strip()
    sells = (request.args.get("sells") or "").strip().lower()   # beer|wine|spirits|hemp|cannabis|rtd_spirits
    q = (request.args.get("q") or "").strip().lower()
    try:
        limit = min(1000, max(1, int(request.args.get("limit") or 300)))
    except Exception:
        limit = 300
    if not (src or state or chain or sells or q):       # the landing view — TTL-cache + warm it
        return jsonify(_ttl("cov_accounts_%d" % limit, 90, lambda: _cov_accounts_data("", "", "", "", "", limit)))
    return jsonify(_cov_accounts_data(src, state, chain, sells, q, limit))


def _cov_accounts_data(src, state, chain, sells, q, limit):
    import warehouse
    where, params = ["1=1"], []
    if src:
        where.append("source = ?"); params.append(src)
    if state:
        where.append("upper(trim(state)) = ?"); params.append(state)
    if chain:
        where.append("chain = ?"); params.append(chain)
    if sells in ("beer", "wine", "spirits", "hemp", "cannabis", "rtd_spirits"):
        where.append("f_%s" % sells)
    if q:
        where.append("(lower(store_name) LIKE ? OR lower(city) LIKE ? OR lower(address) LIKE ?)")
        params += ["%" + q + "%"] * 3
    try:
        rows = warehouse.query("src_outlets", "SELECT source, store_id, store_name, chain, is_chain, f_beer, f_wine, "
                               "f_spirits, f_hemp, f_cannabis, f_rtd_spirits, flag_basis, license_conflict, address, "
                               "city, state, zip, addr_valid, CAST(lat AS DOUBLE) lat, CAST(lng AS DOUBLE) lng, phone, "
                               "hoodie_outlet FROM t WHERE %s "
                               "ORDER BY (nullif(trim(store_name),'') IS NULL), (store_name LIKE '#%%'), store_name "
                               "LIMIT %d"
                               % (" AND ".join(where), limit), params)   # real names first, bare store-#s then empties last
    except Exception as e:
        return dict(ok=False, error=str(e)[:160], accounts=[])
    return dict(ok=True, count=len(rows), accounts=rows)


@app.get("/api/coverage/points")
def coverage_points_ep():
    """One pin per outlet within a map viewport (bbox=west,south,east,north) — so the map renders every
    geocoded account at national scale without shipping all ~200k at once. Optional source / sells filters."""
    import warehouse
    try:
        w, s, e, n = [float(x) for x in (request.args.get("bbox") or "").split(",")]
    except Exception:
        return jsonify(ok=True, points=[], capped=False)
    src = (request.args.get("source") or "").strip()
    sells = (request.args.get("sells") or "").strip().lower()
    cond = ["lat IS NOT NULL", "CAST(lat AS DOUBLE) BETWEEN ? AND ?", "CAST(lng AS DOUBLE) BETWEEN ? AND ?"]
    params = [s, n, w, e]
    if src:
        cond.append("source = ?"); params.append(src)
    if sells in ("beer", "wine", "spirits", "hemp", "cannabis", "rtd_spirits"):
        cond.append("f_%s" % sells)
    try:
        rows = warehouse.query("src_outlets", "SELECT source, store_name, chain, f_beer, f_wine, f_spirits, f_hemp, "
                               "f_rtd_spirits, round(CAST(lat AS DOUBLE),5) lat, round(CAST(lng AS DOUBLE),5) lng "
                               "FROM t WHERE %s LIMIT 45000" % " AND ".join(cond), params)
    except Exception as e:
        return jsonify(ok=False, error=str(e)[:160], points=[])
    return jsonify(ok=True, points=rows, capped=len(rows) >= 45000)


@app.get("/api/coverage/dots")
def coverage_dots_ep():
    """Every geocoded outlet as [lat, lng, source_index] for a ONE-POINT-PER-OUTLET canvas layer — all points
    drawn in one pass, so zoomed out they fill into a coverage/density picture and zoomed in they're distinct.
    Returns a sources[] index for colouring. Optional source / sells filter. Reservoir-sampled above 220k."""
    src = (request.args.get("source") or "").strip()
    sells = (request.args.get("sells") or "").strip().lower()
    multi = request.args.get("multi") in ("1", "true", "yes")   # only outlets CONFIRMED by 2+ sources at one spot
    if not src and not sells and not multi:  # the full-book layer (biggest payload) — TTL-cache + warm it
        return jsonify(_ttl("cov_dots", 90, lambda: _cov_dots_data("", "", False)))
    if multi and not src and not sells:
        return jsonify(_ttl("cov_dots_multi", 90, lambda: _cov_dots_data("", "", True)))
    return jsonify(_cov_dots_data(src, sells, multi))


def _cov_dots_data(src, sells, multi=False):
    import warehouse
    cond = ["lat IS NOT NULL"]
    params = []
    if src:
        cond.append("source = ?"); params.append(src)
    if sells in ("beer", "wine", "spirits", "hemp", "cannabis", "rtd_spirits"):
        cond.append("f_%s" % sells)
    wsql = " AND ".join(cond)
    # multi-source = the same physical outlet seen by >=2 sources. Geocodes vary between sources, so we bucket to a
    # ~110m cell (round lat/lng to 3 dp) and keep cells where >=2 DISTINCT sources land — the corroboration view.
    # warehouse.query exposes the parquet as the view `t`; the multi self-join references that same view.
    join = ("" if not multi else
            " JOIN (SELECT round(CAST(lat AS DOUBLE),3) cy, round(CAST(lng AS DOUBLE),3) cx FROM t "
            "WHERE lat IS NOT NULL GROUP BY 1,2 HAVING count(DISTINCT source) >= 2) m "
            "ON round(CAST(t.lat AS DOUBLE),3)=m.cy AND round(CAST(t.lng AS DOUBLE),3)=m.cx")
    tw = (wsql.replace("lat IS NOT NULL", "t.lat IS NOT NULL").replace("source =", "t.source =")) if multi else wsql
    try:
        n = _wq_cached("src_outlets", "SELECT count(*) c FROM t%s WHERE %s" % (join, tw), params)[0]["c"]
        samp = "" if n <= 220000 else " USING SAMPLE reservoir(210000 ROWS)"
        rows = warehouse.query("src_outlets", "SELECT t.source, round(CAST(t.lat AS DOUBLE),4) y, "
                               "round(CAST(t.lng AS DOUBLE),4) x FROM t%s WHERE %s%s"
                               % (join, tw, samp), params)
    except Exception as e:
        return dict(ok=False, error=str(e)[:160], points=[], sources=[])
    srcs, idx, pts = [], {}, []
    for r in rows:
        s = r["source"]
        if s not in idx:
            idx[s] = len(srcs); srcs.append(s)
        pts.append([r["y"], r["x"], idx[s]])
    return dict(ok=True, total=n, sources=srcs, points=pts)


@app.get("/api/coverage/at")
def coverage_at_ep():
    """The single nearest account to a clicked lat/lng (within a small box) — full record for the detail panel."""
    import warehouse
    try:
        lat = float(request.args.get("lat")); lng = float(request.args.get("lng"))
        r = min(0.2, max(0.0005, float(request.args.get("r", 0.03))))
    except (TypeError, ValueError):
        return jsonify(ok=False, account=None)
    try:
        rows = warehouse.query("src_outlets", "SELECT source, store_id, store_name, chain, is_chain, f_beer, f_wine, "
                               "f_spirits, f_hemp, f_cannabis, f_rtd_spirits, flag_basis, license_conflict, address, "
                               "city, state, zip, addr_valid, CAST(lat AS DOUBLE) lat, CAST(lng AS DOUBLE) lng, phone, "
                               "hoodie_outlet FROM t WHERE lat IS NOT NULL AND CAST(lat AS DOUBLE) BETWEEN ? AND ? "
                               "AND CAST(lng AS DOUBLE) BETWEEN ? AND ? ORDER BY power(CAST(lat AS DOUBLE)-?,2)+"
                               "power(CAST(lng AS DOUBLE)-?,2) LIMIT 1",
                               [lat - r, lat + r, lng - r, lng + r, lat, lng])
    except Exception as e:
        return jsonify(ok=False, error=str(e)[:160], account=None)
    return jsonify(ok=True, account=(rows[0] if rows else None))


@app.get("/api/master/workbench/matches")
def mwb_matches_ep():
    """Dedup candidates at EVERY grain — brand / product / item / supplier — that the Hoodie ID mnemonic blocks
    together (New Amsterdam / New Amsteram; Maker's Mark / Makers Mark), similarity-filtered, confidence + impact
    ranked. These are the matches ABOVE the SKU: corroboration that holds without an exact SKU match. From
    wb_matches (rebuilt each cycle); decided groups drop out. ?grain=brand|product|item|supplier filters."""
    q = (request.args.get("q") or "").strip().lower()
    grain = (request.args.get("grain") or "").strip().lower()
    view = _wb_view("wb_matches")
    if view is None:
        return jsonify(ok=True, count=0, merges=[], by_grain={}, note="wb_matches not built yet — runs on the next master rebuild")
    decided = {d.get("cluster_id") for d in _wq("master_decisions", "SELECT cluster_id, action FROM t")
               if d.get("action") in ("merge", "keep_separate")}
    out, by_grain = [], {}
    for r in view:
        by_grain[r.get("grain")] = by_grain.get(r.get("grain"), 0) + 1
        if r["merge_id"] in decided:
            continue
        if grain and (r.get("grain") or "") != grain:
            continue
        if q and q not in (r.get("canonical") or "").lower():
            continue
        m = dict(r)
        try:
            m["members"] = json.loads(r.get("members") or "[]")
        except Exception:
            m["members"] = []
        out.append(m)
    out.sort(key=lambda x: (-(x.get("confidence") or 0), -(x.get("total_rows") or 0)))   # surest + highest-impact first
    return jsonify(ok=True, count=len(out), by_grain=by_grain, merges=out[:400])


@app.get("/api/master/workbench/image-matches")
def mwb_image_matches_ep():
    """Cross-source candidate pairs the PRODUCT IMAGES agree on (CLIP embedding, img_embed.py) — the ones the
    mnemonic blocking MISSES (e.g. an off-premise store-brand-name row that never blocks with the real brand).
    Each pair is already scored against the item/product attributes: `verdict` (same_item / review) + a combined
    `confidence`, with look-alikes (same label, different varietal) demoted out. SOFT signal for the adjudication
    queue — it proposes a merge to confirm, it never merges on its own or assigns a UPC."""
    return jsonify(_ttl("img_matches_resp", 120, _img_matches_data))


def _img_matches_data():
    try:
        rows = _wq_cached("img_matches", "SELECT * FROM t WHERE verdict IN ('same_item','review') "
                          "ORDER BY confidence DESC LIMIT 500")
    except Exception:
        return dict(ok=True, count=0, matches=[], note="img_matches not built yet — run img_embed build+match")
    if not rows:
        return dict(ok=True, count=0, matches=[])
    need = {}
    for m in rows:
        need.setdefault(m["a_source"], set()).add(str(m["a_sku"]))
        need.setdefault(m["b_source"], set()).add(str(m["b_sku"]))
    nm = {}
    for s, ks in need.items():
        ks = list(ks)
        for i in range(0, len(ks), 900):
            ch = ks[i:i + 900]; ph = ",".join("?" * len(ch))
            for r in _wq("_stage_product", "SELECT CAST(_source_id AS VARCHAR) k, product_name, brand, varietal, "
                         "size_ml FROM t WHERE _source=? AND CAST(_source_id AS VARCHAR) IN (%s)" % ph, [s] + ch):
                nm[(s, r["k"])] = r

    def side(src, sku, upc):
        r = nm.get((src, str(sku)), {})
        return {"source": src, "sku": sku, "upc": upc or "", "name": r.get("product_name") or "",
                "brand": r.get("brand") or "", "varietal": r.get("varietal") or "", "size_ml": r.get("size_ml")}
    out = [{"a": side(m["a_source"], m["a_sku"], m.get("a_upc")),
            "b": side(m["b_source"], m["b_sku"], m.get("b_upc")),
            "cosine": m["cosine"], "confidence": m["confidence"], "verdict": m["verdict"],
            "name_sim": m["name_sim"], "varietal_agree": m["varietal_agree"], "size_agree": m["size_agree"]}
           for m in rows]
    return dict(ok=True, count=len(out), matches=out)


@app.get("/api/master/workbench/merge-detail/<path:mid>")
def mwb_merge_detail_ep(mid):
    """Full detail for one merge group — each member SKU enriched with its image / geo / varietal / ABV /
    vintages / source list, so a reviewer can SEE the products side by side before deciding to merge."""
    view = _wb_view("wb_merges")
    if view is None:
        return jsonify(ok=False, error="not built"), 404
    grp = next((r for r in view if r.get("merge_id") == mid), None)
    if not grp:
        return jsonify(ok=False, error="not found"), 404
    try:
        members = json.loads(grp.get("members") or "[]")
    except Exception:
        members = []
    try:
        attrs = json.loads(grp.get("attrs") or "{}")     # shared product attributes (pre-computed in wb_merges)
    except Exception:
        attrs = {}
    # The group + its shared attributes come straight from the cached wb_merges row — no dim scans on modal open.
    # The ONLY thing that varies per member is the image each source showed for THIS upc (one batched scan).
    # Enrich each member with ALL the attributes we captured (by UPC, one batched scan) — so a reviewer judges
    # the cluster on the FULL data (size / abv / pack / container / origin / varietal / name), not UPC alone.
    # UPC -> attributes comes from the cached merge-UPC map (built once, warmed) — no per-open _stage_product scan.
    upc_attr = _wb_merge_upc_attrs()
    detailed = []
    for m in members:
        a = upc_attr.get(str(m.get("upc"))) or {}
        detailed.append({"sku_key": m.get("sku_key"), "upc": m.get("upc") or "", "gtin": m.get("gtin") or "",
                         "sources": m.get("sources"), "source_rows": m.get("source_rows"),
                         "source_list": m.get("source_list") or a.get("srcs") or [],
                         "variant": a.get("nm") or "", "size_ml": a.get("size_ml"), "abv": a.get("abv"),
                         "container": a.get("container") or "", "pack": a.get("pack") or "",
                         "origin": a.get("origin") or "", "varietal": a.get("varietal") or "",
                         "ttb_id": a.get("source_id") if (a.get("srcs") and "ttb_products" in (a.get("srcs") or [])) else "",
                         "image": a.get("img") or attrs.get("image") or ""})
    hero = next((d for d in detailed if d.get("image")), (detailed[0] if detailed else {}))
    return jsonify(ok=True, merge_id=mid, name=grp.get("name"), brand=grp.get("brand"),
                   reason=grp.get("reason"), distinct_upc=grp.get("distinct_upc"),
                   size_ml=grp.get("size_ml"), n_skus=grp.get("n_skus"), total_rows=grp.get("total_rows"),
                   attrs=attrs, hero=hero, members=detailed)


@app.post("/api/master/workbench/merge")
def mwb_merge_decide_ep():
    """Record a cluster-review decision. `action`: 'merge' (these are one SKU — optionally only the SELECTED
    sku_keys, so a reviewer can merge the 20 they know match and leave the rest), 'split' (the selected sku_keys
    are a distinct SKU from the others), or 'keep_separate' (all genuinely distinct). Persists to
    master_decisions (keyed by merge_id) with the member subset; the master applies it on the next rebuild."""
    import time as _t
    import warehouse
    d = request.get_json(silent=True) or {}
    mid = (d.get("merge_id") or "").strip()
    action = (d.get("action") or "").strip()
    if not mid or action not in ("merge", "keep_separate", "split"):
        return jsonify(ok=False, error="merge_id + action (merge|split|keep_separate) required"), 400
    sku_keys = [str(s) for s in (d.get("sku_keys") or []) if s]
    dec = dict(cluster_id=mid, action=action, tier="merge-review", note=(d.get("note") or "")[:200],
               matched_name=(d.get("name") or "")[:120], steward=(d.get("steward") or "cluster-review"),
               members=json.dumps(sku_keys)[:4000], removed="[]", ts=int(_t.time()))
    existing = [e for e in _wq("master_decisions", "SELECT * FROM t") if e.get("cluster_id") != mid]
    warehouse.write_parquet("master_decisions", existing + [dec])
    return jsonify(ok=True, merge_id=mid, action=action)


@app.post("/api/master/workbench/match")
def mwb_match_ep():
    """Human tier of the cascade — a steward confirms or rejects a candidate the automations + Claude couldn't.
    Records the decision; on 'confirm' the item is promoted into the master as matched_by='manual'."""
    import time as _t
    import warehouse
    d = request.get_json(silent=True) or {}
    cid = (d.get("cluster_id") or "").strip()
    action = (d.get("action") or "").strip()
    if not cid or action not in ("confirm", "reject", "escalate", "split"):
        return jsonify(ok=False, error="cluster_id + action (confirm|reject|escalate|split) required"), 400
    tier = (d.get("tier") or "").strip().lower() if action == "escalate" else ""
    # split = prune wrong members OUT of this cluster (the listed sources don't belong); recorded so the next
    # matcher run re-attributes them instead of collapsing them here. VIP/vpid is never the split anchor.
    members = [str(m) for m in (d.get("members") or []) if str(m).strip()][:200]
    dec = dict(cluster_id=cid, action=action, tier=tier, note=(d.get("note") or "")[:200],
               matched_name=(d.get("matched_name") or "")[:120], steward=(d.get("steward") or "workbench"),
               members="[]", removed=json.dumps(members), ts=int(_t.time()))
    existing = [e for e in _wq("master_decisions", "SELECT * FROM t") if e.get("cluster_id") != cid]
    warehouse.write_parquet("master_decisions", existing + [dec])
    if action == "confirm":
        row = _wq("ttb_review", "SELECT * FROM t WHERE cluster_id = ?", [cid])
        if row:
            r = dict(row[0]); r["matched_by"] = "manual"; r["tier"] = 1
            if dec["matched_name"]:
                r["candidate_name"] = dec["matched_name"]
            master = [m for m in _wq("ttb_master", "SELECT * FROM t") if m.get("cluster_id") != cid]
            warehouse.write_parquet("ttb_master", master + [r])
    return jsonify(ok=True, cluster_id=cid, action=action, tier=tier)


# ── Master data stewardship — human fills missing MASTER fields (geo especially), persisted as ground truth
# in master_overrides. This is the feedback loop: the values reinforce the automated extractors (retailer facets
# + label vision) and overlay the master on rebuild. Geo (origin→region→sub_region→appellation) is master data,
# so it must be fillable at the low level, not curated. ──
_STEWARD_FIELDS = [
    {"key": "origin", "label": "Origin — source of the juice (COO)", "col": "origin"},
    {"key": "bottled_in", "label": "Bottled in", "col": "bottled_in"},
    {"key": "region", "label": "Region", "col": None},
    {"key": "sub_region", "label": "Sub-region", "col": None},
    {"key": "appellation", "label": "Appellation", "col": None},
    {"key": "varietal", "label": "Grape / varietal", "col": None},
]
_STEWARD_KEYS = {f["key"] for f in _STEWARD_FIELDS}


def _steward_overrides():
    return _wq("master_overrides", "SELECT product_key, field, value FROM t")


@app.get("/api/master/steward/summary")
def steward_summary_ep():
    from collections import defaultdict
    cols = [f["col"] for f in _STEWARD_FIELDS if f["col"]]
    prods = _wq_cached("dim_product", "SELECT product_key, %s FROM t" % ", ".join(cols))
    total = len(prods)
    byfield = defaultdict(set)
    for o in _steward_overrides():
        byfield[o.get("field")].add(o.get("product_key"))
    col_filled = {c: {p["product_key"] for p in prods if (p.get(c) or "").strip()} for c in cols}
    out = []
    for f in _STEWARD_FIELDS:
        filled = set(byfield.get(f["key"], set()))
        if f["col"]:
            filled |= col_filled.get(f["col"], set())
        out.append({"key": f["key"], "label": f["label"], "filled": len(filled),
                    "missing": max(0, total - len(filled)), "total": total})
    return jsonify(ok=True, fields=out, total=total)


@app.get("/api/master/steward/queue")
def steward_queue_ep():
    field = (request.args.get("field") or "").strip()
    limit = min(int(request.args.get("limit") or 40), 100)
    fdef = next((f for f in _STEWARD_FIELDS if f["key"] == field), None)
    if not fdef:
        return jsonify(ok=False, error="unknown field"), 400
    brands = _wb_maps()[0]
    # most-corroborated products first — the ones that matter most to get right
    prods = _wq_cached("dim_product", "SELECT product_key, brand_key, product_name, category, origin, bottled_in, "
                "style, sources FROM t ORDER BY sources DESC LIMIT 6000")
    have = {o.get("product_key") for o in _steward_overrides() if o.get("field") == field}
    out = []
    for p in prods:
        pk = p.get("product_key")
        if pk in have:
            continue
        if fdef["col"] and (p.get(fdef["col"]) or "").strip():
            continue
        out.append({"product_key": pk, "name": p.get("product_name"), "brand": brands.get(p.get("brand_key")) or "",
                    "category": p.get("category"), "origin": p.get("origin"), "sources": p.get("sources")})
        if len(out) >= limit:
            break
    return jsonify(ok=True, field=field, label=fdef["label"], count=len(out), items=out)


@app.post("/api/master/steward/set")
def steward_set_ep():
    import time as _t
    import warehouse
    d = request.get_json(silent=True) or {}
    pk = (d.get("product_key") or "").strip()
    field = (d.get("field") or "").strip()
    value = (d.get("value") or "").strip()
    if not pk or field not in _STEWARD_KEYS or not value:
        return jsonify(ok=False, error="product_key + field + value required"), 400
    rec = dict(product_key=pk, field=field, value=value[:120],
               steward=(d.get("steward") or "steward")[:60], ts=int(_t.time()))
    existing = [e for e in _wq("master_overrides", "SELECT * FROM t")
                if not (e.get("product_key") == pk and e.get("field") == field)]
    warehouse.write_parquet("master_overrides", existing + [rec])
    return jsonify(ok=True, product_key=pk, field=field, value=rec["value"])


@app.post("/api/ingest/<dataset>")
def ingest_ep(dataset):
    """The easy way to push data INTO Unifyd. Any script POSTs JSON records here (Bearer
    INGEST_TOKEN) and they land straight in the warehouse as <dataset>, which the whole suite
    already reads. Body: {"records":[{...}], "mode":"append"|"replace", "fields":[...]?}.
    append = merge onto existing rows; replace = overwrite. Returns {ok, dataset, rows}."""
    import warehouse, re
    name = re.sub(r"[^a-z0-9_]", "", (dataset or "").lower())[:60]
    if not name:
        return jsonify(ok=False, error="bad dataset name"), 400
    body = request.get_json(silent=True) or {}
    recs = body.get("records")
    if not isinstance(recs, list) or not recs:
        return jsonify(ok=False, error="records[] required (non-empty list of objects)"), 400
    if not all(isinstance(r, dict) for r in recs):
        return jsonify(ok=False, error="each record must be an object"), 400
    mode = body.get("mode", "append")
    fields = body.get("fields") or None
    try:
        if mode == "append":
            try:
                existing = warehouse.query(name, "SELECT * FROM t")
            except Exception:
                existing = []
            recs = existing + recs
        r = warehouse.write_parquet(name, recs, fields)
        return jsonify(ok=True, dataset=name, rows=r.get("rows"), mode=mode)
    except Exception as e:
        return jsonify(ok=False, error=str(e)[:200]), 500

# ── Connector runner (front-end for pulls, no CLI). Runs the connector as a DETACHED SUBPROCESS
# (immune to gunicorn worker recycling that would kill an in-worker thread) with FILE-BASED status
# (readable from any worker). Kroger is a plain HTTPS API so Fly runs it directly; kroger_api lands
# INCREMENTALLY per zip, so progress is saved + visible and a restart never wipes the whole run.
_CONN_CREDS = {}         # connector -> {id, secret}  (in-memory cache so you don't re-enter every run)
_CONN_RUNNABLE = {"kroger"}   # extend as API connectors land (target, …)
_CONN_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "agent_state", "conn")


def _pid_alive(pid):
    try:
        os.kill(int(pid), 0); return True
    except Exception:
        return False


@app.post("/api/connector/run")
def connector_run():
    body = request.get_json(silent=True) or {}
    connector = (body.get("connector") or "").strip().lower()
    if connector not in _CONN_RUNNABLE:
        return jsonify(ok=False, error="unknown/unsupported connector (runnable: %s)" % ", ".join(sorted(_CONN_RUNNABLE))), 400
    os.makedirs(_CONN_DIR, exist_ok=True)
    jid = "cj-%d-%d" % (int(time.time()), random.randint(100, 999))
    here = os.path.dirname(os.path.abspath(__file__))
    if connector == "kroger":
        cid = body.get("client_id") or _CONN_CREDS.get("kroger", {}).get("id") or os.environ.get("KROGER_CLIENT_ID", "")
        sec = body.get("client_secret") or _CONN_CREDS.get("kroger", {}).get("secret") or os.environ.get("KROGER_CLIENT_SECRET", "")
        if not (cid and sec):
            return jsonify(ok=False, error="Kroger Client ID + Secret required"), 400
        _CONN_CREDS["kroger"] = {"id": cid, "secret": sec}
        import kroger_api
        zips = (body.get("zips") or "").strip() or ",".join(kroger_api.DEFAULT_ZIPS)
        terms = (body.get("terms") or "").strip() or ",".join(kroger_api.DEFAULT_TERMS)
        env = dict(os.environ, KROGER_CLIENT_ID=cid, KROGER_CLIENT_SECRET=sec)
        args = [sys.executable, os.path.join(here, "kroger_api.py"), "--zips", zips, "--terms", terms]
        lf = open(os.path.join(_CONN_DIR, jid + ".log"), "w")
        p = subprocess.Popen(args, env=env, stdout=lf, stderr=subprocess.STDOUT, start_new_session=True, cwd=here)
        json.dump({"connector": connector, "status": "running", "pid": p.pid, "started": int(time.time())},
                  open(os.path.join(_CONN_DIR, jid + ".json"), "w"))
        return jsonify(ok=True, jobId=jid, status="running")
    return jsonify(ok=False, error="not runnable: " + connector), 400


@app.get("/api/connector/status/<jid>")
def connector_status(jid):
    jid = re.sub(r"[^0-9a-z-]", "", jid or "")
    mp = os.path.join(_CONN_DIR, jid + ".json")
    if not os.path.exists(mp):
        return jsonify(ok=False, error="no such job"), 404
    try:
        m = json.load(open(mp))
    except Exception:
        return jsonify(ok=False, error="job state unreadable"), 500
    log = ""
    lp = os.path.join(_CONN_DIR, jid + ".log")
    if os.path.exists(lp):
        try: log = open(lp, "r", errors="replace").read()[-4000:]
        except Exception: pass
    if m.get("status") == "running":
        # completion detected from the LOG (a finished detached child becomes a zombie, so the pid can
        # still look 'alive' — the log markers are the reliable signal).
        done = re.search(r"DONE (\d+) products across (\d+) stores", log)
        if done:
            m["status"] = "done"; m["result"] = {"products": int(done.group(1)), "stores": int(done.group(2))}
        elif "401" in log:
            m["status"] = "error"; m["error"] = ("Kroger rejected the Client ID/Secret (401). The secret is shown once "
                                                 "at app creation — regenerate it in your Kroger app and retry.")
        elif "[kroger] OFF" in log or "token failed" in log:
            m["status"] = "error"; m["error"] = ("Kroger auth failed — confirm the Client ID/Secret and that the app "
                                                 "has the product.compact scope and is a production app.")
        elif m.get("pid") and not _pid_alive(m["pid"]) and log:
            m["status"] = "error"; m["error"] = "run ended without completing — see log."
        if m["status"] != "running":
            try: json.dump(m, open(mp, "w"))
            except Exception: pass
    m["log"] = log
    prog = re.findall(r"zip (\d+)/(\d+) \([^)]*\) . (\d+) products, (\d+) stores", log)   # live progress
    if prog:
        z, zt, n, st = prog[-1]
        r = m.get("result") or {}
        r.update({"products": int(n), "stores": int(st), "progress": "%s/%s zips" % (z, zt)})
        m["result"] = r
    return jsonify(ok=True, jobId=jid, **m)


@app.get("/api/connector/creds")
def connector_creds():
    """Which connectors already have creds this process can reuse (so the UI can pre-check them)."""
    have = {c: bool(_CONN_CREDS.get(c, {}).get("id") or (c == "kroger" and os.environ.get("KROGER_CLIENT_ID")))
            for c in _CONN_RUNNABLE}
    return jsonify(ok=True, have=have)


def _conn_last_run(meta):
    """The last run for a connector, from the SINGLE most authoritative source available: the connector's own
    <x>_runs warehouse table first, then the in-process RUNS log, then (last resort) a live data-table count
    that proves we HAVE data even when run history is missing. This is what makes 'what's running' legible."""
    rt = meta.get("runs")
    if rt:
        try:
            import warehouse
            rows = warehouse.query(rt, "SELECT * FROM t")
            if rows:
                last = rows[-1]
                cnt = last.get("products") or last.get("skus") or last.get("rows") or last.get("total")
                return {"run_id": last.get("run_id"), "at": last.get("at"), "status": last.get("status", "ok"),
                        "count": cnt, "source": rt}
        except Exception:
            pass
    r = next((x for x in RUNS if x.get("connId") == meta["id"]), None)
    if r:
        return {"run_id": r.get("id"), "at": r.get("finishedAt"), "status": r.get("status"),
                "count": r.get("total"), "source": "runs.json"}
    if meta.get("data"):
        n = _wh_count(meta["data"])
        if n:
            return {"run_id": None, "at": None, "status": "data-present", "count": n, "source": meta["data"]}
    return None


@app.get("/api/connectors")
def connectors_list():
    """ONE board for every source: what exists, whether it's enabled (on/off), and its true last run — merged
    across warehouse <x>_runs, the runs log, and live data counts. The Pulls console renders off this."""
    en = _conn_enabled()
    out = []
    for m in CONNECTORS_META:
        out.append({"id": m["id"], "label": m.get("label"), "group": m.get("group"),
                    "heavy": bool(m.get("heavy")), "needs_creds": bool(m.get("needs_creds")),
                    "toggle": bool(m.get("toggle")), "data": m.get("data"),
                    "enabled": en.get(m["id"], True),
                    "runnable": m["id"] in VALID_CONNS,
                    "last_run": _conn_last_run(m)})
    return jsonify(ok=True, connectors=out, groups=sorted({m.get("group", "") for m in CONNECTORS_META}))


@app.post("/api/connectors/toggle")
def connectors_toggle():
    """Turn a source on/off (e.g. DoorDash, Google). Disabled sources are skipped by the scheduler and shown
    off in the UI; they stay manually runnable. Persisted like the rest of the agent state."""
    body = request.get_json(silent=True) or {}
    cid = (body.get("id") or "").strip()
    if cid not in {m["id"] for m in CONNECTORS_META}:
        return jsonify(ok=False, error="unknown connector"), 400
    st = load("connectors.json", {}) or {}
    en = st.setdefault("enabled", {})
    en[cid] = bool(body.get("enabled", True))
    _save_json("connectors.json", st)
    return jsonify(ok=True, id=cid, enabled=en[cid])

@app.get("/api/ttb-label/<ttbid>")
def ttb_label_ep(ttbid):
    """Serve a stored TTB label thumbnail (the enrichment uploads them to Tigris as
    label_images/<ttbid>.jpg). The original TTB attachment URLs are F5-gated and won't load in a
    browser, so the tracker's label modal points at this instead. 404 if we haven't captured it yet."""
    import warehouse, re
    tid = re.sub(r"[^0-9]", "", ttbid or "")[:20]
    if not tid:
        return ("bad id", 400)
    data = warehouse.get_bytes("label_images/%s.jpg" % tid)
    if not data:
        return ("not captured", 404)
    headers = {"Content-Type": "image/jpeg", "Cache-Control": "private, max-age=86400"}
    if request.args.get("download"):
        headers["Content-Disposition"] = 'attachment; filename="ttb_%s.jpg"' % tid
    return Response(data, headers=headers)

@app.get("/api/scrape/<sid>/rows")
def scrape_rows_ep(sid):
    """A sample of the actual rows for ANY tracked scrape, straight from its warehouse table — so the
    tracker can show the data (not just counts) for every source, not only Walmart. Optional ?q= filters
    across all columns; ?limit caps rows (default 200)."""
    import warehouse
    s = next((x for x in _SCRAPES if x["id"] == sid), None)
    if not s:
        return jsonify(ok=False, error="unknown scrape"), 404
    table = s["table"]
    q = (request.args.get("q") or "").strip()
    try:
        limit = max(1, min(1000, int(request.args.get("limit", "200"))))
    except ValueError:
        limit = 200
    # column list (works even when a filter yields 0 rows)
    try:
        cols = [r["column_name"] for r in warehouse.query(
            table, "SELECT column_name FROM information_schema.columns "
                   "WHERE table_name='t' ORDER BY ordinal_position")]
    except Exception as e:
        return jsonify(ok=True, landed=False, error=str(e)[:140], columns=[], rows=[]), 200
    # newest-first when the table carries an obvious timestamp column
    order = next((c for c in ("pulled_at", "at", "approval_date", "modified", "ts") if c in cols), None)
    sql = "SELECT * FROM t"
    params = []
    if q:
        like = "%" + q.lower() + "%"
        sql += " WHERE " + " OR ".join("lower(CAST(%s AS VARCHAR)) LIKE ?" % c for c in cols)
        params = [like] * len(cols)
    if order:
        sql += ' ORDER BY "%s" DESC' % order
    sql += " LIMIT %d" % limit
    try:
        rows = warehouse.query(table, sql, params)
    except Exception as e:
        return jsonify(ok=True, landed=True, columns=cols, rows=[], error=str(e)[:140]), 200
    # JSON-safe: stringify anything non-primitive (dates, decimals)
    for r in rows:
        for k, v in list(r.items()):
            if v is not None and not isinstance(v, (str, int, float, bool)):
                r[k] = str(v)
    return jsonify(ok=True, landed=True, table=table, columns=cols, rows=rows,
                   count=len(rows), filtered=bool(q))

@app.get("/api/chains")
def chains_ep():
    """The bev-alc chains registry — each chain's pricing/inventory feasibility + method + auth. The
    source-candidate catalog: any chain with pricing and/or inventory reachable is a source we can build."""
    import warehouse
    try:
        rows = warehouse.query("bevalc_chains",
            "SELECT chain, premise, channel, banners, est_locations, website, pricing, inventory, "
            "method, auth, note, is_source, yields, family FROM t ORDER BY is_source DESC, est_locations DESC")
    except Exception as e:
        return jsonify(ok=True, landed=False, error=str(e)[:140], chains=[]), 200
    from collections import Counter
    src = [r for r in rows if r["is_source"]]
    summary = dict(total=len(rows), sources=len(src),
                   off=sum(1 for r in rows if r["premise"] == "off"),
                   on=sum(1 for r in rows if r["premise"] == "on"),
                   pricing=sum(1 for r in rows if r["pricing"] not in ("", "no")),
                   inventory=sum(1 for r in rows if r["inventory"] not in ("", "no")),
                   by_family=dict(Counter(r["family"] for r in rows)),
                   by_channel=dict(Counter(r["channel"] for r in rows)))
    return jsonify(ok=True, landed=True, summary=summary, chains=rows)

@app.get("/api/walmart/match")
def walmart_match_ep():
    """Cross-source matches for one Walmart product — TTB COLA filings by brand (potential same product)."""
    import warehouse
    brand = (request.args.get("brand") or "").strip()
    if not brand:
        return jsonify(ok=True, brand="", matches=[])
    try:
        rows = warehouse.query("ttb_cola_detail",
            "SELECT ttb_id, CAST(brand_name AS VARCHAR) brand_name, CAST(class_type_desc AS VARCHAR) class_type, "
            "CAST(net_contents AS VARCHAR) net_contents, CAST(approval_date AS VARCHAR) approval_date "
            "FROM t WHERE lower(CAST(brand_name AS VARCHAR)) LIKE ? "
            "ORDER BY approval_date DESC LIMIT 12", ["%" + brand.lower() + "%"])
        for r in rows:
            r["source"] = "TTB COLA"; r["how"] = "brand match"
        return jsonify(ok=True, brand=brand, matches=rows)
    except Exception as e:
        return jsonify(ok=True, brand=brand, matches=[], error=str(e)[:120])

# ── Field mapping — persist source.field → master.field crosswalks, with pre/post transforms ──
# The PRODUCT master is the WIDE hierarchy schema — every field declares its GRAIN (brand|product|item|
# sku|supplier). The apply engine SHREDS one mapped source into dim_brand → dim_product → dim_item →
# dim_sku by grain (+ dim_supplier). Attribute lives at its level: brand_group@brand, flavor/abv@product,
# size@item, upc@sku. Vintage/edition are sku ATTRIBUTES but excluded from the sku identity (releases
# collapse to one sku). Price is NOT here — it's the pricing FACT, held separately.
DEFAULT_MASTER_FIELDS = [
    {"name": "brand", "grain": "brand", "type": "string", "desc": "Brand (e.g. Absolut)"},
    {"name": "brand_group", "grain": "brand", "type": "string", "desc": "Brand-group / family section within the brand"},
    {"name": "product_name", "grain": "product", "type": "string", "desc": "Product / variant (e.g. Absolut Blue)"},
    {"name": "flavor", "grain": "product", "type": "string", "desc": "Flavor / variant (applied at product level)"},
    {"name": "style", "grain": "product", "type": "string", "desc": "Style / sub-type"},
    {"name": "category", "grain": "product", "type": "string", "desc": "Category / class-type"},
    {"name": "abv", "grain": "product", "type": "number", "desc": "Alcohol % by volume (product level)"},
    {"name": "origin", "grain": "product", "type": "string", "desc": "Country / region of origin"},
    {"name": "size_ml", "grain": "item", "type": "number", "desc": "Net contents in mL (item grain)"},
    {"name": "container", "grain": "item", "type": "string", "desc": "Container (bottle / can / keg)"},
    {"name": "packsize", "grain": "item", "type": "string", "desc": "Pack / container size (as filed)"},
    {"name": "pack", "grain": "sku", "type": "string", "desc": "Pack configuration (single / case / n-pack)"},
    {"name": "upc", "grain": "sku", "type": "string", "desc": "UPC / GTIN — auto-normalized to GTIN-14", "normalize": "upc"},
    {"name": "vintage", "grain": "sku", "type": "string", "desc": "Vintage year — attribute, NOT part of sku identity", "normalize": "vintage"},
    {"name": "edition", "grain": "sku", "type": "string", "desc": "Special/seasonal edition — attribute, NOT part of sku identity"},
    {"name": "supplier", "grain": "supplier", "type": "string", "desc": "Supplier / owner (brand↔supplier SCD, not a hierarchy level)"},
]
# The OUTLET master schema — the canonical, atomic outlet fields every outlet source maps INTO
# (the 'lowest level of consistency'). Same machinery as products, different entity.
DEFAULT_OUTLET_FIELDS = [
    {"name": "outlet_name", "type": "string", "desc": "Outlet / account name", "normalize": "name"},
    {"name": "dba", "type": "string", "desc": "Doing-business-as / display name", "normalize": "name"},
    {"name": "address", "type": "string", "desc": "Street address"},
    {"name": "city", "type": "string", "desc": "City"},
    {"name": "state", "type": "string", "desc": "State (2-letter)", "normalize": "state"},
    {"name": "zip5", "type": "string", "desc": "ZIP (5-digit)", "normalize": "zip5"},
    {"name": "zip4", "type": "string", "desc": "ZIP+4 add-on", "normalize": "zip4"},
    {"name": "county_fips", "type": "string", "desc": "County FIPS"},
    {"name": "lat", "type": "number", "desc": "Latitude"},
    {"name": "lng", "type": "number", "desc": "Longitude"},
    {"name": "phone", "type": "string", "desc": "Phone (digits)", "normalize": "phone"},
    {"name": "outlet_type", "type": "string", "desc": "On/off-premise or channel"},
    {"name": "license_num", "type": "string", "desc": "State license / permit number"},
    {"name": "source_ref", "type": "string", "desc": "Source outlet id (e.g. AB vpid) — strong key when present"},
    {"name": "carriage", "type": "string", "desc": "Brands/products seen at this outlet (carriage signal)"},
]
# Master entities: name → (default schema, state-file basenames). product keeps the legacy filenames
# for back-compat; outlet gets its own. Adding an entity = one row here.
ENTITIES = {
    "product": {"defaults": DEFAULT_MASTER_FIELDS, "schema": "master_schema.json",
                "maps": "field_mappings.json", "mapmeta": "mappings_meta.json", "schemameta": "schema_meta.json"},
    "outlet":  {"defaults": DEFAULT_OUTLET_FIELDS, "schema": "master_schema_outlet.json",
                "maps": "field_mappings_outlet.json", "mapmeta": "mappings_meta_outlet.json", "schemameta": "schema_meta_outlet.json"},
}
def _entity(v):
    return v if v in ENTITIES else "product"

# Seed mappings for OUR OWN outlet connectors — so the outlet master builds turnkey (no hand-mapping
# each source; a saved mapping always overrides). Key = dataset name; a trailing '_' key = name prefix.
SEED_OUTLET_MAPPINGS = {
    "ab_outlets": [
        {"source_field": "VPID", "master_field": "source_ref"},
        {"source_field": "Name", "master_field": "outlet_name"},
        {"source_field": "Address", "master_field": "address"},
        {"source_field": "City", "master_field": "city"},
        {"source_field": "State", "master_field": "state"},
        {"source_field": "Zip", "master_field": "zip5"},
        {"source_field": "Zip", "master_field": "zip4"},
        {"source_field": "Lat", "master_field": "lat"},
        {"source_field": "Lng", "master_field": "lng"},
        {"source_field": "AB_Brands", "master_field": "carriage"},
    ],
    "vtinfo_": [   # prefix — every vtinfo_<brand> dataset
        {"source_field": "Account", "master_field": "outlet_name"},
        {"source_field": "Street", "master_field": "address"},
        {"source_field": "City", "master_field": "city"},
        {"source_field": "State", "master_field": "state"},
        {"source_field": "Zip", "master_field": "zip5"},
        {"source_field": "Phone", "master_field": "phone"},
        {"source_field": "Lat", "master_field": "lat"},
        {"source_field": "Lng", "master_field": "lng"},
        {"source_field": "StoreType", "master_field": "outlet_type"},
        {"source_field": "Brand", "master_field": "carriage"},
    ],
}
def _seed_for(ds):
    if ds in SEED_OUTLET_MAPPINGS:
        return SEED_OUTLET_MAPPINGS[ds]
    for k, v in SEED_OUTLET_MAPPINGS.items():
        if k.endswith("_") and ds.startswith(k):
            return v
    return None

def _outlet_maps_effective():
    """Persisted outlet mappings + seed mappings for any recognized warehouse dataset lacking one, so
    the outlet master builds turnkey. A saved mapping always wins over its seed."""
    out = dict(load(ENTITIES["outlet"]["maps"], {}))
    try:
        for d in warehouse.list_datasets():
            nm = d.get("name", "")
            if nm and nm not in out and _seed_for(nm):
                out[nm] = _seed_for(nm)
    except Exception as e:
        app.logger.warning("outlet seed merge failed: %s", e)
    return out

# ── FACT tables: inventory + pricing (sku × outlet × date). A fact schema = the FK-identity fields
# (outlet + product, to compute the keys) + measures + obs_date. Same config-driven build as dimensions.
_FACT_IDENTITY = [
    {"name": "source_ref", "grain": "outlet", "type": "string", "desc": "Outlet id (vpid) — outlet key"},
    {"name": "outlet_name", "grain": "outlet", "type": "string", "desc": "Outlet name", "normalize": "name"},
    {"name": "address", "grain": "outlet", "type": "string", "desc": "Outlet street"},
    {"name": "city", "grain": "outlet", "type": "string", "desc": "Outlet city"},
    {"name": "state", "grain": "outlet", "type": "string", "desc": "Outlet state", "normalize": "state"},
    {"name": "zip5", "grain": "outlet", "type": "string", "desc": "Outlet ZIP5", "normalize": "zip5"},
    {"name": "brand", "grain": "brand", "type": "string", "desc": "Brand (→ brand/sku key)"},
    {"name": "product_name", "grain": "product", "type": "string", "desc": "Product / variant"},
    {"name": "size_ml", "grain": "item", "type": "number", "desc": "Size mL (→ sku grain)"},
    {"name": "upc", "grain": "sku", "type": "string", "desc": "UPC (→ sku key)", "normalize": "upc"},
    {"name": "obs_date", "grain": "fact", "type": "string", "desc": "Observation date (else build date)"},
]
DEFAULT_INVENTORY_FIELDS = _FACT_IDENTITY + [
    {"name": "in_stock", "grain": "fact", "type": "string", "desc": "In stock (Y/N)"},
    {"name": "on_hand", "grain": "fact", "type": "number", "desc": "On-hand units"},
    {"name": "units", "grain": "fact", "type": "number", "desc": "Units / movement"},
    {"name": "carriage", "grain": "fact", "type": "string", "desc": "Brands/products carried (presence)"},
]
DEFAULT_PRICING_FIELDS = _FACT_IDENTITY + [
    {"name": "price", "grain": "fact", "type": "number", "desc": "Price"},
    {"name": "promo", "grain": "fact", "type": "string", "desc": "Promo / strike price"},
    {"name": "currency", "grain": "fact", "type": "string", "desc": "Currency"},
]
FACTS = {
    "inventory": {"defaults": DEFAULT_INVENTORY_FIELDS, "schema": "fact_schema_inventory.json",
                  "maps": "field_mappings_inventory.json", "mapmeta": "mappings_meta_inventory.json"},
    "pricing":   {"defaults": DEFAULT_PRICING_FIELDS, "schema": "fact_schema_pricing.json",
                  "maps": "field_mappings_pricing.json", "mapmeta": "mappings_meta_pricing.json"},
}
def _fact(v):
    return v if v in FACTS else "inventory"

# Seed fact mappings for our own connectors — inventory carriage from AB (brand-list) + VTInfo (per-brand).
SEED_FACT_MAPPINGS = {
    "inventory": {
        "ab_outlets": [{"source_field": "VPID", "master_field": "source_ref"},
                       {"source_field": "Name", "master_field": "outlet_name"},
                       {"source_field": "Address", "master_field": "address"},
                       {"source_field": "City", "master_field": "city"},
                       {"source_field": "State", "master_field": "state"},
                       {"source_field": "Zip", "master_field": "zip5"},
                       {"source_field": "AB_Brands", "master_field": "carriage"}],
        "vtinfo_": [{"source_field": "Account", "master_field": "outlet_name"},
                    {"source_field": "Street", "master_field": "address"},
                    {"source_field": "City", "master_field": "city"},
                    {"source_field": "State", "master_field": "state"},
                    {"source_field": "Zip", "master_field": "zip5"},
                    {"source_field": "Brand", "master_field": "brand"},
                    {"source_field": "Brand", "master_field": "carriage"}],
        "walmart_products": [{"source_field": "store_loc", "master_field": "outlet_name"},
                             {"source_field": "store", "master_field": "address"},
                             {"source_field": "brand", "master_field": "brand"},
                             {"source_field": "product_name", "master_field": "product_name"},
                             {"source_field": "size_ml", "master_field": "size_ml"},
                             {"source_field": "upc", "master_field": "upc"},
                             {"source_field": "in_stock", "master_field": "in_stock"}],
    },
    "pricing": {   # price lists (bc_liquor / or_pricing / Total Wine) — mapped in the UI per their headers
        "walmart_products": [{"source_field": "store_loc", "master_field": "outlet_name"},
                             {"source_field": "brand", "master_field": "brand"},
                             {"source_field": "product_name", "master_field": "product_name"},
                             {"source_field": "size_ml", "master_field": "size_ml"},
                             {"source_field": "upc", "master_field": "upc"},
                             {"source_field": "price", "master_field": "price"},
                             {"source_field": "list_price", "master_field": "promo"}],
    },
}
def _fact_seed_for(fact, ds):
    seeds = SEED_FACT_MAPPINGS.get(fact, {})
    if ds in seeds:
        return seeds[ds]
    for k, v in seeds.items():
        if k.endswith("_") and ds.startswith(k):
            return v
    return None

def _fact_schema(fact):
    spec = FACTS[_fact(fact)]
    fields = load(spec["schema"], spec["defaults"])
    for f in fields:
        if isinstance(f, dict) and not f.get("normalize") and f.get("name") in _NORMALIZE_DEFAULTS:
            f["normalize"] = _NORMALIZE_DEFAULTS[f["name"]]
    return fields

def _fact_maps_effective(fact):
    out = dict(load(FACTS[_fact(fact)]["maps"], {}))
    try:
        for d in warehouse.list_datasets():
            nm = d.get("name", "")
            if nm and nm not in out and _fact_seed_for(fact, nm):
                out[nm] = _fact_seed_for(fact, nm)
    except Exception as e:
        app.logger.warning("fact seed merge failed: %s", e)
    return out

# Data-dictionary transforms that can apply BEFORE (on the source value) or AFTER (on the mapped
# master value). Names only here — the apply-engine consumes them when materializing the master.
MAP_TRANSFORMS = ["none", "trim", "upper", "lower", "title_case", "digits_only",
                  "size_to_ml", "year_from_date"]
# Canonical normalizers keyed by master field NAME — ensured on read so the 'lowest level of
# consistency' contract holds even for a schema saved before a normalizer was added.
_NORMALIZE_DEFAULTS = {"upc": "upc", "gtin": "upc", "zip5": "zip5", "zip4": "zip4",
                       "state": "state", "phone": "phone", "outlet_name": "name", "dba": "name"}
# CONTROLLED fields — a field with a DOMAIN accepts only these permissible options (enum). Seeded here
# as sensible starters (editable per-field via /api/master/domain); a dictionary maps raw source strings
# INTO one of these values. Seeded only when the field carries no `domain` key yet (edits/clears win).
_DOMAIN_DEFAULTS = {
    "container": ["Bottle", "Can", "Keg", "Bag-in-Box", "Pouch", "Cask", "Growler"],
    "pack": ["Single", "2-Pack", "4-Pack", "6-Pack", "8-Pack", "12-Pack", "15-Pack",
             "18-Pack", "24-Pack", "30-Pack", "Case"],
    "outlet_type": ["On-Premise", "Off-Premise"],
    # top-level alcohol category — the clean roll-up of COLA Class/Type
    "category": ["Wine", "Beer", "Spirits", "Cider", "Mead", "Sake", "Seltzer/RTD", "Non-Alcoholic"],
    # sub-type within a category — a curated starter (prune per your taxonomy)
    "style": ["Red", "White", "Rosé", "Sparkling", "Dessert/Fortified",
              "Lager", "Ale", "IPA", "Stout", "Porter", "Pilsner", "Wheat", "Sour",
              "Bourbon", "Rye", "Scotch", "Irish Whiskey", "Blended Whiskey",
              "Vodka", "Gin", "Tequila", "Mezcal", "White Rum", "Dark Rum", "Spiced Rum",
              "Brandy", "Cognac", "Liqueur"],
    # country of origin — canonical country vocabulary (US-state origins roll up to United States
    # via a dictionary; add states here if you want state-level origin instead). Seeded from the real
    # COLA Origin distribution (France/Italy/Spain/Argentina/… dominate imports).
    "origin": ["United States", "France", "Italy", "Spain", "Argentina", "Australia", "Germany",
               "Chile", "Portugal", "New Zealand", "South Africa", "Austria", "Mexico", "Scotland",
               "Canada", "Japan", "Greece", "Belgium", "Ireland", "Netherlands", "England", "Hungary",
               "Israel", "Georgia", "Croatia", "Slovenia", "Lebanon", "Brazil", "Peru", "China",
               "India", "South Korea", "Russia", "Poland", "Sweden"],
}

def _master_schema(entity="product"):
    spec = ENTITIES[_entity(entity)]
    fields = [dict(f) for f in load(spec["schema"], spec["defaults"])]   # copy — never mutate the defaults
    for f in fields:
        if not f.get("normalize") and f.get("name") in _NORMALIZE_DEFAULTS:
            f["normalize"] = _NORMALIZE_DEFAULTS[f["name"]]
        if "domain" not in f and f.get("name") in _DOMAIN_DEFAULTS:
            f["domain"] = list(_DOMAIN_DEFAULTS[f["name"]])
    return fields

@app.get("/api/master/schema")
def master_schema_get():
    entity = _entity(request.args.get("entity", "product"))
    return jsonify(ok=True, entity=entity, entities=list(ENTITIES),
                   fields=_master_schema(entity), transforms=MAP_TRANSFORMS)

# ── users (each gets a short CODE) + created/updated audit stamps ──
def _mk_user_code(email, users):
    base = re.sub(r"[^a-z0-9]", "", (email.split("@")[0] if email else "usr")).upper()[:3] or "USR"
    taken = {u.get("code") for u in users.values()}
    code, i = base, 1
    while code in taken:
        code = base + str(i); i += 1
    return code

def _user_rec():
    """The signed-in user (from the OIDC session) with a stable short code; auto-registers on first
    sight so every created/updated stamp is attributable. (Distinct from the admin _current_user,
    which returns just the email.)"""
    email = (session.get("email") or "").lower()
    users = load("users.json", {})
    u = users.get(email)
    if email and not u:
        u = {"email": email, "code": _mk_user_code(email, users), "name": email.split("@")[0]}
        users[email] = u; _save_json("users.json", users)
    return u or {"email": "", "code": "SYS", "name": "system"}

def _stamp(meta_name, key):
    """created_at/created_by (first write) + updated_at/updated_by (every write) for a config record."""
    now = int(time.time()); who = _user_rec().get("code", "SYS")
    allmeta = load(meta_name, {})
    m = allmeta.get(key, {})
    if not m.get("created_at"):
        m["created_at"] = now; m["created_by"] = who
    m["updated_at"] = now; m["updated_by"] = who
    allmeta[key] = m; _save_json(meta_name, allmeta)
    return m

@app.get("/api/whoami")
def whoami_ep():
    return jsonify(ok=True, user=_user_rec())

@app.get("/api/users")
def users_get():
    return jsonify(ok=True, users=load("users.json", {}))

@app.post("/api/users")
def users_post():
    """Assign / update a user's code + name (keyed by email)."""
    b = request.get_json(silent=True) or {}
    email = (b.get("email") or "").lower().strip()
    if not email:
        return jsonify(ok=False, error="email required"), 400
    users = load("users.json", {})
    u = users.get(email, {"email": email})
    if b.get("code"): u["code"] = re.sub(r"[^A-Za-z0-9]", "", b["code"]).upper()[:8]
    if b.get("name"): u["name"] = b["name"]
    if not u.get("code"): u["code"] = _mk_user_code(email, users)
    users[email] = u; _save_json("users.json", users)
    return jsonify(ok=True, user=u)

@app.post("/api/master/schema")
def master_schema_post():
    """Create a master field (start building the master schema) or replace the whole set. ?entity=."""
    body = request.get_json(silent=True) or {}
    entity = _entity(body.get("entity") or request.args.get("entity", "product"))
    spec = ENTITIES[entity]
    fields = load(spec["schema"], spec["defaults"])
    if isinstance(body.get("fields"), list):
        fields = body["fields"]
    else:
        nm = (body.get("name") or "").strip()
        if not nm:
            return jsonify(ok=False, error="name required"), 400
        if not any(f.get("name") == nm for f in fields):
            fields.append({"name": nm, "type": body.get("type", "string"), "desc": body.get("desc", ""),
                           **({"normalize": body["normalize"]} if body.get("normalize") else {})})
    _save_json(spec["schema"], fields)
    return jsonify(ok=True, entity=entity, fields=fields, meta=_stamp(spec["schemameta"], "schema"))

# ── Controlled vocabularies — a master field's DOMAIN (permissible options). A field with options is an
# enum; the Dictionary maps raw source values into one of them (off-domain → unmapped). ──
@app.get("/api/master/domains")
def master_domains_get():
    """All controlled fields for an entity → {field: [permissible options]}. ?entity=product|outlet."""
    entity = _entity(request.args.get("entity", "product"))
    fields = _master_schema(entity)
    return jsonify(ok=True, entity=entity,
                   domains={f["name"]: f["domain"] for f in fields if f.get("domain")})

@app.post("/api/master/domain")
def master_domain_post():
    """Set the permissible options for ONE master field. Body: {entity?, field, options:[...]}. An empty
    list clears the constraint (the field becomes free-text again). De-dupes case-insensitively, keeps order."""
    b = request.get_json(silent=True) or {}
    entity = _entity(b.get("entity") or "product")
    spec = ENTITIES[entity]
    field = (b.get("field") or "").strip()
    if not field:
        return jsonify(ok=False, error="field required"), 400
    opts, seen = [], set()
    for o in (b.get("options") or []):
        s = str(o).strip()
        if s and s.lower() not in seen:
            seen.add(s.lower()); opts.append(s)
    fields = [dict(f) for f in load(spec["schema"], spec["defaults"])]   # copy — never mutate the defaults
    if not any(f.get("name") == field for f in fields):
        return jsonify(ok=False, error="unknown field"), 400
    for f in fields:
        if f.get("name") == field:
            f["domain"] = opts                       # explicit key (even []) — blocks the seed default
    _save_json(spec["schema"], fields)
    return jsonify(ok=True, entity=entity, field=field, options=opts,
                   meta=_stamp(spec["schemameta"], "schema"))

@app.post("/api/master/domain/check")
def master_domain_check():
    """Fit a source field's real values against a set of permissible options. Body: {dataset, field,
    options:[...]}. Returns distinct source values split into IN-domain (already a permissible option,
    case-insensitive exact) and OUT (need a dictionary rule), with row counts — so you can see how much
    of the source already conforms and what still has to be mapped."""
    b = request.get_json(silent=True) or {}
    ds = (b.get("dataset") or "ttb_cola").strip()
    field = (b.get("field") or "").strip()
    if field not in _ds_columns(ds):
        return jsonify(ok=False, error="unknown field for dataset"), 400
    allowed = {str(o).strip().lower() for o in (b.get("options") or []) if str(o).strip()}
    try:
        vals = _cola_q(ds, 'SELECT CAST("%s" AS VARCHAR) v, count(*) n FROM t '
                       "WHERE CAST(\"%s\" AS VARCHAR)<>'' GROUP BY 1 ORDER BY n DESC LIMIT 400" % (field, field))
    except Exception as e:
        return jsonify(ok=False, error=str(e)[:160]), 200
    ins, out, in_rows, out_rows = [], [], 0, 0
    for r in vals:
        if str(r["v"]).strip().lower() in allowed:
            ins.append(r); in_rows += r["n"]
        else:
            out.append(r); out_rows += r["n"]
    return jsonify(ok=True, dataset=ds, field=field, in_domain=ins[:60], out_domain=out[:60],
                   in_rows=in_rows, out_rows=out_rows,
                   fit_pct=round(100 * in_rows / max(1, in_rows + out_rows), 1))

@app.get("/api/mappings")
def mappings_get():
    """Persisted field mappings + audit meta. ?entity= (product|outlet), ?dataset= for one source's rows."""
    entity = _entity(request.args.get("entity", "product"))
    spec = ENTITIES[entity]
    ds = (request.args.get("dataset") or "").strip()
    m = load(spec["maps"], {})
    meta = load(spec["mapmeta"], {})
    if ds:
        rows = m.get(ds)
        seeded = False
        if rows is None and entity == "outlet":                  # pre-fill our own connectors' mapping
            rows = _seed_for(ds); seeded = rows is not None
        return jsonify(ok=True, entity=entity, dataset=ds, mappings=rows or [],
                       seeded=seeded, meta=meta.get(ds))
    return jsonify(ok=True, entity=entity, dataset=None, mappings=m, meta=meta)

@app.post("/api/mappings")
def mappings_post():
    """Persist a source dataset's mapping rows: [{source_field, master_field, pre, post}]. Stamps
    created/updated + user, and — when a row maps to master.upc — auto-drives the UPC healer. ?entity=."""
    body = request.get_json(silent=True) or {}
    entity = _entity(body.get("entity") or request.args.get("entity", "product"))
    spec = ENTITIES[entity]
    ds = (body.get("dataset") or "").strip()
    if not ds:
        return jsonify(ok=False, error="dataset required"), 400
    m = load(spec["maps"], {})
    m[ds] = body.get("mappings", [])
    _save_json(spec["maps"], m)
    meta = _stamp(spec["mapmeta"], ds)
    # auto-drive the UPC healer if a source field is mapped to master.upc (base data untouched)
    upc_rows = [r for r in m[ds] if r.get("master_field") == "upc" and r.get("source_field")]
    healed = False
    if upc_rows:
        healed = True
        ucol = upc_rows[0]["source_field"]
        appl = next((r["source_field"] for r in m[ds] if r.get("master_field") == "supplier"), None)
        def _heal():
            try:
                import upc_heal
                rep = upc_heal.heal(ds, ucol, appl, log=lambda x: app.logger.info("UPCHEAL %s", x))
                reports = load("upc_reports.json", {}); reports[ds] = rep; _save_json("upc_reports.json", reports)
            except Exception as e:
                app.logger.warning("upc heal %s failed: %s", ds, e)
        threading.Thread(target=_heal, daemon=True).start()
    return jsonify(ok=True, dataset=ds, count=len(m[ds]), meta=meta, upc_heal_started=healed)

@app.get("/api/upc/report")
def upc_report_ep():
    """The UPC health report for a dataset (auto-generated when a upc mapping is assigned)."""
    ds = (request.args.get("dataset") or "").strip()
    reports = load("upc_reports.json", {})
    return jsonify(ok=True, dataset=ds or None, report=(reports.get(ds) if ds else reports))

@app.post("/api/master/preview")
def master_preview_ep():
    """Apply ONE derivation rule to a sample of the dataset → [{raw, derived}], so a rule can be
    verified before it's committed. Body: {dataset, rule:{source_field,mode,pre,post,pattern,group,map,expr}}."""
    body = request.get_json(silent=True) or {}
    ds = (body.get("dataset") or "").strip()
    if not ds:
        return jsonify(ok=False, error="dataset required"), 400
    try:
        import master_apply
        rule = body.get("rule") or {}
        rule = _resolve_dict_refs({ds: [rule]})[ds][0]      # a {mode:'dict', dict:<id>} rule → inline entries
        fct = body.get("fact")                              # fact preview uses the fact schema's normalizers
        fields = _fact_schema(fct) if fct in FACTS else _master_schema(_entity(body.get("entity") or "product"))
        nz = next((f.get("normalize") for f in fields if isinstance(f, dict) and f.get("name") == rule.get("master_field")), None)
        rows = master_apply.preview(ds, rule, limit=int(body.get("limit", 12)), normalize=nz)
        return jsonify(ok=True, dataset=ds, normalize=nz, rows=rows)
    except Exception as e:
        return jsonify(ok=False, error=str(e)[:180]), 200

@app.post("/api/master/apply")
def master_apply_ep():
    """Materialize dim_<entity> + dim_<entity>_resolved from ALL persisted mappings + the master schema
    (one DuckDB pass per source over Parquet → UNION → resolve → warehouse). ?entity=product|outlet.
    Returns per-source counts + any skipped sources."""
    try:
        import master_apply
        body = request.get_json(silent=True) or {}
        entity = _entity(body.get("entity") or request.args.get("entity", "product"))
        spec = ENTITIES[entity]
        fields = _master_schema(entity)
        maps = _outlet_maps_effective() if entity == "outlet" else load(spec["maps"], {})
        maps = _resolve_dict_refs(maps)            # expand {mode:'dict', dict:<id>} rows to inline entries
        who = _user_rec().get("code", "SYS")
        res = master_apply.build(fields, maps, entity=entity, built_by=who, built_at=int(time.time()),
                                 log=lambda mm: app.logger.info("APPLY[%s] %s", entity, mm))
        return jsonify(ok=True, entity=entity, built_by=who, **res)
    except Exception as e:
        app.logger.exception("apply failed")
        return jsonify(ok=False, error=str(e)[:200]), 200

# ── FACT tables (inventory + pricing) — same config-driven build, keyed sku × outlet × date ──
@app.get("/api/master/fact/schema")
def fact_schema_get():
    fact = _fact(request.args.get("fact", "inventory"))
    return jsonify(ok=True, fact=fact, facts=list(FACTS),
                   fields=_fact_schema(fact), transforms=MAP_TRANSFORMS)

@app.get("/api/master/fact/mappings")
def fact_mappings_get():
    fact = _fact(request.args.get("fact", "inventory"))
    spec = FACTS[fact]
    ds = (request.args.get("dataset") or "").strip()
    m = load(spec["maps"], {})
    if ds:
        rows = m.get(ds)
        seeded = False
        if rows is None:
            rows = _fact_seed_for(fact, ds); seeded = rows is not None
        return jsonify(ok=True, fact=fact, dataset=ds, mappings=rows or [], seeded=seeded)
    return jsonify(ok=True, fact=fact, dataset=None, mappings=m)

@app.post("/api/master/fact/mappings")
def fact_mappings_post():
    body = request.get_json(silent=True) or {}
    fact = _fact(body.get("fact") or request.args.get("fact", "inventory"))
    spec = FACTS[fact]
    ds = (body.get("dataset") or "").strip()
    if not ds:
        return jsonify(ok=False, error="dataset required"), 400
    m = load(spec["maps"], {})
    m[ds] = body.get("mappings", [])
    _save_json(spec["maps"], m)
    return jsonify(ok=True, fact=fact, dataset=ds, count=len(m[ds]),
                   meta=_stamp(spec["mapmeta"], ds))

# ── MDM workbench: browse the OUTLET book (src_outlets — the full account universe, every store-bearing
# source). "carriage" shows what the outlet SELLS (the sell-flags); "sources" shows its source. ──
_OUTM_SELL = ("concat_ws(', ', CASE WHEN f_beer THEN 'Beer' END, CASE WHEN f_wine THEN 'Wine' END, "
              "CASE WHEN f_spirits THEN 'Spirits' END, CASE WHEN f_hemp THEN 'Hemp' END, "
              "CASE WHEN f_cannabis THEN 'Cannabis' END)")
_OUTM_SEL = ("source||'|'||store_id outlet_key, store_name outlet_name, address, city, state, zip zip5, "
             "'' zip4, chain outlet_type, %s carriage, source sources, source source_list, "
             "CAST(lat AS DOUBLE) lat, CAST(lng AS DOUBLE) lng, phone, addr_valid, flag_basis, "
             "license_conflict, hoodie_outlet source_ref, is_chain" % _OUTM_SELL)


@app.get("/api/master/outlets")
def master_outlets_browse_ep():
    """Paged, filterable view of the outlet book (src_outlets — 220k+ accounts across every source). ?q=
    (name/addr/city), ?state=, ?page=, ?size=. Rows + total + a state facet for the workbench."""
    q = (request.args.get("q") or "").strip().lower()
    state = (request.args.get("state") or "").strip().upper()
    try:
        page = max(0, int(request.args.get("page", 0))); size = min(200, max(1, int(request.args.get("size", 50))))
    except ValueError:
        page, size = 0, 50
    if not q and not state and page == 0:       # the landing page — TTL-cache + warm it
        return jsonify(_ttl("outlets_resp_%d" % size, 90, lambda: _outlets_data("", "", 0, size)))
    return jsonify(_outlets_data(q, state, page, size))


def _outlets_data(q, state, page, size):
    import warehouse
    where, params = [], []
    if state:
        where.append("upper(CAST(state AS VARCHAR)) = ?"); params.append(state)
    if q:
        where.append("(lower(CAST(store_name AS VARCHAR)) LIKE ? OR lower(CAST(address AS VARCHAR)) LIKE ? OR lower(CAST(city AS VARCHAR)) LIKE ?)")
        qq = "%" + q + "%"; params += [qq, qq, qq]
    wsql = (" WHERE " + " AND ".join(where)) if where else ""
    fwhere = " WHERE " + " AND ".join(where + ["nullif(trim(CAST(state AS VARCHAR)),'') IS NOT NULL"])
    try:
        # count + state facet are rebuild-stable per filter → cache them; only the page rows are fetched live.
        total = _wq_cached("src_outlets", "SELECT count(*) c FROM t" + wsql, params)[0]["c"]
        rows = warehouse.query("src_outlets",
            "SELECT %s FROM t%s ORDER BY (nullif(trim(store_name),'') IS NULL), (lat IS NULL), state, city, "
            "store_name LIMIT %d OFFSET %d" % (_OUTM_SEL, wsql, size, page * size), params)
        facet = _wq_cached("src_outlets",
            "SELECT CAST(state AS VARCHAR) state, count(*) n FROM t%s GROUP BY 1 ORDER BY n DESC LIMIT 60" % fwhere, params)
    except Exception as e:
        return dict(ok=True, landed=False, error=str(e)[:140], outlets=[], total=0)
    return dict(ok=True, landed=True, total=total, page=page, size=size,
                outlets=rows, states=[{"state": r["state"], "n": r["n"]} for r in facet])

@app.get("/api/master/outlet")
def master_outlet_one_ep():
    """One outlet's full record (from src_outlets) for the detail panel — keyed source|store_id."""
    import warehouse
    key = (request.args.get("key") or "").strip()
    if not key or "|" not in key:
        return jsonify(ok=False, error="key required"), 400
    src, sid = key.split("|", 1)
    try:
        rows = warehouse.query("src_outlets", "SELECT %s FROM t WHERE source = ? AND CAST(store_id AS VARCHAR) = ? LIMIT 1"
                               % _OUTM_SEL, [src, sid])
    except Exception as e:
        return jsonify(ok=False, error=str(e)[:140]), 200
    return jsonify(ok=True, outlet=(rows[0] if rows else None))

# ── MDM workbench: CONFIRM / CONFORM — reconcile the same outlet across sources into Tier-1 truth ──
# The deterministic resolve merges on exact key; cross-VENDOR duplicates (AB vpid vs VTInfo, which lacks
# zip) slip through, so we FUZZY-match: cross-source pairs in the same ~1km cell, high name similarity,
# within ~0.5 mi. Accepting a candidate records that the two outlet_keys are the SAME real outlet.
_CONFIRM_SQL = (
    "WITH o AS (SELECT outlet_key, outlet_name, address, city, state, carriage, "
    " try_cast(lat AS DOUBLE) lat, try_cast(lng AS DOUBLE) lng, source_list "
    " FROM t WHERE try_cast(lat AS DOUBLE) IS NOT NULL) "
    "SELECT a.outlet_key ka, a.outlet_name na, a.address aa, a.city, a.state, CAST(a.carriage AS VARCHAR) ca, a.source_list sa, "
    " b.outlet_key kb, b.outlet_name nb, b.address ba, CAST(b.carriage AS VARCHAR) cb, b.source_list sb, "
    " round(jaro_winkler_similarity(lower(a.outlet_name),lower(b.outlet_name)),3) sim, "
    " round(sqrt(power((a.lat-b.lat)*69,2)+power((a.lng-b.lng)*54.6,2)),2) dist_mi "
    "FROM o a JOIN o b ON a.outlet_key<b.outlet_key AND round(a.lat,2)=round(b.lat,2) AND round(a.lng,2)=round(b.lng,2) "
    "WHERE a.source_list<>b.source_list "
    " AND jaro_winkler_similarity(lower(a.outlet_name),lower(b.outlet_name)) > ? "
    " AND sqrt(power((a.lat-b.lat)*69,2)+power((a.lng-b.lng)*54.6,2)) < 0.5 "
    "ORDER BY sim DESC, dist_mi ASC LIMIT ?")

@app.get("/api/master/confirm/candidates")
def master_confirm_candidates_ep():
    """Fuzzy cross-source outlet-match candidates (same place, different source) to review + confirm.
    ?min= name-similarity threshold (default .86), ?limit=. Already-decided pairs are dropped."""
    import warehouse
    try:
        mn = float(request.args.get("min", 0.86)); lim = min(500, int(request.args.get("limit", 150)))
    except ValueError:
        mn, lim = 0.86, 150
    try:
        rows = warehouse.query("dim_outlet_resolved", _CONFIRM_SQL, [mn, lim + 60])
    except Exception as e:
        return jsonify(ok=True, landed=False, error=str(e)[:160], candidates=[]), 200
    decided = load("outlet_confirms.json", {})
    out = []
    for r in rows:
        key = "|".join(sorted([str(r["ka"]), str(r["kb"])]))
        if key in decided:
            continue
        out.append(r)
        if len(out) >= lim:
            break
    return jsonify(ok=True, landed=True, count=len(out), candidates=out)

@app.post("/api/master/confirm")
def master_confirm_post_ep():
    """Record a confirm decision on a candidate pair: {a, b, decision:'same'|'different'}. 'same' = the
    two outlet_keys are one real outlet (Tier-1 confirmed); 'different' = not a match (won't resurface)."""
    b = request.get_json(silent=True) or {}
    ka, kb = (b.get("a") or "").strip(), (b.get("b") or "").strip()
    if not (ka and kb):
        return jsonify(ok=False, error="a + b required"), 400
    key = "|".join(sorted([ka, kb]))
    confirms = load("outlet_confirms.json", {})
    confirms[key] = {"a": ka, "b": kb, "decision": (b.get("decision") or "same"),
                     "at": int(time.time()), "by": _user_rec().get("code", "SYS")}
    _save_json("outlet_confirms.json", confirms)
    same = sum(1 for v in confirms.values() if v.get("decision") == "same")
    return jsonify(ok=True, decided=len(confirms), confirmed=same)

@app.get("/api/master/confirm/summary")
def master_confirm_summary_ep():
    """The reconciliation picture: outlet count, auto-confirmed (2+ sources), fuzzy candidates, decisions."""
    import warehouse
    out = {"outlets": None, "auto_confirmed": None, "candidates": None}
    try:
        out["outlets"] = warehouse.query("dim_outlet_resolved", "SELECT count(*) c FROM t")[0]["c"]
        out["auto_confirmed"] = warehouse.query("dim_outlet_resolved", "SELECT count(*) c FROM t WHERE sources>=2")[0]["c"]
        out["candidates"] = warehouse.query("dim_outlet_resolved", "SELECT count(*) c FROM (%s)" % _CONFIRM_SQL, [0.86, 100000])[0]["c"]
    except Exception as e:
        out["error"] = str(e)[:120]
    confirms = load("outlet_confirms.json", {})
    out["accepted"] = sum(1 for v in confirms.values() if v.get("decision") == "same")
    out["rejected"] = sum(1 for v in confirms.values() if v.get("decision") == "different")
    return jsonify(ok=True, **out)

# ── MDM workbench: browse the SKU/product catalog (dim_sku joined up the hierarchy) ──
def _sku_join():
    import warehouse
    u = lambda n: warehouse.uri(n).replace("'", "")
    return ("LEFT JOIN read_parquet('%s') it ON t.item_key=it.item_key "
            "LEFT JOIN read_parquet('%s') p ON it.product_key=p.product_key "
            "LEFT JOIN read_parquet('%s') b ON p.brand_key=b.brand_key"
            % (u("dim_item"), u("dim_product"), u("dim_brand")))
_SKU_SEL = ("t.sku_key, t.upc, t.pack, t.vintage, t.edition, t.sources, t.source_list, "
            "p.product_name, p.category, p.origin, p.abv, it.size_ml, b.brand, b.brand_group")
_CATALOG_COLS = ("sku_key, upc, pack, vintage, edition, sources, source_list, product_name, category, "
                 "origin, abv, size_ml, brand, brand_group")


@app.get("/api/master/skus")
def master_skus_ep():
    """Paged, filterable SKU catalog. Pages the FLAT pre-sorted `catalog_skus` table (built at rebuild) so the
    landing load is a cheap LIMIT/OFFSET, not a 1M-row 4-table join+sort per request. ?q= (brand/product/upc)."""
    q = (request.args.get("q") or "").strip().lower()
    try:
        page = max(0, int(request.args.get("page", 0))); size = min(200, max(1, int(request.args.get("size", 50))))
    except ValueError:
        page, size = 0, 50
    if not q and page == 0:                 # the landing page — TTL-cache + warm it
        return jsonify(_ttl("skus_resp_%d" % size, 90, lambda: _skus_data("", 0, size)))
    return jsonify(_skus_data(q, page, size))


def _skus_data(q, page, size):
    import warehouse
    where, params = [], []
    if q:
        where.append("(brand_lc LIKE ? OR product_lc LIKE ? OR upc_lc LIKE ?)")
        qq = "%" + q + "%"; params += [qq, qq, qq]
    wsql = (" WHERE " + " AND ".join(where)) if where else ""
    try:                                    # flat catalog: pre-sorted, so no ORDER BY needed on read
        total = _wq_cached("catalog_skus", "SELECT count(*) c FROM t" + wsql, params)[0]["c"]
        rows = warehouse.query("catalog_skus", "SELECT %s FROM t%s LIMIT %d OFFSET %d"
                               % (_CATALOG_COLS, wsql, size, page * size), params)
        return dict(ok=True, landed=True, total=total, page=page, size=size, skus=rows)
    except Exception:
        pass
    # Fallback: the flat table isn't built yet — do the live join (slow, but correct).
    jn = _sku_join()
    w2 = []
    if q:
        w2.append("(lower(CAST(b.brand AS VARCHAR)) LIKE ? OR lower(CAST(p.product_name AS VARCHAR)) LIKE ? "
                  "OR lower(CAST(t.upc AS VARCHAR)) LIKE ?)")
    w2sql = (" WHERE " + " AND ".join(w2)) if w2 else ""
    try:
        total = warehouse.query("dim_sku", "SELECT count(*) c FROM t %s%s" % (jn, w2sql), params)[0]["c"]
        rows = warehouse.query("dim_sku", "SELECT %s FROM t %s%s "
                               "ORDER BY (CASE WHEN b.brand IS NULL OR CAST(b.brand AS VARCHAR)='' THEN 1 ELSE 0 END), "
                               "b.brand, p.product_name LIMIT %d OFFSET %d"
                               % (_SKU_SEL, jn, w2sql, size, page * size), params)
    except Exception as e:
        return dict(ok=True, landed=False, error=str(e)[:140], skus=[], total=0)
    return dict(ok=True, landed=True, total=total, page=page, size=size, skus=rows)

@app.get("/api/master/sku")
def master_sku_one_ep():
    import warehouse
    key = (request.args.get("key") or "").strip()
    if not key:
        return jsonify(ok=False, error="key required"), 400
    try:
        rows = warehouse.query("dim_sku", "SELECT %s, t.master_created_at, t.updated_by FROM t %s WHERE t.sku_key = ? LIMIT 1"
                               % (_SKU_SEL, _sku_join()), [key])
    except Exception as e:
        return jsonify(ok=False, error=str(e)[:140]), 200
    return jsonify(ok=True, sku=(rows[0] if rows else None))

@app.get("/api/master/counts")
def master_counts_ep():
    """Row counts of each master table — for the workbench overview + catalog header."""
    out = {}
    for t in ("dim_brand", "dim_product", "dim_item", "dim_sku", "dim_supplier", "dim_outlet_resolved",
              "fact_inventory", "fact_pricing"):
        try:
            out[t] = _wq_cached(t, "SELECT count(*) c FROM t")[0]["c"]   # rebuild-stable → cached
        except Exception:
            out[t] = None
    return jsonify(ok=True, counts=out)

@app.get("/api/master/outlets/geo")
def master_outlets_geo_ep():
    """Light geocoded points from the resolved outlet master for the workbench map. Optional ?bbox=
    west,south,east,north (viewport) + ?q= / ?state=. Ships {outlet_key,lat,lng,name,state} only."""
    bb = request.args.get("bbox", "")
    st = (request.args.get("state") or "").strip().upper()
    q = (request.args.get("q") or "").strip().lower()
    if not (bb or st or q):                     # default full map — TTL-cache + warm
        return jsonify(_ttl("outlets_geo", 90, lambda: _outlets_geo_data("", "", "")))
    return jsonify(_outlets_geo_data(bb, st, q))


def _outlets_geo_data(bb, st, q):
    import warehouse
    # Exclude Null Island (0,0) — a sentinel some scrapers write instead of NULL. Measured live on
    # ubereats/postmates; left in, the coverage map draws pins in the Atlantic and the auto-fit
    # stretches the viewport across the ocean, which is part of why those sources "render wrong".
    where = ["try_cast(lat AS DOUBLE) IS NOT NULL", "try_cast(lng AS DOUBLE) IS NOT NULL",
             "NOT (try_cast(lat AS DOUBLE) = 0 AND try_cast(lng AS DOUBLE) = 0)"]
    params = []
    if bb:
        try:
            w, s, e, n = (float(x) for x in bb.split(","))
            where.append("try_cast(lat AS DOUBLE) BETWEEN ? AND ? AND try_cast(lng AS DOUBLE) BETWEEN ? AND ?")
            params += [s, n, w, e]
        except (ValueError, TypeError):
            pass
    if st:
        where.append("upper(CAST(state AS VARCHAR)) = ?"); params.append(st)
    if q:
        where.append("(lower(CAST(store_name AS VARCHAR)) LIKE ? OR lower(CAST(city AS VARCHAR)) LIKE ?)")
        qq = "%" + q + "%"; params += [qq, qq]
    wsql = " WHERE " + " AND ".join(where)
    cap = 20000
    # WHY THIS IS NOT A BARE `LIMIT 20000` ANY MORE.
    # It used to be, with no ORDER BY — so against a 1.76M-row src_outlets DuckDB returned whichever
    # 20k rows the first Parquet parts happened to hold. Parts are grouped by source, so the map
    # showed a handful of sources at full density and silently omitted the rest: that is why UberEats
    # "doesn't render correctly" in coverage. The count shipped was len(rows), i.e. the cap itself, so
    # a truncated map reported a complete-looking footprint and nothing anywhere said 1.74M points
    # had been dropped.
    #
    # Two fixes: report the TRUE total, and make the sample uniform across the whole matched set
    # instead of positional. hash(outlet_key) % m is deterministic (same viewport → same points, so
    # the map doesn't shimmer on refresh) and independent of storage order, so every source lands in
    # the sample in proportion to its real size.
    try:
        total = warehouse.query("src_outlets", "SELECT count(*) n FROM t%s" % wsql, params)[0]["n"]
    except Exception as e:
        return dict(ok=True, landed=False, error=str(e)[:140], points=[])
    stride = 1 if total <= cap else -(-total // cap)          # ceil
    samp = ("" if stride <= 1 else
            " AND (hash(CAST(source AS VARCHAR)||'|'||CAST(store_id AS VARCHAR)) %% %d) = 0" % stride)
    try:
        rows = warehouse.query("src_outlets",
            "SELECT source||'|'||store_id outlet_key, try_cast(lat AS DOUBLE) AS lat, try_cast(lng AS DOUBLE) AS lng, "
            "CAST(store_name AS VARCHAR) AS \"name\", CAST(state AS VARCHAR) AS state, "
            "CAST(source AS VARCHAR) AS source FROM t%s%s LIMIT %d" % (wsql, samp, cap), params)
        # Per-source truth, so a source that is under-drawn on the map is visibly under-drawn in the
        # numbers rather than being mistaken for a source we don't hold.
        by_src = warehouse.query("src_outlets",
            "SELECT CAST(source AS VARCHAR) AS source, count(*) n FROM t%s GROUP BY 1 ORDER BY 2 DESC" % wsql,
            params)
    except Exception as e:
        return dict(ok=True, landed=False, error=str(e)[:140], points=[])
    return dict(ok=True, landed=True, count=len(rows), total=total,
                sampled=stride > 1, stride=stride, sources=by_src, points=rows)

@app.post("/api/master/fact/apply")
def fact_apply_ep():
    """Materialize fact_<fact> (inventory|pricing) from its mappings — one DuckDB pass per source over
    Parquet → compute outlet_key + sku/brand key inline → group by (outlet, product, date)."""
    try:
        import master_facts
        body = request.get_json(silent=True) or {}
        fact = _fact(body.get("fact") or request.args.get("fact", "inventory"))
        fields = _fact_schema(fact)
        maps = _resolve_dict_refs(_fact_maps_effective(fact))    # expand dict refs to inline entries
        who = _user_rec().get("code", "SYS")
        res = master_facts.build(fact, fields, maps, built_by=who, built_at=int(time.time()),
                                 log=lambda mm: app.logger.info("FACT[%s] %s", fact, mm))
        return jsonify(ok=True, fact=fact, built_by=who, **res)
    except Exception as e:
        app.logger.exception("fact apply failed")
        return jsonify(ok=False, error=str(e)[:200]), 200

# ── "needs dictionary" flags — map a field now, flag it to go back and build its dictionary later ──
@app.get("/api/dict/needs")
def dict_needs_get():
    """Fields flagged as needing a data dictionary. ?dataset= for one source's flags, else all."""
    ds = (request.args.get("dataset") or "").strip()
    needs = load("dict_needs.json", {})
    items = list(needs.values())
    if ds:
        items = [n for n in items if n.get("dataset") == ds]
    return jsonify(ok=True, dataset=ds or None, count=len(items), needs=items)

@app.post("/api/dict/needs")
def dict_needs_post():
    """Toggle a flag: {dataset, field, target, on}. Keyed target|dataset|field."""
    b = request.get_json(silent=True) or {}
    ds = (b.get("dataset") or "").strip(); field = (b.get("field") or "").strip()
    if not (ds and field):
        return jsonify(ok=False, error="dataset + field required"), 400
    target = b.get("target") or "product"
    key = "%s|%s|%s" % (target, ds, field)
    needs = load("dict_needs.json", {})
    if b.get("on", True) and key not in needs:
        needs[key] = {"dataset": ds, "field": field, "target": target,
                      "at": int(time.time()), "by": _user_rec().get("code", "SYS")}
    elif not b.get("on", True):
        needs.pop(key, None)
    _save_json("dict_needs.json", needs)
    return jsonify(ok=True, count=len(needs), on=key in needs)

@app.get("/api/datasets/download")
def dataset_download():
    """Stream the COMPLETE pulled dataset (all rows, not the UI sample) as CSV or JSON."""
    did = (request.args.get("dataset") or "").strip()
    fmt = (request.args.get("format") or "csv").lower()
    full = load_full(did)                                   # full rows if we kept them (FL/COLA)
    if full:
        header, rows = full.get("header") or [], full.get("rows") or []
    elif did in DATASETS:                                   # else the in-memory set (chains: already complete)
        ds = DATASETS[did]; header, rows = ds.get("header") or [], ds.get("rows") or []
    else:
        return jsonify(error="unknown dataset: " + did), 404
    if fmt == "json":
        body = json.dumps([dict(zip(header, r)) for r in rows]); mime, ext = "application/json", "json"
    else:
        buf = io.StringIO(); w = csv.writer(buf); w.writerow(header); w.writerows(rows)
        body, mime, ext = buf.getvalue(), "text/csv", "csv"
    return Response(body, mimetype=mime + "; charset=utf-8",
                    headers={"Content-Disposition": 'attachment; filename="%s.%s"' % (did, ext)})

@app.get("/api/runs")
def runs():
    return jsonify(RUNS[:200])


# ─────────────────────────────────────────────────────────────────────────────
# Registry pipeline — EXECUTE the canonical source_registry sources IN THE APP.
#
# The point (owner, 2026-07-22): data operations should run in the deployed app, which holds the
# secrets ONCE in its Fly config — not in a re-provisioned sandbox or re-added to CI every time. So
# these endpoints run run_sources.run_one server-side using THIS process's environment (its secrets,
# its warehouse creds, its proxy), land to the shared source_runs ledger, and report expected-vs-landed
# COVERAGE. Behind the OIDC gate like every /api/* route. This is distinct from /api/run (the older
# connId pull surface); it drives the 31-source canonical registry that run_sources + health use.
# ─────────────────────────────────────────────────────────────────────────────
_REG_JOBS = {}


@app.get("/api/registry/sources")
def registry_sources_ep():
    """The canonical source list + each source's expected-vs-landed coverage + which creds are absent
    HERE (so you can see what this executor can actually run). Reads the app's own warehouse/ledger."""
    import source_registry as reg
    import warehouse
    try:
        import coverage as cov
    except Exception:
        cov = None
    def _cost_class(s):
        # what a source actually needs: 'free' (direct HTTP / API / sitemap — $0), 'free-api-key' (a free
        # key), or 'anti-bot' (the ONLY class that even tempts a paid proxy — and it should try the free
        # local-browser / mobile-UA path first).
        note = (s.get("note") or "").lower()
        if s["klass"] == "mac" or any(w in note for w in
                ("perimeterx", "imperva", "incapsula", "datadome", "cloudflare", "waf", "akamai", "forter")):
            return "anti-bot"
        return "free-api-key" if s["klass"] == "creds" else "free"
    out = []
    for s in reg.SOURCES:
        try:
            cv = cov.assess(s) if cov else None
        except Exception:
            cv = None
        out.append({"id": s["id"], "label": s["label"], "klass": s["klass"], "cadence": s.get("cadence"),
                    "enabled": bool(s.get("enabled")), "tables": s.get("tables"),
                    "cost_class": _cost_class(s),
                    "requires": s.get("requires") or [],
                    "missing_creds": [e for e in (s.get("requires") or []) if not os.environ.get(e)],
                    "coverage": cv})
    n_free = sum(1 for x in out if x["cost_class"] != "anti-bot" and x["enabled"])
    return jsonify(ok=True, count=len(out), free_or_keyed=n_free, remote=warehouse.remote(), sources=out)


@app.get("/api/registry/coverage")
def registry_coverage_ep():
    """Expected-vs-landed store/item coverage for every source, from the app's coverage_log — the honest
    'did the last run actually cover the catalog' read that a cumulative row-count can't give."""
    import source_registry as reg
    import coverage as cov
    out = {}
    for s in reg.SOURCES:
        try:
            out[s["id"]] = cov.assess(s)
        except Exception as e:
            out[s["id"]] = {"error": str(e)[:120]}
    return jsonify(ok=True, coverage=out)


def _reg_run_job(jid, ids):
    import run_sources as rs
    import source_registry as reg
    job = _REG_JOBS[jid]

    def _log(*a):
        job["log"].append(" ".join(str(x) for x in a))
        del job["log"][:-200]
    for i in ids:
        s = reg.by_id(i)
        if not s:
            continue
        job["current"] = i
        try:
            rec = rs.run_one(s, log=_log)
            rs._land_runs([rec], log=lambda *a: None)     # shared ledger → health + /api/runs see it
        except Exception as e:
            rec = {"source": i, "status": "failed", "error": str(e)[:300]}
        job["records"].append(rec)
    job["current"] = None
    job["status"] = "done"
    job["finishedAt"] = int(time.time() * 1000)


@app.post("/api/registry/run")
def registry_run_ep():
    """Run registry source(s) SERVER-SIDE with the app's env/secrets. Body: {id} or {ids:[...]} or
    {cadence}; {include_mac:true} to include the anti-bot browser sources (default: headless/creds only,
    since a server without a warmed browser can't run them); {sync:true} to block for a single source
    (else a background job → poll /api/registry/run/progress?id=)."""
    import source_registry as reg
    import run_sources as rs
    body = request.get_json(silent=True) or {}
    ids = body.get("ids") or ([body["id"]] if body.get("id") else None)
    if not ids:
        cad = body.get("cadence")
        ok_klass = (lambda k: True) if body.get("include_mac") else (lambda k: k != "mac")
        ids = [s["id"] for s in reg.SOURCES if s.get("enabled") and ok_klass(s["klass"])
               and (not cad or s.get("cadence") == cad or cad == "all")]
    ids = [i for i in ids if reg.by_id(i)]
    if not ids:
        return jsonify(ok=False, error="no matching enabled sources"), 400
    if body.get("sync") and len(ids) == 1:
        rec = rs.run_one(reg.by_id(ids[0]))
        rs._land_runs([rec], log=lambda *a: None)
        return jsonify(ok=True, sync=True, record=rec)
    jid = "reg-%d" % int(time.time() * 1000)
    _REG_JOBS[jid] = {"id": jid, "ids": ids, "status": "running", "current": None,
                      "startedAt": int(time.time() * 1000), "finishedAt": None, "log": [], "records": []}
    for old in sorted(_REG_JOBS, key=lambda k: _REG_JOBS[k]["startedAt"])[:-50]:
        _REG_JOBS.pop(old, None)                              # keep the last ~50 jobs
    threading.Thread(target=_reg_run_job, args=(jid, ids), daemon=True).start()
    return jsonify(ok=True, jobId=jid, running=ids), 202


@app.get("/api/registry/run/progress")
def registry_run_progress_ep():
    """Poll a background registry run: status, the source in flight, run records so far (with coverage),
    and the recent log tail."""
    jid = (request.args.get("id") or "").strip()
    job = _REG_JOBS.get(jid)
    if not job:
        return jsonify(ok=False, error="unknown job"), 404
    return jsonify(ok=True, id=job["id"], status=job["status"], current=job["current"],
                   startedAt=job["startedAt"], finishedAt=job["finishedAt"],
                   records=job["records"], log=job["log"][-60:])


# ─────────────────────────────────────────────────────────────────────────────
# In-app SCHEDULER — the app runs due sources on their cadence, itself.
#
# So capture doesn't depend on GitHub Actions (which need secrets re-added and were dying mid-run) or a
# human/CLI trigger: a daemon thread in the single gunicorn worker ticks every SCHEDULER_INTERVAL, asks
# run_sources which sources are DUE (past their interval_h in the shared ledger), and runs them with the
# app's own env/secrets — landing to the same ledger + coverage the rest of the pipeline reads. It takes
# the SAME machine-global file lock run_sources' --due uses, so an in-app tick and a CLI/CI pass dedupe
# instead of stacking. Default headless/creds only (a server has no warmed browser); SCHEDULER_MAC=1
# opts the anti-bot browser sources in (only where the image actually has Chrome + a proxy).
#
# Env: SCHEDULER=1 auto-starts it at boot; SCHEDULER_INTERVAL (s, default 1800); SCHEDULER_MAC=1.
# Runtime toggle without redeploy: POST /api/registry/scheduler {enabled, interval_s, run_now}.
# ─────────────────────────────────────────────────────────────────────────────
def _truthy(v):
    return str(v or "").strip().lower() in ("1", "true", "yes", "on")


# The cloud RUNNER process (fly.toml [processes].runner — 8gb, off the serving box) self-drives the
# dispatcher so the whole pipeline runs in-cloud on cadence: NOTHING LOCAL. It runs headless/creds sources
# AND the derived builds (the single dim_* writer), which used to run only on the Mac --due tick. The serving
# 'app' process never schedules. SCHEDULER=0/1 still overrides explicitly (e.g. to pause the runner).
_RUNNER = os.environ.get("FLY_PROCESS_GROUP") == "runner"
_SCHED_ENV = os.environ.get("SCHEDULER")
_SCHED = {"enabled": _truthy(_SCHED_ENV) if _SCHED_ENV is not None else _RUNNER,
          "interval_s": max(60, int(os.environ.get("SCHEDULER_INTERVAL", "1800") or 1800)),
          "mac": _truthy(os.environ.get("SCHEDULER_MAC")),
          # run derived builds in the tick — on the build-writer host only (the runner), so builds move off
          # the Mac. SCHEDULER_BUILDS=1 forces it on elsewhere; the shared source_runs ledger keeps a build
          # from double-running across hosts during the Mac→cloud cutover.
          "builds": _RUNNER or _truthy(os.environ.get("SCHEDULER_BUILDS")),
          "thread_started": False, "running": False, "ticks": 0,
          "last_tick": None, "next_tick": None, "last_due": [], "last_skip": None,
          "runs": [], "error": None}
_SCHED_LOCK = threading.Lock()


def _sched_tick():
    """One dispatcher pass: run every DUE source once, land it, record the outcome. No-ops cleanly when
    nothing is due or another pass holds the machine-global lock."""
    import run_sources as rs
    with _SCHED_LOCK:
        if _SCHED["running"]:
            return
        _SCHED["running"] = True
    lock = None
    try:
        lock = rs._acquire_lock()                    # dedupe with any CLI/CI --due pass on this host
        if not lock:
            _SCHED["last_skip"] = "lock held by another pass"
            return
        try:
            due = rs.due_sources()
        except Exception as e:
            _SCHED["error"] = "due_sources: %s" % str(e)[:160]
            due = []
        if not _SCHED["mac"]:
            due = [s for s in due if s["klass"] != "mac"]
        _SCHED["last_due"] = [s["id"] for s in due]
        # COST GUARD: an unattended tick forces per-GB residential OFF (flat ISP pool only) so the
        # scheduler can never run up a metered proxy tab. Opt in deliberately with SCHEDULER_PAYGO=1.
        overlay = {} if _truthy(os.environ.get("SCHEDULER_PAYGO")) else {"RESI_ISP_ONLY": "1"}
        for s in due:
            try:
                rec = rs.run_one(s, extra_env=overlay)
                rs._land_runs([rec], log=lambda *a: None)
                _SCHED["runs"].insert(0, {"source": rec.get("source"), "status": rec.get("status"),
                                          "delta": rec.get("delta"), "cov_items": rec.get("cov_items"),
                                          "cov_stores": rec.get("cov_stores"), "at": int(time.time())})
                del _SCHED["runs"][80:]
            except Exception as e:
                _SCHED["error"] = "%s: %s" % (s["id"], str(e)[:160])
        # Derived master builds (dim_sku chain, master_quality, item_identity, the canon head-to-head) AFTER
        # the source landings that trigger them — on the build-writer host only, so the cloud runner owns the
        # builds that used to run on the Mac --due tick. Same run_one + ledger treatment as a source.
        if _SCHED["builds"]:
            try:
                for b in rs.due_builds():
                    rec = rs.run_one(b)
                    rs._land_runs([rec], log=lambda *a: None)
                    _SCHED["runs"].insert(0, {"source": rec.get("source"), "status": rec.get("status"),
                                              "at": int(time.time())})
                del _SCHED["runs"][80:]
            except Exception as e:
                _SCHED["error"] = "builds: %s" % str(e)[:160]
        _SCHED["ticks"] += 1
        _SCHED["last_tick"] = int(time.time())
    finally:
        if lock:
            try:
                lock.close()                          # release the machine-global lock
            except Exception:
                pass
        _SCHED["running"] = False


def _scheduler_loop():
    while True:
        try:
            if _SCHED["enabled"]:
                _sched_tick()
        except Exception as e:
            _SCHED["error"] = "loop: %s" % str(e)[:160]
        _SCHED["next_tick"] = int(time.time()) + _SCHED["interval_s"]
        time.sleep(_SCHED["interval_s"])


def _ensure_scheduler_thread():
    """Start the single scheduler daemon (idempotent). Started at import when SCHEDULER=1, or on first
    runtime enable via the API — so you can turn capture on without a redeploy."""
    with _SCHED_LOCK:
        if _SCHED["thread_started"]:
            return
        _SCHED["thread_started"] = True
    threading.Thread(target=_scheduler_loop, name="hoodie-scheduler", daemon=True).start()


@app.get("/api/registry/scheduler")
def registry_scheduler_status_ep():
    import run_sources as rs
    try:
        due = [s["id"] for s in rs.due_sources()]
    except Exception:
        due = []
    try:
        import resi
        proxy = {"usage": resi.usage(), "scheduler_paygo": _truthy(os.environ.get("SCHEDULER_PAYGO"))}
    except Exception:
        proxy = None
    return jsonify(ok=True, enabled=_SCHED["enabled"], interval_s=_SCHED["interval_s"], mac=_SCHED["mac"],
                   running=_SCHED["running"], ticks=_SCHED["ticks"], last_tick=_SCHED["last_tick"],
                   next_tick=_SCHED["next_tick"], thread_started=_SCHED["thread_started"],
                   due_now=due, last_due=_SCHED["last_due"], last_skip=_SCHED["last_skip"],
                   proxy=proxy, recent=_SCHED["runs"][:20], error=_SCHED["error"])


@app.get("/api/registry/proxy")
def registry_proxy_status_ep():
    """Residential-proxy COST view: this month's per-GB session usage + the guard state (flat ISP pool
    size, isp_only, monthly cap, whether per-GB is currently allowed). The ISP pool bills flat; only the
    per-GB tier runs up a tab, and this is where you see/limit it."""
    import resi
    return jsonify(ok=True, usage=resi.usage(),
                   env={"FETCH_POLICY": resi.fetch_policy(), "RESI_ISP_ONLY": resi.isp_only(),
                        "RESI_MONTHLY_MAX_SESSIONS": resi.monthly_cap(),
                        "SCHEDULER_PAYGO": _truthy(os.environ.get("SCHEDULER_PAYGO"))})


@app.post("/api/registry/proxy")
def registry_proxy_control_ep():
    """Set the cost dial at runtime (in-memory; the env vars are the durable default). Body:
    {policy:'free'|'flat'|'paid', isp_only:bool, max_sessions:int, reset:bool}. policy='free' spends
    nothing (direct/browser only); 'flat' adds the fixed ISP pool; 'paid' opts into the per-GB tier."""
    import resi
    b = request.get_json(silent=True) or {}
    if b.get("policy") in ("free", "flat", "paid"):
        os.environ["FETCH_POLICY"] = b["policy"]
        os.environ.pop("FETCH_FREE_ONLY", None)
    if "isp_only" in b:
        os.environ["RESI_ISP_ONLY"] = "1" if b["isp_only"] else "0"
    if "max_sessions" in b:
        try:
            os.environ["RESI_MONTHLY_MAX_SESSIONS"] = str(max(0, int(b["max_sessions"])))
        except (ValueError, TypeError):
            pass
    if b.get("reset"):
        resi.reset_month()
    return jsonify(ok=True, usage=resi.usage(),
                   env={"FETCH_POLICY": resi.fetch_policy(), "RESI_ISP_ONLY": resi.isp_only(),
                        "RESI_MONTHLY_MAX_SESSIONS": resi.monthly_cap()})


@app.post("/api/registry/scheduler")
def registry_scheduler_control_ep():
    """Toggle/adjust the scheduler at runtime (in-memory; SCHEDULER env is the durable default). Body:
    {enabled:bool, interval_s:int, mac:bool, run_now:bool}. run_now fires one tick immediately."""
    b = request.get_json(silent=True) or {}
    if "enabled" in b:
        _SCHED["enabled"] = bool(b["enabled"])
    if "mac" in b:
        _SCHED["mac"] = bool(b["mac"])
    if "interval_s" in b:
        try:
            _SCHED["interval_s"] = max(60, int(b["interval_s"]))
        except (ValueError, TypeError):
            pass
    if _SCHED["enabled"]:
        _ensure_scheduler_thread()
    if b.get("run_now"):
        threading.Thread(target=_sched_tick, name="hoodie-scheduler-kick", daemon=True).start()
    return jsonify(ok=True, enabled=_SCHED["enabled"], interval_s=_SCHED["interval_s"], mac=_SCHED["mac"],
                   thread_started=_SCHED["thread_started"], kicked=bool(b.get("run_now")))


if _SCHED["enabled"]:                                 # auto-start under gunicorn when SCHEDULER=1
    _ensure_scheduler_thread()


# ─────────────────────────────────────────────────────────────────────────────
# Service desk — data-quality reports (raised from the suite, tracked in the CRM)
# ─────────────────────────────────────────────────────────────────────────────
_SERVICE_STATES = ("open", "acknowledged", "reported", "resolved")

def save_service():
    name = "service_reports.json"
    if STATE_BUCKET:
        try:
            _s3().put_object(Bucket=STATE_BUCKET, Key=_key(name),
                             Body=json.dumps(SERVICE).encode("utf-8"),
                             ContentType="application/json")
        except Exception as e:
            app.logger.warning("S3 save %s failed: %s", name, e)
    else:
        try:
            json.dump(SERVICE, open(os.path.join(STATE_DIR, name), "w"))
        except Exception as e:
            app.logger.warning("save %s failed: %s", name, e)

@app.get("/api/service/reports")
def service_list():
    st = (request.args.get("status") or "").strip().lower()
    items = SERVICE if st not in _SERVICE_STATES else [r for r in SERVICE if r.get("status") == st]
    counts = {s: sum(1 for r in SERVICE if r.get("status") == s) for s in _SERVICE_STATES}
    return jsonify(ok=True, reports=items[:500], total=len(SERVICE), counts=counts)

@app.post("/api/service/reports")
def service_create():
    b = request.get_json(force=True, silent=True) or {}
    now = int(time.time() * 1000)
    by = (b.get("reporter") or "field").strip()[:120]
    rec = {
        "id": "SR-%d" % now,
        "kind": (b.get("kind") or "data-quality").strip()[:40],
        "connId": (b.get("connId") or "").strip()[:80],
        "source": (b.get("source") or b.get("connId") or "—").strip()[:160],
        "dataset": (b.get("dataset") or "").strip()[:160],
        "severity": (b.get("severity") or "degraded").strip()[:40],
        "warnings": [str(w)[:300] for w in (b.get("warnings") or [])][:20],
        "note": (b.get("note") or "").strip()[:2000],
        "url": (b.get("url") or "").strip()[:400],
        "reporter": by,
        "status": "open",
        "createdAt": now, "updatedAt": now,
        "history": [{"at": now, "status": "open", "by": by}],
    }
    SERVICE.insert(0, rec); del SERVICE[1000:]; save_service()
    return jsonify(ok=True, report=rec)

@app.patch("/api/service/reports/<rid>")
def service_update(rid):
    b = request.get_json(force=True, silent=True) or {}
    rec = next((r for r in SERVICE if r.get("id") == rid), None)
    if not rec:
        return jsonify(ok=False, error="not found"), 404
    now = int(time.time() * 1000)
    st = (b.get("status") or "").strip().lower()
    if st in _SERVICE_STATES:
        rec["status"] = st
        rec.setdefault("history", []).append({"at": now, "status": st, "by": (b.get("by") or "ops")})
    if "note" in b:
        rec["note"] = (b.get("note") or "").strip()[:2000]
    rec["updatedAt"] = now
    save_service()
    return jsonify(ok=True, report=rec)


# ─────────────────────────────────────────────────────────────────────────────
# Per-account planograms — shelf facings that persist (Planogram tab in the CRM)
# ─────────────────────────────────────────────────────────────────────────────
def save_planogram():
    name = "planograms.json"
    if STATE_BUCKET:
        try:
            _s3().put_object(Bucket=STATE_BUCKET, Key=_key(name),
                             Body=json.dumps(PLANOGRAM).encode("utf-8"),
                             ContentType="application/json")
        except Exception as e:
            app.logger.warning("S3 save %s failed: %s", name, e)
    else:
        try:
            json.dump(PLANOGRAM, open(os.path.join(STATE_DIR, name), "w"))
        except Exception as e:
            app.logger.warning("save %s failed: %s", name, e)

def _plano_summary(rec):
    rows = rec.get("rows") or []
    ours = sum((r.get("facings") or 0) for r in rows)
    bench = sum((r.get("benchmark") or 0) for r in rows)
    return {"accountId": rec.get("accountId"), "accountName": rec.get("accountName"),
            "skus": len(rows), "facings": ours, "benchmark": bench, "gap": ours - bench,
            "source": rec.get("source") or "manual", "updatedAt": rec.get("updatedAt")}

@app.get("/api/planogram/accounts")
def planogram_list():
    return jsonify(ok=True, planograms=[_plano_summary(r) for r in PLANOGRAM.values()])

@app.get("/api/planogram/accounts/<aid>")
def planogram_get(aid):
    rec = PLANOGRAM.get(aid) or {"accountId": aid, "accountName": None, "rows": [],
                                  "notes": "", "source": "manual", "updatedAt": None}
    return jsonify(ok=True, planogram=rec)

@app.put("/api/planogram/accounts/<aid>")
def planogram_put(aid):
    b = request.get_json(force=True, silent=True) or {}
    now = int(time.time() * 1000)
    rows = []
    for r in (b.get("rows") or [])[:200]:
        rows.append({
            "item": (r.get("item") or "").strip()[:160],
            "facings": max(0, int(r.get("facings") or 0)),
            "benchmark": max(0, int(r.get("benchmark") or 0)),
        })
    PLANOGRAM[aid] = {
        "accountId": aid,
        "accountName": (b.get("accountName") or (PLANOGRAM.get(aid) or {}).get("accountName")),
        "rows": rows,
        "notes": (b.get("notes") or "").strip()[:2000],
        "source": (b.get("source") or "manual").strip()[:20],
        "updatedAt": now,
    }
    save_planogram()
    return jsonify(ok=True, planogram=PLANOGRAM[aid])


# ─────────────────────────────────────────────────────────────────────────────
# Admin console — usage reporting, feature flags, access, ops snapshot
# ─────────────────────────────────────────────────────────────────────────────
def _current_user():
    try:
        return session.get("email") or None
    except Exception:
        return None

def _managed_allowlist():
    """UI-managed access list (durable admin_allowlist.json — same STATE_BUCKET path as admin_flags, so it
    survives Fly redeploys). auth_gate merges this with the ALLOWED_EMAILS env bootstrap + the admins."""
    d = load("admin_allowlist.json", {"emails": []})
    return d.get("emails", []) if isinstance(d, dict) else []

auth_gate.set_allowlist_provider(_managed_allowlist)   # effective allowlist = env bootstrap + this + admins

import re as _re_admin
_EMAIL_RE = _re_admin.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

def _require_admin():
    """Guard for /api/admin/*. Returns a 403 response for non-admins, else None to proceed. When the gate
    is OFF (local dev / pre-secrets) everything is open, matching auth_gate's 'unconfigured -> open'."""
    if not auth_gate.enabled():
        return None
    if not auth_gate.is_admin(_current_user()):
        return jsonify(ok=False, error="admin access required"), 403
    return None

def _save_json(name, obj):
    if STATE_BUCKET:
        try:
            _s3().put_object(Bucket=STATE_BUCKET, Key=_key(name),
                             Body=json.dumps(obj).encode("utf-8"), ContentType="application/json")
        except Exception as e:
            app.logger.warning("S3 save %s failed: %s", name, e)
    else:
        try:
            json.dump(obj, open(os.path.join(STATE_DIR, name), "w"))
        except Exception as e:
            app.logger.warning("save %s failed: %s", name, e)

@app.post("/api/usage/event")
def usage_event():
    b = request.get_json(force=True, silent=True) or {}
    app_id = (b.get("app") or "").strip()[:60]
    if not app_id:
        return jsonify(ok=False, error="app required"), 400
    USAGE.append({"app": app_id, "user": (_current_user() or b.get("user") or "local")[:120],
                  "ts": int(time.time() * 1000)})
    del USAGE[:-20000]                     # keep a rolling window
    _save_json("usage.json", USAGE)
    return jsonify(ok=True)

@app.get("/api/usage/summary")
def usage_summary():
    try:
        days = max(1, min(365, int(request.args.get("days", "30"))))
    except Exception:
        days = 30
    cutoff = int(time.time() * 1000) - days * 86400000
    ev = [e for e in USAGE if (e.get("ts") or 0) >= cutoff]
    by_app, by_user, by_day = {}, {}, {}
    for e in ev:
        by_app[e["app"]] = by_app.get(e["app"], 0) + 1
        by_user[e["user"]] = by_user.get(e["user"], 0) + 1
        day = datetime.datetime.utcfromtimestamp((e.get("ts") or 0) / 1000).strftime("%Y-%m-%d")
        by_day[day] = by_day.get(day, 0) + 1
    top = lambda d: sorted(([k, v] for k, v in d.items()), key=lambda x: -x[1])
    return jsonify(ok=True, days=days, totalOpens=len(ev), totalAllTime=len(USAGE),
                   activeUsers=len(by_user), byApp=top(by_app), byUser=top(by_user),
                   byDay=sorted(([k, v] for k, v in by_day.items())))

@app.get("/api/admin/flags")
def admin_flags_get():
    g = _require_admin()
    if g: return g
    return jsonify(ok=True, flags=FLAGS)

@app.put("/api/admin/flags")
def admin_flags_put():
    g = _require_admin()
    if g: return g
    b = request.get_json(force=True, silent=True) or {}
    if isinstance(b.get("apps"), dict):
        FLAGS["apps"] = b["apps"]
    if isinstance(b.get("features"), dict):
        FLAGS["features"] = b["features"]
    _save_json("admin_flags.json", FLAGS)
    return jsonify(ok=True, flags=FLAGS)

@app.get("/api/admin/access")
def admin_access():
    g = _require_admin()
    if g: return g
    env_boot = sorted({e.strip().lower() for e in os.environ.get("ALLOWED_EMAILS", "").split(",") if e.strip()})
    managed = sorted({str(e).strip().lower() for e in _managed_allowlist() if str(e).strip()})
    admins = sorted(auth_gate._admin_emails())
    seen = sorted({e.get("user") for e in USAGE if e.get("user")})
    try:
        allow = sorted(auth_gate._allowed_emails())
    except Exception:
        allow = []
    return jsonify(ok=True, gated=auth_gate.enabled(), currentUser=_current_user(),
                   isAdmin=(auth_gate.is_admin(_current_user()) or not auth_gate.enabled()),
                   admins=admins, bootstrap=env_boot, managed=managed,
                   allowlist=allow, seenUsers=seen)

@app.post("/api/admin/users")
def admin_users_add():
    """Add an email to the UI-managed allowlist (durable). Admin-only. Bootstrap (env) + admins are managed
    elsewhere and are always allowed regardless."""
    g = _require_admin()
    if g: return g
    b = request.get_json(force=True, silent=True) or {}
    email = (b.get("email") or "").strip().lower()
    if not _EMAIL_RE.match(email):
        return jsonify(ok=False, error="enter a valid email address"), 400
    d = load("admin_allowlist.json", {"emails": []})
    emails = {str(e).strip().lower() for e in (d.get("emails") or []) if str(e).strip()}
    emails.add(email)
    _save_json("admin_allowlist.json", {"emails": sorted(emails)})
    return jsonify(ok=True, added=email, managed=sorted(emails))

@app.post("/api/admin/users/remove")
def admin_users_remove():
    """Remove an email from the UI-managed allowlist. Admin-only. Can't remove a bootstrap(env) email or an
    admin here — those are edited via the ALLOWED_EMAILS / ADMIN_EMAILS secrets so a UI slip can't lock out."""
    g = _require_admin()
    if g: return g
    b = request.get_json(force=True, silent=True) or {}
    email = (b.get("email") or "").strip().lower()
    boot = {e.strip().lower() for e in os.environ.get("ALLOWED_EMAILS", "").split(",") if e.strip()}
    if email in boot:
        return jsonify(ok=False, error="bootstrap email — remove it from the ALLOWED_EMAILS secret instead"), 400
    if auth_gate.is_admin(email):
        return jsonify(ok=False, error="admin — cannot be removed here"), 400
    d = load("admin_allowlist.json", {"emails": []})
    emails = {str(e).strip().lower() for e in (d.get("emails") or []) if str(e).strip()}
    emails.discard(email)
    _save_json("admin_allowlist.json", {"emails": sorted(emails)})
    return jsonify(ok=True, removed=email, managed=sorted(emails))

@app.get("/api/admin/overview")
def admin_overview():
    g = _require_admin()
    if g: return g
    # book value (best-effort) + open service reports + run count + 7d usage
    book_val = None
    try:
        import book as _book
        s = _book.summary() if hasattr(_book, "summary") else None
        if isinstance(s, dict):
            book_val = s.get("revenue") or s.get("total") or s.get("value")
    except Exception:
        pass
    week = int(time.time() * 1000) - 7 * 86400000
    ev7 = [e for e in USAGE if (e.get("ts") or 0) >= week]
    return jsonify(ok=True,
                   apiOk=True,
                   warehouse=bool(STATE_BUCKET) or True,
                   opens7d=len(ev7),
                   activeUsers7d=len({e.get("user") for e in ev7}),
                   openReports=sum(1 for r in SERVICE if r.get("status") == "open"),
                   runs=len(RUNS),
                   planograms=len(PLANOGRAM),
                   bookValue=book_val)


# ─────────────────────────────────────────────────────────────────────────────
# Hoodie Relations — per-account delivery day + goals (goals matched to the master)
# ─────────────────────────────────────────────────────────────────────────────
def _rel(aid):
    return RELATIONS.get(aid) or {"accountId": aid, "deliveryDay": None, "goals": []}

@app.get("/api/relations/accounts")
def relations_list():
    # reportable roll-up: every account with extras
    out = []
    for aid, r in RELATIONS.items():
        out.append({"accountId": aid, "accountName": r.get("accountName"),
                    "deliveryDay": r.get("deliveryDay"), "goals": len(r.get("goals") or [])})
    return jsonify(ok=True, accounts=out)

@app.get("/api/relations/accounts/<aid>")
def relations_get(aid):
    return jsonify(ok=True, account=_rel(aid))

@app.put("/api/relations/accounts/<aid>")
def relations_put(aid):
    b = request.get_json(force=True, silent=True) or {}
    now = int(time.time() * 1000)
    cur = RELATIONS.get(aid) or {}
    rec = {
        "accountId": aid,
        "accountName": (b.get("accountName") or cur.get("accountName")),
        "deliveryDay": (b.get("deliveryDay") if "deliveryDay" in b else cur.get("deliveryDay")),
        "goals": (b.get("goals") if "goals" in b else (cur.get("goals") or [])),
        "updatedAt": now,
    }
    # bound the goals list a little
    if isinstance(rec["goals"], list):
        rec["goals"] = rec["goals"][:50]
    RELATIONS[aid] = rec
    _save_json("relations.json", RELATIONS)
    return jsonify(ok=True, account=rec)

@app.post("/api/relations/goals/match")
def relations_goal_match():
    b = request.get_json(force=True, silent=True) or {}
    import relations
    match = relations.match_goal((b.get("text") or ""), b.get("accountName"))
    return jsonify(ok=True, match=match, llm=relations.llm_enabled())


# ─────────────────────────────────────────────────────────────────────────────
# Generalized scraper — Claude reads a target URL and returns the data + how to scrape it
# ─────────────────────────────────────────────────────────────────────────────
@app.post("/api/scraper/analyze")
def scraper_analyze():
    b = request.get_json(force=True, silent=True) or {}
    import source_analyzer, recipes
    url = b.get("url") or ""
    result = source_analyzer.analyze(url, b.get("goal"))
    # If the live page is bot-walled but we already learned this source, fall back to the
    # saved recipe instead of dead-ending — the API + field map are already known.
    if "error" in result and result.get("blocked"):
        synth = recipes.analysis_from_recipe(recipes.get(RECIPES, url))
        if synth:
            result = synth
        else:
            # Some hosts are OWNED connectors with a dedicated scraper (search-form / bespoke) —
            # the generalized analyzer can't read them. Point to Pulls instead of a raw fetch error.
            host = recipes.host_of(url)
            conn = OWNED_HOSTS.get(host)
            if conn:
                result = {"error": "This is an owned connector (%s) with a dedicated scraper — the "
                          "generalized analyzer can't read its page. Run it from Pulls (Hoodie MDM → "
                          "Pulls), where it pulls through the engine pipeline." % conn,
                          "owned_connector": conn, "blocked": True, "attempts": result.get("attempts")}
    recipe = None
    if "error" not in result:
        if not result.get("from_recipe"):           # stash a freshly-read config as a candidate
            recipes.save_config(RECIPES, url, result, int(time.time() * 1000))
            _save_json("recipes.json", RECIPES)
        rec = recipes.get(RECIPES, url)
        recipe = {"host": rec["host"], "platform": rec.get("platform"), "status": rec["status"],
                  "platform_proven_on": recipes.platform_proven_on(RECIPES, url, result)} if rec else None
    return jsonify(ok=("error" not in result), analysis=result, llm=source_analyzer.llm_enabled(), recipe=recipe)

@app.post("/api/scraper/extract")
def scraper_extract():
    b = request.get_json(force=True, silent=True) or {}
    import source_analyzer, recipes
    url = b.get("url") or ""
    prompt = b.get("prompt") or ""
    api = dict(b.get("api") or {})
    # Reuse a proven field map (learned on an earlier run) so this pull normalizes columns
    # deterministically — no LLM call at all.
    rec0 = recipes.get(RECIPES, url)
    if not api.get("field_map") and isinstance(rec0, dict):
        fm = (rec0.get("config") or {}).get("field_map")
        if fm:
            api["field_map"] = fm
    result = source_analyzer.extract(url, prompt, b.get("pages", 1), b.get("limit", 3000), api=api)
    # Fold the run into the recipe book: validate vs baseline, promote/demote. Pass what
    # was actually run so a recipe proven by extraction alone stays runnable recipe-first.
    _, verdict = recipes.record_run(RECIPES, url, result, int(time.time() * 1000),
                                    used={"prompt": prompt, "target": url})
    # Persist a freshly-learned field map onto the recipe so future pulls reuse it (no LLM).
    fm = result.get("field_map") if isinstance(result, dict) else None
    if fm:
        rec = RECIPES.get(recipes.host_of(url))
        if isinstance(rec, dict):
            rec.setdefault("config", {})["field_map"] = fm
    # Self-heal: on drift/broken, re-analyze and write the fresh config back into the recipe.
    if verdict.get("needs_heal") and source_analyzer.llm_enabled():
        try:
            fresh = source_analyzer.analyze(url)
            if "error" not in fresh:
                recipes.apply_heal(RECIPES, url, fresh, int(time.time() * 1000))
                verdict["healed"] = True
        except Exception as e:
            app.logger.warning("recipe self-heal failed for %s: %s", url, e)
    _save_json("recipes.json", RECIPES)
    # Highlights — a 'what did this scrape yield' summary for the completion card + Hoodie Intelligence.
    try:
        hl = source_analyzer.highlights(result.get("rows"), host=recipes.host_of(url), engine=result.get("engine"))
        if hl:
            hl["ts"] = int(time.time() * 1000)
            HIGHLIGHTS.insert(0, hl); del HIGHLIGHTS[30:]
            _save_json("highlights.json", HIGHLIGHTS)
            result["highlights"] = hl
    except Exception as e:
        app.logger.warning("highlights failed for %s: %s", url, e)
    return jsonify(ok=("error" not in result), recipe=verdict, **result)

@app.get("/api/highlights")
def highlights_list():
    return jsonify(highlights=HIGHLIGHTS[:int(request.args.get("limit", 20) or 20)])

@app.post("/api/query")
def data_query():
    """Declarative run from a node in the Estate model: given a dataset + field + op, resolve it
    against the live data (the in-memory sample of the pulled dataset). Deterministic."""
    b = request.get_json(force=True, silent=True) or {}
    ds, field = b.get("dataset"), b.get("field")
    op = (b.get("op") or "distinct").lower()
    limit = max(1, min(int(b.get("limit", 60) or 60), 500))
    d = DATASETS.get(ds)
    if not isinstance(d, dict):
        return jsonify(error="unknown dataset"), 404
    header = d.get("header") or []
    rows = d.get("rows") or []
    if field not in header:
        return jsonify(error="unknown field"), 400
    ix = header.index(field)
    vals = []
    for r in rows:
        v = (r[ix] if isinstance(r, list) and len(r) > ix else (r.get(field) if isinstance(r, dict) else None))
        if v not in (None, ""):
            vals.append(str(v).strip())
    sampled, total = len(rows), (d.get("total") or len(rows))
    counts = {}
    for v in vals:
        counts[v] = counts.get(v, 0) + 1
    base = {"dataset": ds, "field": field, "op": op, "sampled": sampled, "total": total, "distinct": len(counts)}
    if op == "count":
        return jsonify(**base, result=len(vals))
    if op == "top":
        items = sorted(counts.items(), key=lambda x: -x[1])[:limit]
        return jsonify(**base, items=[{"value": k, "n": n} for k, n in items])
    dv = sorted(counts.keys())
    return jsonify(**base, values=dv[:limit], truncated=len(dv) > limit)

@app.post("/api/scraper/fingerprint")
def scraper_fingerprint():
    """Discover: given a list of chains, classify each one's platform + whether we have a
    native pull path — so we can enumerate who's on Algolia/Shopify/Yext/… at a glance."""
    b = request.get_json(force=True, silent=True) or {}
    import source_analyzer
    urls = b.get("urls") or ([b["url"]] if b.get("url") else [])
    urls = [u for u in urls if isinstance(u, str) and u.strip()][:40]   # polite cap
    out = []
    for u in urls:
        try:
            fp = source_analyzer.fingerprint(u)
        except Exception as e:
            fp = {"url": u, "platform": None, "signals": [], "native": False, "note": str(e)[:80]}
        if fp:
            out.append(fp)
    return jsonify(results=out)

# The full recipe LIBRARY — every deterministic scraper/connector we've built, not just the hosts that have
# been run through the analyzer. platform recipes (multi-retailer) + site/chain/locator/control-state/reference
# connectors. The estate shows all of these so the recipe picture is the real one, not a sliver.
_RECIPE_CONNECTORS = [
    # (name, platform/impl, status, kind)
    ("Total Wine", "getProduct API", "proven", "chain"), ("ABC FWS", "BigCommerce", "proven", "chain"),
    ("Binny's", "site", "proven", "chain"), ("Spec's", "site", "proven", "chain"),
    ("Kroger", "products API", "proven", "chain"), ("Target", "RedSky", "proven", "chain"),
    ("Walmart", "Bright Data", "proven", "chain"), ("City Hive", "container API", "building", "platform"),
    ("AB InBev locator", "GraphQL", "proven", "locator"), ("VTInfo · Tito's", "locator", "proven", "locator"),
    ("Outback locator", "Yext", "proven", "locator"), ("DoorDash", "GraphQL aggregator", "proven", "aggregator"),
    ("Instacart", "location aggregator", "building", "aggregator"),
    ("TTB COLA", "federal registry", "proven", "reference"), ("Census ACS/CBP", "Census API", "proven", "reference"),
    ("Oregon OLCC", "control state", "proven", "control"), ("Montgomery MD", "control state", "proven", "control"),
    ("BC Liquor", "control state", "proven", "control"), ("Utah DABS", "control state", "proven", "control"),
    ("Idaho / NC / MT / AL / ME", "control state", "proven", "control"),
    ("TX / IL / NY / CO / MO / CT", "Socrata outlets", "proven", "license"),
    ("CA ABC", "license roster", "proven", "license"), ("FL DBPR/ABT", "license roster", "proven", "license"),
]


@app.get("/api/recipes")
def recipes_list():
    import recipes
    seen, lib = set(), []
    # 1) the platform recipes (multi-retailer, dispatched by run_census)
    try:
        import off_premise
        for plat, impl in sorted(off_premise.RECIPE_PLATFORMS.items()):
            k = plat.replace("cityhive", "city hive")
            if k in seen:
                continue
            seen.add(k)
            lib.append({"host": plat.title(), "platform": plat, "status": "proven", "kind": "platform", "impl": impl})
    except Exception:
        pass
    # 2) the built site / chain / locator / control-state / reference connectors
    for name, impl, st, kind in _RECIPE_CONNECTORS:
        lib.append({"host": name, "platform": impl, "status": st, "kind": kind, "impl": impl})
    # 3) analyzer-learned per-host recipes (recipes.json) — merge on top (they carry field maps + run history)
    for r in RECIPES.values():
        lib.append(dict(r, kind=r.get("kind", "analyzed")))
    order = {"proven": 0, "live": 0, "building": 1, "candidate": 2}
    lib.sort(key=lambda r: (order.get((r.get("status") or "").lower(), 3), r.get("host", "")))
    return jsonify(recipes=lib, stats=recipes.stats(RECIPES))

@app.delete("/api/recipes/<host>")
def recipes_delete(host):
    existed = RECIPES.pop(host, None) is not None
    if existed:
        _save_json("recipes.json", RECIPES)
    return jsonify(ok=existed, host=host)

@app.post("/api/recipes/<host>/reset")
def recipes_reset(host):
    """Re-prove a recipe from scratch — keep the config/field-map, clear the proof state
    so it must earn 'proven' again over the next clean runs."""
    rec = RECIPES.get(host)
    if not rec:
        return jsonify(ok=False, error="no such recipe"), 404
    rec["status"] = "candidate"
    rec["clean_runs"] = 0
    rec["baseline"] = None
    rec["proven_baseline"] = None
    _save_json("recipes.json", RECIPES)
    return jsonify(ok=True, host=host, status="candidate")

@app.get("/api/scraper/recipe")
def scraper_recipe():
    import recipes
    rec = recipes.get(RECIPES, request.args.get("url", ""))
    return jsonify(recipe=rec)

@app.get("/api/hierarchy")
def hierarchy():
    # derived from the live data when present; else the bundled/curated seed
    return jsonify(derive_hierarchy(DATASETS) or HIERARCHY)

@app.post("/api/run")
def run():
    body = request.get_json(force=True, silent=True) or {}
    conn = body.get("connId")
    if conn not in VALID_CONNS and conn not in FL_CONN:
        return jsonify(error="unknown connId"), 400
    # On/off gate: a SCHEDULED run of a disabled source is skipped (manual runs still go through, so you can
    # always force one). This is what the DoorDash/Google (and any) toggle controls.
    if body.get("trigger") == "scheduled" and not _conn_is_enabled(conn):
        return jsonify(connId=conn, status="skipped", reason="connector disabled"), 200
    # Async (opt-in via {"stream":true} or ?async=1): run in a thread, stream progress via
    # /api/run/progress. This is what Hoodie Pulls' live console uses.
    if body.get("stream") or request.args.get("async"):
        jid = _new_job(conn)
        threading.Thread(target=_run_job, args=(jid, conn, body), daemon=True).start()
        return jsonify(jobId=jid, connId=conn, status="running"), 202
    # Legacy synchronous path (blocks until the pull finishes) — kept for other callers.
    try:
        rec = _dispatch_pull(conn, body)
    except Exception as e:
        app.logger.exception("run failed")
        msg = ("%s: %s" % (type(e).__name__, e)).replace("\n", " ").strip()[:400]
        if "407" in msg:
            msg = "Proxy rejected the credentials (407) — check BRIGHTDATA_PROXY_USER/PASS. · " + msg
        rec = {"id": "R-ERR", "connId": conn, "startedAt": int(time.time()*1000),
               "finishedAt": int(time.time()*1000), "durationMs": 0, "status": "failed",
               "trigger": body.get("trigger", "manual"), "total": 0, "extracts": [],
               "warnings": [msg], "error": msg}
    if rec is None:
        return jsonify(error="unknown connId"), 400
    RUNS.insert(0, rec); del RUNS[200:]; save()
    _emit_pull_highlights(conn, rec)   # estate model pulses + feed shows owned pulls (sync path too)
    return jsonify(rec)

@app.get("/api/run/progress")
def run_progress():
    """Live status of a background run job — log lines, elapsed, and the final record once done."""
    jid = (request.args.get("id") or "").strip()
    job = JOBS.get(jid)
    if not job:
        return jsonify(error="unknown job: %s" % jid), 404
    now = int(time.time() * 1000)
    return jsonify(id=jid, connId=job["connId"], status=job["status"],
                   startedAt=job["startedAt"], finishedAt=job["finishedAt"],
                   elapsedMs=(job["finishedAt"] or now) - job["startedAt"],
                   log=job["log"], run=job["run"], error=job["error"])

_PROG_RE = re.compile(r"(\d[\d,]*)\s*/\s*(\d[\d,]*)")


def _job_progress(job):
    """Parse live progress out of the job's log — scrapers emit 'fetched 50/1000', 'page 3/100', 'N/800
    parsed' etc. Latest N/M wins → current/total/pct + rows/sec + ETA + the current-stage line."""
    cur = tot = 0
    last = ""
    for e in (job.get("log") or [])[-60:]:
        msg = (e.get("msg") if isinstance(e, dict) else str(e)) or ""
        if msg.strip():
            last = msg
        mm = list(_PROG_RE.finditer(msg))
        if mm:
            try:
                c = int(mm[-1].group(1).replace(",", "")); t = int(mm[-1].group(2).replace(",", ""))
                if 0 < c <= t:
                    cur, tot = c, t
            except Exception:
                pass
    now = int(time.time() * 1000)
    el = ((job.get("finishedAt") or now) - job["startedAt"]) / 1000.0
    rate = (cur / el) if (cur and el > 0) else 0
    eta = int(((tot - cur) / rate) * 1000) if (rate > 0 and tot > cur) else None
    return {"current": cur, "total": tot, "pct": (round(100 * cur / tot) if tot else None),
            "lastMsg": last[:160], "ratePerSec": round(rate, 1), "etaMs": eta}


@app.get("/api/jobs")
def jobs_ep():
    """Live board of all pull jobs (running + recent) with parsed progress/ETA/rows-per-sec — the active-run
    tracker. Poll this to watch pulls in flight, not just their last-run status."""
    now = int(time.time() * 1000)
    out = []
    for jid, job in sorted(JOBS.items(), key=lambda kv: -kv[1]["startedAt"])[:30]:
        p = _job_progress(job)
        out.append(dict(id=jid, connId=job["connId"], status=job["status"], startedAt=job["startedAt"],
                        finishedAt=job.get("finishedAt"), elapsedMs=(job.get("finishedAt") or now) - job["startedAt"],
                        error=job.get("error"), landed=(job.get("run") or {}).get("total"), **p))
    # runs registered via runlog (scrapes that run OFF the server box — laptops, crons, batch sweeps). One file
    # per run (glob), so concurrent runs don't clobber each other's heartbeat.
    try:
        import warehouse as _wh
        rows = _wh.query_parts("scrape_runs", "SELECT * FROM t WHERE updatedAt >= ? ORDER BY startedAt DESC "
                               "LIMIT 40", [now - 24 * 3600 * 1000])
        seen = {j["id"] for j in out}
        for r in rows:
            if r.get("run_id") in seen:
                continue
            n, tot = r.get("n") or 0, r.get("total")
            elapsed = (r.get("finishedAt") or r.get("updatedAt") or now) - (r.get("startedAt") or now)
            status = r.get("status")
            hb_age = now - (r.get("updatedAt") or now)             # since last heartbeat
            prog_age = now - (r.get("progressedAt") or r.get("updatedAt") or now)   # since n last advanced
            if status == "running":
                if hb_age > 180000:
                    status = "stalled"                             # no heartbeat = the process died/hung
                elif prog_age > 90000:
                    status = "paced"                               # alive + heartbeating, just politely throttled
            rate = (n / (elapsed / 1000.0)) if (elapsed > 0 and n) else None
            eta = ((tot - n) / rate * 1000) if (rate and tot and tot > n) else None
            out.append(dict(id=r.get("run_id"), connId=r.get("connId"), status=status,
                            startedAt=r.get("startedAt"), finishedAt=r.get("finishedAt"), elapsedMs=elapsed,
                            done=n, total=tot, pct=(round(100 * n / tot) if (tot and n) else None),
                            rate=(round(rate, 1) if rate else None), etaMs=(int(eta) if eta else None),
                            host=r.get("host"), note=r.get("note"), landed=n))
    except Exception:
        pass
    out.sort(key=lambda j: -(j.get("startedAt") or 0))
    return jsonify(ok=True, active=sum(1 for j in out if j.get("status") in ("running", "paced")), jobs=out[:40])


@app.post("/api/analyze")
def analyze_ep():
    """Read an uploaded dataset → context-aware first pass + (when it fits) Report Builder
    specs. Front-end sends the RB vocabulary (dimensions/measures/viz) since it lives in
    the dashboard. Falls back gracefully: 503 when no API key, 502 on model error."""
    body = request.get_json(force=True, silent=True) or {}
    header, rows = body.get("header") or [], body.get("rows") or []
    if not header or not rows:
        return jsonify(error="need header + rows"), 400
    result = analyze.analyze(header, rows, filename=body.get("filename", "dataset.csv"),
                             registries=body.get("registries") or {},
                             full=bool(body.get("full", True)))
    if "error" not in result:
        return jsonify(result)
    return jsonify(result), (503 if result["error"] == "llm-disabled" else 502)

@app.post("/api/ai-read")
def ai_read_ep():
    """AI analysis of data opened on its OWN terms (file-as-root). Privacy-preserving: the
    browser sends the PROFILE + COMPUTED AGGREGATES only (no raw rows). 503 without a key."""
    body = request.get_json(force=True, silent=True) or {}
    profile = body.get("profile")
    if not profile:
        return jsonify(error="need profile"), 400
    result = analyze.ai_read(profile, summary=body.get("summary"),
                             filename=body.get("filename", "dataset.csv"),
                             header=body.get("header"), rows=body.get("rows"))
    if "error" not in result:
        return jsonify(result)
    return jsonify(result), (503 if result["error"] == "llm-disabled" else 502)


@app.post("/api/parse-upload")
def parse_upload_ep():
    """Fast parse-only (no Claude): raw CSV/TSV/XLSX upload → header (auto-detected past preamble) + rows +
    per-field profile. For multi-file ingest (e.g. the Order Hub taking 50 supplier menus) where a Claude read
    per file would be too slow/costly — heuristic column mapping happens client-side."""
    f = request.files.get("file")
    if not f:
        return jsonify(error="no file"), 400
    try:
        raw = f.read()
        parsed = analyze.parse_upload(f.filename or "upload.csv", raw)
    except Exception as e:
        return jsonify(error="parse-failed", detail=str(e)[:200]), 400
    if not parsed.get("header"):
        return jsonify(error="parse-failed", detail="couldn't find a header row"), 400
    prof = analyze.profile_columns(parsed["header"], parsed["rows"])
    return jsonify(filename=f.filename, header=parsed["header"], rows=parsed["rows"][:20000],
                   row_src=parsed.get("row_src", [])[:20000], col_src=parsed.get("col_src", []),
                   header_row=parsed["header_row"], sheet=parsed.get("sheet"),
                   row_count=len(parsed["rows"]), profile=prof)


@app.post("/api/refill-xlsx")
def refill_xlsx_ep():
    """Refill an ORIGINAL xlsx WITH ITS FORMATTING: load the uploaded file writable (openpyxl preserves styles)
    and write the quantities at their physical rows. `orders` = {physical_0based_row: qty} (from parse-upload's
    row_src). Landing: `qty_col` (physical 0-based col, from parse-upload's col_src) drops quantities INTO that
    existing column — so a menu whose own formulas hang off its units column (e.g. Curaleaf's Base Total $ =
    Total Units × Base $) computes when the supplier opens it. Without `qty_col`, appends an order column
    (`header_row` places its label). Returns the formatted .xlsx."""
    import io
    import openpyxl
    from flask import send_file
    f = request.files.get("file")
    if not f:
        return jsonify(error="no file"), 400
    try:
        orders = json.loads(request.form.get("orders") or "{}")
    except Exception:
        orders = {}
    label = (request.form.get("label") or "ORDER QTY")[:40]
    try:
        header_row = int(request.form.get("header_row") or 0)
    except Exception:
        header_row = 0
    try:
        qty_col = int(request.form.get("qty_col"))                 # physical 0-based landing column
    except (TypeError, ValueError):
        qty_col = -1
    try:
        data_rows = json.loads(request.form.get("rows") or "[]")   # parsed data rows (row_src) — clearing scope
    except Exception:
        data_rows = []
    try:
        orders_by_col = json.loads(request.form.get("orders_by_col") or "null")   # {phys_col: {phys_row: qty}}
    except Exception:
        orders_by_col = None

    def _fill_col(ws, qcol, col_orders, clear_rows):
        # Phantom-order guard: a pre-filled landing column (par/suggested quantities, or a
        # re-uploaded prior order) must not survive as silent order lines the buyer never sees.
        # Clear LITERAL values on every data row not ordered; leave formulas alone — they
        # compute from their own inputs (e.g. =SUM(LOC1:LOC4)) and evaluate 0 when empty.
        for r0 in clear_rows:
            try:
                r0 = int(r0)
            except (TypeError, ValueError):
                continue
            if str(r0) in col_orders:
                continue
            cell = ws.cell(row=r0 + 1, column=qcol)
            if cell.data_type != "f" and cell.value not in (None, ""):
                cell.value = None
        for k, v in col_orders.items():
            if v in (None, ""):
                continue
            try:
                r0 = int(k)
                q = int(v) if str(v).strip().lstrip("-").isdigit() else v
                ws.cell(row=r0 + 1, column=qcol, value=q)
            except Exception:
                pass

    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()))          # writable → preserves original formatting
        ws = wb[wb.sheetnames[0]]
        if isinstance(orders_by_col, dict) and orders_by_col:
            # multi-column landing: per-store order columns (Dispensary 1..N / LOC1..4 / MED·AU pairs),
            # each written into its own physical column so the file's own totals compute per store
            for cs, col_orders in orders_by_col.items():
                try:
                    qcol = int(cs) + 1
                except (TypeError, ValueError):
                    continue
                if isinstance(col_orders, dict):
                    _fill_col(ws, qcol, col_orders, data_rows)
        elif 0 <= qty_col:
            _fill_col(ws, qty_col + 1, orders, data_rows)          # land in the file's own column (header stays)
        else:
            qcol = ws.max_column + 1                               # append the order column after the used range
            ws.cell(row=header_row + 1, column=qcol, value=label)  # header at its physical (1-based) row
            for k, v in orders.items():
                if v in (None, "") or v in (0, "0"):
                    continue                                       # append mode: blank already means no order
                try:
                    r0 = int(k)
                    q = int(v) if str(v).strip().lstrip("-").isdigit() else v
                    ws.cell(row=r0 + 1, column=qcol, value=q)
                except Exception:
                    pass
        out = io.BytesIO(); wb.save(out); out.seek(0)
    except Exception as e:
        return jsonify(error="refill-failed", detail=str(e)[:200]), 400
    fn = (f.filename or "order").rsplit(".", 1)[0] + " - ORDER.xlsx"
    return send_file(out, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=fn)


# Known brand → public Lingo kit (kit uuid). Menus from these suppliers can auto-attach product photography.
LINGO_BRANDS = {
    "curaleaf-ny": {"label": "Curaleaf NY", "kit": "A4EAE236-742F-443B-BC4E-5A94D1A04B12",
                    "portal": "4PkjwR", "sections": ["Product Photography", "Strain Cards"],
                    "match": ["curaleaf", "ocm-xrod", "whole flower", "20 to 1", "20to1"]},
}
_LINGO_MAN = {}  # kit uuid -> manifest (in-process cache; Lingo kits are stable)


def _lingo_manifest(kit, sections=None):
    key = kit
    if key not in _LINGO_MAN:
        import lingo_assets
        _LINGO_MAN[key] = lingo_assets.manifest(kit, only_sections=sections)
    return _LINGO_MAN[key]


@app.post("/api/lingo-images")
def lingo_images_ep():
    """Attach brand product photography to an inbound menu. POST {brand|kit|portal, products:[{name}], sections?}
    → {kit, asset_total, matched, matches:[{name, matched, img, thumb, full, score}]}. Images come from a public
    Lingo DAM portal (see unifyd/lingo_assets.py); URLs carry a durable public token and render directly in <img>."""
    import lingo_assets
    b = request.get_json(force=True, silent=True) or {}
    brand = (b.get("brand") or "").strip().lower()
    kit = (b.get("kit") or "").strip()
    portal = (b.get("portal") or "").strip()
    sections = b.get("sections")
    products = b.get("products") or []
    reg = LINGO_BRANDS.get(brand)
    if reg:
        kit = kit or reg["kit"]; sections = sections or reg["sections"]
    if not kit and portal:
        try:
            kit = kit or (lingo_assets.portal(portal) or {}).get("uuid")  # portal metadata only; kit still needed
        except Exception:
            pass
    if not kit:
        return jsonify(error="no-kit", detail="pass brand, kit, or a resolvable portal"), 400
    try:
        man = _lingo_manifest(kit, sections)
        matches = lingo_assets.match_products(products, man) if products else []
    except Exception as e:
        return jsonify(error="lingo-failed", detail=str(e)[:200]), 502
    return jsonify(kit=kit, asset_total=man.get("asset_total", 0),
                   matched=sum(1 for m in matches if m.get("matched")), matches=matches)


@app.post("/api/ai-read-upload")
def ai_read_upload_ep():
    """File-mode reader: accepts a raw CSV/TSV/XLSX upload (multipart 'file'), auto-detects the header row
    (handles preamble rows), profiles, and has Claude read it on its own terms. Returns the analysis PLUS a
    `parse` block (detected header, header-row index, sheet, per-field profile) so the UI can show editable
    field metadata (datatype + role) and hand off to the Report Builder. 503 without a key."""
    f = request.files.get("file")
    if not f:
        return jsonify(error="no file"), 400
    try:
        raw = f.read()
        parsed = analyze.parse_upload(f.filename or "upload.csv", raw)
    except Exception as e:
        return jsonify(error="parse-failed", detail=str(e)[:200]), 400
    if not parsed.get("header"):
        return jsonify(error="parse-failed", detail="couldn't find a header row"), 400
    prof = analyze.profile_columns(parsed["header"], parsed["rows"])
    result = analyze.ai_read(prof, filename=f.filename, header=parsed["header"], rows=parsed["rows"])
    if not isinstance(result, dict):
        result = {"error": "ai-read-failed"}
    result["parse"] = {"header": parsed["header"], "header_row": parsed["header_row"],
                       "sheet": parsed.get("sheet"), "row_count": len(parsed["rows"]),
                       "rows": parsed["rows"][:8000], "profile": prof}
    if "error" not in result:
        return jsonify(result)
    return jsonify(result), (503 if result["error"] == "llm-disabled" else 502)

# ── Hoodie Intelligence analyst — real Claude answers + a pause switch ────────────────
_AI_PAUSE_FILE = os.path.join(STATE_DIR, "ai_paused.json")

def _ai_paused():
    try: return bool(json.load(open(_AI_PAUSE_FILE)).get("paused"))
    except Exception: return False

def _ai_set_paused(v):
    try: json.dump({"paused": bool(v)}, open(_AI_PAUSE_FILE, "w"))
    except Exception: pass
    return bool(v)

@app.get("/api/ai/status")
def ai_status_ep():
    """Is the live Claude analyst available, and is it currently paused? Drives the toggle."""
    return jsonify(enabled=hi_analyst.enabled(), paused=_ai_paused())

@app.post("/api/ai/pause")
def ai_pause_ep():
    """Pause/resume the live analyst — the 'switch it off while I work on the app' control.
    Persists so it stays put across restarts until flipped back."""
    b = request.get_json(force=True, silent=True) or {}
    return jsonify(paused=_ai_set_paused(b.get("paused", True)))

@app.post("/api/ask")
def ai_ask_ep():
    """Answer a free-form Q&A question with Claude, grounded in the exact figures the front-end
    computed (rbComputeMeasure). Returns {paused:true} when switched off → the browser falls
    back to its deterministic synthesizer. 503 without a key."""
    if _ai_paused():
        return jsonify(paused=True)
    b = request.get_json(force=True, silent=True) or {}
    q, facts = (b.get("question") or "").strip(), b.get("facts") or []
    if not q or not facts:
        return jsonify(error="need question + facts"), 400
    result = hi_analyst.ask(q, facts, vocab=b.get("vocab") or {})
    if "error" not in result:
        return jsonify(result)
    return jsonify(result), (503 if result["error"] == "llm-disabled" else 502)

@app.post("/api/enrich/census")
def enrich_census_ep():
    """Reference-data ENRICH (not master ingest): join census_acs county demographics onto the
    outlet master by county+state → lands `outlets_census`. Auto-picks the biggest outlet dataset
    (has county+state) unless `outlet_dataset` is given. Run the US Census pull first."""
    packs = [v for k, v in DATASETS.items()
             if k.startswith("census_") and isinstance(v, dict) and v.get("rows")]
    if not packs:
        return jsonify(error="no census data yet — run the US Census pull first"), 400
    census_ds = enrich.merge_census(packs)   # merge demographic + economic + housing → one county table
    body = request.get_json(silent=True) or {}
    # FL DBPR extracts all share one header, so items/brands "look" like outlets. Restrict to the
    # ACTUAL outlet tables (retail/wholesale/permits + ABC store cells + the places accounts), never
    # items/registrants/brands, then pick whichever genuinely matches census best.
    OUTLET_IDS = {"outlets_master", "bd4006lic", "bd4005lic", "bd4002lic", "abc_store_cells", "tx_outlets", "il_outlets", "ct_outlets", "ny_outlets", "co_outlets", "mo_outlets"}
    NON_OUTLET = {"bd4008lic", "bd4011lic", "abtbrands", "census_acs", "outlets_census"}
    def _places(d):
        h = [str(x).lower() for x in (d.get("header") or [])]
        return "county" in h and ("account_id" in h or "premise" in h)
    want = body.get("outlet_dataset")
    if want and isinstance(DATASETS.get(want), dict):
        cands = [(want, DATASETS[want])]
    else:
        cands = [(k, v) for k, v in DATASETS.items()
                 if isinstance(v, dict) and v.get("rows") and k not in NON_OUTLET
                 and (k in OUTLET_IDS or _places(v))]
    if not cands:
        return jsonify(error="no outlet dataset to join — run FL Outlets (bd4006lic) or the Orlando accounts pull first"), 400
    best = None
    for k, v in cands:
        r = enrich.join_census_to_outlets(census_ds, v)
        if "error" in r:
            continue
        if best is None or r["matched"] > best[2]["matched"]:
            best = (k, v, r)
    if best is None:
        return jsonify(error="outlet datasets present but none have a joinable county/state column"), 400
    oid, outlets, res = best
    DATASETS.update(_absorb({"outlets_census": {"header": res["header"], "rows": res["rows"][:800],
                             "total": len(res["rows"]), "_rows_full": res["rows"]}})); save()
    cov = round(100.0 * res["matched"] / res["total"], 1) if res["total"] else 0
    return jsonify(ok=True, outlet_dataset=oid, matched=res["matched"], total=res["total"],
                   coverage=cov, counties=res["counties_indexed"], demo_cols=res["demo_cols"],
                   landed="outlets_census", header=res["header"], sample=res["rows"][:5])

@app.post("/api/master/outlets/build")
def master_outlets_ep():
    """Unify the per-state outlet pulls (FL bd4006lic + TX tx_outlets + …) into ONE normalized
    outlets_master — the base outlet spine the coverage map + census-join-at-scale need. Uses FULL
    rows from the on-demand store when present. Run the state outlet pulls first."""
    sources = {}
    for did in master.SOURCES:
        full = load_full(did)
        if full and full.get("rows"):
            sources[did] = full
        elif isinstance(DATASETS.get(did), dict) and DATASETS[did].get("rows"):
            sources[did] = DATASETS[did]
    if not sources:
        return jsonify(error="no outlet pulls yet — run FL Outlets and/or Texas TABC first"), 400
    res = master.build(sources)
    DATASETS.update(_absorb({"outlets_master": {"header": res["header"], "rows": res["rows"][:800],
                             "total": res["total"], "_rows_full": res["rows"]}})); save()
    return jsonify(ok=True, landed="outlets_master", total=res["total"], by_state=res["by_state"],
                   sources=res["sources"], header=res["header"], sample=res["rows"][:5])

@app.get("/api/benchmark")
def benchmark_ep():
    """Market NORM for the Planogram app — format-aware. Deterministic (no LLM), so it always
    answers; the seam where real SipSource depletion norms plug in."""
    result = planogram.benchmark(request.args.get("market"), request.args.get("format", "off"))
    return jsonify(result), (404 if result.get("error") == "unknown-market" else 200)


@app.post("/api/shelf-vision")
def shelf_vision_ep():
    """Photo of a shelf/back-bar -> facings per category x tier (Claude vision). 503 without a key."""
    body = request.get_json(force=True, silent=True) or {}
    result = planogram.shelf_count(body.get("image"), media_type=body.get("media_type", "image/jpeg"),
                                   categories=body.get("categories"), tiers=body.get("tiers"))
    if "error" not in result:
        return jsonify(result)
    code = 503 if result["error"] == "llm-disabled" else (400 if result["error"] == "need-image" else 502)
    return jsonify(result), code


@app.post("/api/pitch")
def pitch_ep():
    """Narrate the computed shelf gaps into a buyer pitch (numbers in, wording out). 503 without a key."""
    body = request.get_json(force=True, silent=True) or {}
    result = planogram.pitch(body.get("account"), body.get("market"), body.get("format", "off"),
                             body.get("deltas") or [])
    if "error" not in result:
        return jsonify(result)
    return jsonify(result), (503 if result["error"] == "llm-disabled" else 502)


@app.get("/api/prism")
def prism_ep():
    """Prism mobile app's data contract — the book cut every way, plus the pulse feed.
    Deterministic (no LLM), so it always answers and the app works offline once cached."""
    return jsonify(prism.bundle(request.args.get("measure")))


@app.post("/api/upc")
def upc_ep():
    """Assess one or many UPC/EAN codes. Deterministic QC always (check digit, placeholder,
    restricted ranges); owner-agreement + confidence when the prefix->owner crosswalk is loaded.
    Body: {"upc": "...", "applicant": "..."} OR {"items": [{"upc","applicant"}, ...]}."""
    body = request.get_json(force=True, silent=True) or {}
    items = body.get("items")
    if items is None:
        items = [{"upc": body.get("upc", ""), "applicant": body.get("applicant", "")}]
    xw = UPC_XWALK or None
    out = [upc.assess((it or {}).get("upc", ""), applicant=(it or {}).get("applicant"), crosswalk=xw)
           for it in items[:5000]]
    return jsonify(ok=True, crosswalk_loaded=bool(UPC_XWALK), count=len(out), results=out)


GEO_CAP = 500000
@app.get("/api/outlets/geo")
def outlets_geo_ep():
    """LIGHT geocoded points for the coverage map — enough to render, colour, filter and search the
    whole footprint at once, not just one state. ?dataset=outlets_master (default). Each point is
    {i,lat,lng,name,type,area,county_fips} where i is the row index used by /api/outlets/one to
    fetch the FULL record lazily on click (so we ship ~7 fields/point, not the whole record ×100k).
    Any dataset carrying latitude/longitude works (IL/Chicago native; TX/CT/FL geocoded)."""
    did = request.args.get("dataset", "outlets_master")
    full = load_full(did) or DATASETS.get(did)
    if not isinstance(full, dict) or not full.get("rows"):
        return jsonify(ok=True, dataset=did, count=0, points=[], note="no data pulled yet")
    header = full.get("header", []); rows = full.get("rows", [])
    idx = {str(h).lower(): i for i, h in enumerate(header)}
    def gi(*names):
        for n in names:
            if n in idx: return idx[n]
        return -1
    la, lo = gi("latitude", "lat"), gi("longitude", "lng", "lon")
    if la < 0 or lo < 0:
        return jsonify(ok=True, dataset=did, count=0, points=[], note="dataset has no lat/lng — geocode first")
    ni = gi("dba", "name", "trade_name")
    ti = gi("license_types", "license_type", "type", "credential")
    ai = gi("community_area", "county")
    ci = gi("county_fips", "fips")
    il_cook = "17031" if did == "il_outlets" else None   # Chicago ships lat/lng, no FIPS col → all Cook
    # Optional viewport filter: ?bbox=west,south,east,north — the map sends its current bounds so we
    # ship only the dots in view (a few thousand), not the whole 200k+ footprint on every pan. in_view
    # is the true count matching the bbox (before the cap) so the client can say "zoom in for all".
    west = south = east = north = None
    try:
        bb = request.args.get("bbox", "")
        if bb:
            west, south, east, north = (float(x) for x in bb.split(","))
    except (ValueError, TypeError):
        west = None
    # Return EVERY dot in view (up to a large safety cap) — the whole point of the map is to SEE the
    # gaps, and a sample fills them in. The client draws all of them on a single canvas layer (fast for
    # 100k+), so density/coverage is truthful; the reservoir only trips past GEO_CAP as a runaway guard.
    cap = GEO_CAP
    pts = []; in_view = 0
    for ri, r in enumerate(rows):
        if la >= len(r) or lo >= len(r):
            continue
        try:
            lat, lng = float(r[la]), float(r[lo])
        except (ValueError, TypeError):
            continue
        if not (lat and lng and -90 < lat < 90 and -180 < lng < 180):
            continue
        if west is not None and not (south <= lat <= north and west <= lng <= east):
            continue
        in_view += 1
        p = {"i": ri, "lat": round(lat, 6), "lng": round(lng, 6)}
        if 0 <= ni < len(r) and r[ni]: p["name"] = r[ni]
        if 0 <= ti < len(r) and r[ti]: p["type"] = r[ti]
        if 0 <= ai < len(r) and r[ai]: p["area"] = r[ai]
        cf = (r[ci] if 0 <= ci < len(r) else "") or il_cook
        if cf: p["county_fips"] = cf
        if len(pts) < cap:
            pts.append(p)
        else:
            j = random.randint(0, in_view - 1)      # reservoir: uniform sample across the viewport
            if j < cap:
                pts[j] = p
    return jsonify(ok=True, dataset=did, count=len(pts), in_view=in_view,
                   points=pts, capped=in_view > len(pts))


@app.get("/api/outlets/one")
def outlets_one_ep():
    """The FULL record for one outlet, fetched lazily when a dot is clicked. ?dataset=&i=<row index>
    (the index the /api/outlets/geo point carries). Returns {record: {column: value}} straight from
    the full rows — the detail panel renders the outlet's own data from this."""
    did = request.args.get("dataset", "outlets_master")
    try:
        i = int(request.args.get("i", "-1"))
    except ValueError:
        i = -1
    full = load_full(did) or DATASETS.get(did)
    if not isinstance(full, dict) or not full.get("rows") or not (0 <= i < len(full["rows"])):
        return jsonify(ok=False, error="not found"), 404
    header = full.get("header", []); row = full["rows"][i]
    rec = {str(h): (row[j] if j < len(row) else "") for j, h in enumerate(header)}
    return jsonify(ok=True, dataset=did, record=rec)


@app.get("/api/outlets/table")
def outlets_table_ep():
    """The outlet master AS A TABLE — the data underneath the coverage map. Server-side paged +
    filtered (the master is 255k+ rows, can't ship to the DOM whole). ?dataset=&offset=&limit=&q=
    (name contains) &state= (exact) &type= (contains). Each row carries its index `i` + lat/lng +
    county_fips so a table click reuses the same detail panel + map pan as a dot click."""
    did = request.args.get("dataset", "outlets_master")
    try:
        offset = max(0, int(request.args.get("offset", "0")))
    except ValueError:
        offset = 0
    try:
        limit = min(500, max(1, int(request.args.get("limit", "100"))))
    except ValueError:
        limit = 100
    q = (request.args.get("q", "") or "").strip().lower()
    fstate = (request.args.get("state", "") or "").strip().lower()
    ftype = (request.args.get("type", "") or "").strip().lower()
    full = load_full(did) or DATASETS.get(did)
    if not isinstance(full, dict) or not full.get("rows"):
        return jsonify(ok=True, dataset=did, total=0, filtered=0, offset=offset, limit=limit, rows=[])
    header = full["header"]; rows = full["rows"]
    idx = {str(h).lower(): i for i, h in enumerate(header)}
    def gi(*names):
        for n in names:
            if n in idx:
                return idx[n]
        return -1
    ni = gi("name", "dba", "trade_name"); ti = gi("license_type", "license_types", "credential", "type")
    oi = gi("owner", "backer", "owner name"); cyi = gi("city"); coi = gi("county"); si = gi("state")
    zi = gi("zip"); li = gi("license_num", "credential", "license_id", "outlet_id"); sri = gi("source")
    cfi = gi("county_fips", "fips"); lai = gi("latitude", "lat"); loi = gi("longitude", "lng", "lon")
    def v(r, i):
        return str(r[i]) if 0 <= i < len(r) and r[i] is not None else ""
    out = []; matched = 0
    for ri, r in enumerate(rows):
        nm = v(r, ni); ty = v(r, ti); stt = v(r, si)
        if q and q not in nm.lower():
            continue
        if fstate and fstate != stt.lower():
            continue
        if ftype and ftype not in ty.lower():
            continue
        matched += 1
        if matched <= offset or len(out) >= limit:
            continue
        out.append({"i": ri, "name": nm, "type": ty, "owner": v(r, oi), "city": v(r, cyi),
                    "county": v(r, coi), "state": stt, "zip": v(r, zi), "license_num": v(r, li),
                    "source": v(r, sri), "county_fips": v(r, cfi), "lat": v(r, lai), "lng": v(r, loi)})
    return jsonify(ok=True, dataset=did, total=len(rows), filtered=matched,
                   offset=offset, limit=limit, rows=out)


@app.get("/api/outlets/context")
def outlets_context_ep():
    """Everything the Census reference layer knows about one outlet's geography — the market
    behind a single dot on the coverage map. ?county_fips=17031 (5-digit state+county). Returns the
    county's population (PEP), beer/wine/liquor retailer + bev-alc wholesaler + drinking-place
    establishment/employment/payroll counts (CBP), the state-level nonemployer count (the small/
    independent operators CBP misses), and retailers-per-10k-residents density. Suppressed or
    uncovered cells come back null — never a fabricated zero — so the panel can say 'not reported'."""
    fips = (request.args.get("county_fips") or "").strip()
    if not (fips.isdigit() and len(fips) == 5):
        return jsonify(ok=False, error="county_fips must be a 5-digit FIPS"), 400
    st = fips[:2]
    try:
        import census_ref
    except Exception as e:
        return jsonify(ok=False, error="census layer unavailable: %s" % e), 200

    def cell(dataset, geo_level, gfips, naics, metric):
        try:
            rows = census_ref.query(dataset=dataset, geo_level=geo_level, geo_fips=gfips,
                                    naics=naics, metric=metric, limit=5)
        except Exception:
            return None
        for r in rows:
            if r.get("suppressed"):
                return None
            try:
                return int(float(r.get("metric_value")))
            except (TypeError, ValueError):
                return None
        return None

    NAICS_LABELS = (("44531", "retailers"), ("4248", "wholesalers"), ("722", "onpremise"))
    market = {}
    for naics, label in NAICS_LABELS:
        market[label] = {
            "estab": cell("cbp", "county", fips, naics, "estab"),
            "emp": cell("cbp", "county", fips, naics, "emp"),
            "payann": cell("cbp", "county", fips, naics, "payann"),
            "nonemployer_state": cell("nonemployer", "state", st, naics, "nestab"),
        }
    pop = cell("pep", "county", fips, None, "population")
    ret = market["retailers"]["estab"]
    dens = round(ret / pop * 10000, 2) if (pop and ret) else None
    return jsonify(ok=True, county_fips=fips, state_fips=st, population=pop,
                   per_10k_retailers=dens, market=market)


GEO_JOBS = {}
@app.post("/api/outlets/geocode/build")
def outlets_geocode_build():
    """Geocode an outlet dataset's addresses (free US Census batch) → append latitude/longitude/
    county_fips, so the Coverage Map + census join work for it (county_fips also fixes the FL
    numeric-county gap). Background job (~1 min per 10k). ?dataset=outlets_master (default);
    poll /api/outlets/geocode/progress?id=<jobId>."""
    body = request.get_json(silent=True) or {}
    did = body.get("dataset") or request.args.get("dataset", "outlets_master")
    full = load_full(did) or DATASETS.get(did)
    if not isinstance(full, dict) or not full.get("rows"):
        return jsonify(ok=False, error="no data pulled for %s" % did), 400
    import random as _rnd
    jid = "GEO-" + "".join(_rnd.choices("ABCDEFGHJKMNPQRSTUVWXYZ23456789", k=6))
    GEO_JOBS[jid] = {"jobId": jid, "dataset": did, "status": "running",
                     "total": full.get("total", len(full["rows"])), "matched": 0, "requested": 0, "log": []}
    header, rows = full["header"], full["rows"]
    def run():
        try:
            import geocode
            nh, nr, stats = geocode.geocode_outlets(header, rows,
                                                    log=lambda m: GEO_JOBS[jid]["log"].append(m))
            DATASETS[did] = {"header": nh, "rows": nr[:800], "total": len(nr)}
            save_full(did, nh, nr); save()
            GEO_JOBS[jid].update(status="done", matched=stats["matched"], requested=stats["requested"])
        except Exception as e:
            app.logger.exception("geocode failed")
            GEO_JOBS[jid].update(status="failed", error=str(e)[:200])
    threading.Thread(target=run, daemon=True).start()
    return jsonify(ok=True, jobId=jid, dataset=did, total=GEO_JOBS[jid]["total"])


@app.get("/api/outlets/geocode/progress")
def outlets_geocode_progress():
    j = GEO_JOBS.get(request.args.get("id", ""))
    if not j:
        return jsonify(error="unknown job"), 404
    return jsonify(jobId=j["jobId"], dataset=j.get("dataset", ""), status=j["status"], matched=j.get("matched", 0),
                   requested=j.get("requested", 0), total=j.get("total", 0), error=j.get("error"), log=j["log"][-6:])


@app.post("/api/census/build")
def census_build_ep():
    """Build the census_reference layer (CBP + Nonemployer + PEP → warehouse). Background; needs
    CENSUS_API_KEY. Poll /api/outlets/geocode/progress?id=<jobId> (shared job store)."""
    if not os.environ.get("CENSUS_API_KEY", "").strip():
        return jsonify(ok=False, error="CENSUS_API_KEY not set"), 400
    import random as _r
    jid = "CEN-" + "".join(_r.choices("ABCDEFGHJKMNPQRSTUVWXYZ23456789", k=6))
    GEO_JOBS[jid] = {"jobId": jid, "status": "running", "log": []}
    def run():
        try:
            import census_ref
            res = census_ref.build(log=lambda m: GEO_JOBS[jid]["log"].append(m))
            GEO_JOBS[jid].update(status="done", rows=res["rows"], uri=res["uri"])
        except Exception as e:
            app.logger.exception("census build failed")
            GEO_JOBS[jid].update(status="failed", error=str(e)[:200])
    threading.Thread(target=run, daemon=True).start()
    return jsonify(ok=True, jobId=jid)


@app.get("/api/census/reference")
def census_reference_ep():
    """Query census_reference by geo/naics/metric/vintage — the territory.html join helper. e.g.
    ?dataset=cbp&geo_level=county&geo_fips=17031&naics=44531&metric=estab"""
    import census_ref
    a = request.args
    try:
        rows = census_ref.query(dataset=a.get("dataset"), geo_level=a.get("geo_level"),
                                geo_fips=a.get("geo_fips"), naics=a.get("naics"), metric=a.get("metric"),
                                vintage=a.get("vintage"), limit=int(a.get("limit", 5000)))
        return jsonify(ok=True, count=len(rows), rows=rows)
    except Exception as e:
        return jsonify(ok=False, error=str(e)[:160]), 500


@app.get("/api/cex/spend")
def cex_spend_ep():
    """BLS CEX mean annual alcohol $ per consumer unit by income bracket.
    ?item=ALCBEVG|ALCHOME|ALCAWAY|TOTALEXP|INCBEFTX  ?bracket=21  ?year=2023"""
    import cex_ref
    try:
        yr = int(request.args["year"]) if request.args.get("year") else None
        rows = cex_ref.query(item=request.args.get("item") or None,
                             bracket=request.args.get("bracket") or None, year=yr)
        return jsonify(ok=True, landed=True, count=len(rows), rows=rows)
    except Exception:
        return jsonify(ok=True, landed=False, count=0, rows=[],
                       note="cex_reference not landed yet (run the cex source: cex_ref.build)")


@app.get("/api/census/demand")
def census_demand_ep():
    """Trade-area alcohol demand $ for one geo — CEX mean spend × ACS B19001 households, with the
    at-home (off-premise) / away (on-premise) split. ?geo_fips=12095[&geo_level=county|state|zcta].
    Bare 5-digit ids try county then fall back to ZIP (they collide); pass geo_level=zcta to force ZIP.
    Honest {ok:false, reason} until both cex_reference and the ACS brackets are landed."""
    import cex_ref
    fips = (request.args.get("geo_fips") or "").strip()
    if not fips:
        return jsonify(ok=False, error="geo_fips required (county FIPS, ZIP, or 2-digit state)"), 400
    try:
        return jsonify(**cex_ref.demand_for(fips, geo_level=request.args.get("geo_level") or None))
    except Exception as e:
        return jsonify(ok=False, error=str(e)[:160]), 500


@app.get("/api/census/demand/top")
def census_demand_top_ep():
    """Rank geos by trade-area demand (the demand.html leaderboard). ?geo_level=zcta|county|state
    ?prefix=328 (geo_fips prefix — e.g. Orlando-area ZIPs) ?by=demand_total_usd|demand_per_hh_usd|
    demand_index_vs_us|households ?limit=25 ?min_households=500 (floor keeps tiny-ZCTA noise out of
    per-HH/index rankings)."""
    import warehouse
    lvl = request.args.get("geo_level") or "county"
    by = request.args.get("by") or "demand_total_usd"
    if by not in ("demand_total_usd", "demand_per_hh_usd", "demand_index_vs_us", "households"):
        return jsonify(ok=False, error="bad ?by"), 400
    if lvl not in ("zcta", "county", "state"):
        return jsonify(ok=False, error="bad ?geo_level"), 400
    try:
        limit = max(1, min(500, int(request.args.get("limit", 25))))
        min_hh = int(request.args.get("min_households", 500))
    except ValueError:
        return jsonify(ok=False, error="bad numeric param"), 400
    where, params = ["geo_level = ?", "households >= ?"], [lvl, min_hh]
    pref = (request.args.get("prefix") or "").strip()
    if pref:
        where.append("geo_fips LIKE ?"); params.append(pref + "%")
    sql = ("SELECT * FROM t WHERE " + " AND ".join(where)
           + " ORDER BY %s DESC LIMIT %d" % (by, limit))
    try:
        rows = warehouse.query("trade_area_demand", sql, params)
        return jsonify(ok=True, count=len(rows), rows=rows)
    except Exception:
        return jsonify(ok=True, count=0, rows=[],
                       note="trade_area_demand not landed yet (run cex_ref.build_demand)")


@app.get("/api/cpi")
def cpi_ep():
    """BLS CPI-U alcohol series. ?item=SAF116&area=0000&year=2024&annual=1 — or ?real=1 for the
    item rebased against all-items (relative real price, the deflator/premiumization read)."""
    import cpi_ref
    a = request.args
    try:
        if a.get("real") in ("1", "true", "yes"):
            yr = int(a["base_year"]) if a.get("base_year") else None
            return jsonify(**cpi_ref.real_series(item=a.get("item") or "SAF116",
                                                 area=a.get("area") or "0000", base_year=yr))
        yr = int(a["year"]) if a.get("year") else None
        rows = cpi_ref.query(item=a.get("item") or None, area=a.get("area") or None, year=yr,
                             annual=a.get("annual") in ("1", "true", "yes"))
        return jsonify(ok=True, landed=True, count=len(rows), rows=rows)
    except ValueError as e:
        return jsonify(ok=False, error="bad numeric param: %s" % e), 400
    except Exception:
        return jsonify(ok=True, landed=False, count=0, rows=[],
                       note="cpi_reference not landed yet (run the cpi source: cpi_ref.build)")


@app.get("/api/fred")
def fred_ep():
    """FRED macro series. ?series=MRTSSM4453USN|liquor_store_sales&year=2026."""
    import fred_ref
    try:
        yr = int(request.args["year"]) if request.args.get("year") else None
        rows = fred_ref.query(series=request.args.get("series") or None, year=yr,
                              limit=int(request.args.get("limit", 5000)))
        return jsonify(ok=True, landed=True, count=len(rows), rows=rows)
    except ValueError as e:
        return jsonify(ok=False, error="bad numeric param: %s" % e), 400
    except Exception:
        return jsonify(ok=True, landed=False, count=0, rows=[],
                       note="fred_reference not landed yet (run the fred source: fred_ref.build, needs FRED_API_KEY)")


@app.get("/api/bea")
def bea_ep():
    """BEA regional income. ?metric=personal_income_per_capita&geo_level=county&geo_fips=12095&year=2023."""
    import bea_ref
    try:
        yr = int(request.args["year"]) if request.args.get("year") else None
        rows = bea_ref.query(metric=request.args.get("metric") or None,
                             geo_level=request.args.get("geo_level") or None,
                             geo_fips=request.args.get("geo_fips") or None, year=yr)
        return jsonify(ok=True, landed=True, count=len(rows), rows=rows)
    except ValueError as e:
        return jsonify(ok=False, error="bad numeric param: %s" % e), 400
    except Exception:
        return jsonify(ok=True, landed=False, count=0, rows=[],
                       note="bea_reference not landed yet (run the bea source: bea_ref.build — the BEA key "
                            "must be activated via BEA's email link)")


@app.get("/api/places")
def places_ep():
    """Query the pulled on-premise accounts (Orlando) from the warehouse. Filters:
    ?premise=on|off|unknown  ?q=<name substring>  ?limit=N. Graceful before the first pull."""
    premise = request.args.get("premise")
    q = (request.args.get("q") or "").strip()
    try:
        limit = max(1, min(2000, int(request.args.get("limit", 200))))
    except ValueError:
        limit = 200
    sql, params = "SELECT * FROM t", []
    where = []
    if premise in ("on", "off", "unknown"):
        where.append("premise = ?"); params.append(premise)
    if q:
        where.append("lower(name) LIKE ?"); params.append("%" + q.lower() + "%")
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY name LIMIT ?"; params.append(limit)
    try:
        rows = warehouse.query("orlando_accounts", sql, params)
        return jsonify(ok=True, count=len(rows), remote=warehouse.remote(), accounts=rows)
    except Exception as e:
        # No parquet yet (no pull has run) or storage not reachable — answer gracefully.
        return jsonify(ok=True, count=0, accounts=[], note="no data yet — run the orlando-accounts pull (%s)" % str(e)[:120])


@app.post("/api/places/enrich")
def places_enrich_ep():
    """Match the stored accounts against open POI (FSQ/Overture) and re-write enriched.
    Never blocks the spine — self-reports poi-failed / no-accounts. ?source=fsq|overture."""
    body = request.get_json(force=True, silent=True) or {}
    return jsonify(places.enrich_orlando(source=body.get("source", "fsq")))


# ---- canonical book (star schema): synthetic data foundation for the production app ----
@app.post("/api/seed/build")
def seed_build_ep():
    """Generate the coherent synthetic book (dim_product/account/date + fact_depletion)
    into the warehouse. The data to build the production app against until real facts flow."""
    import seed
    body = request.get_json(force=True, silent=True) or {}
    return jsonify(ok=True, **seed.build(**{k: body[k] for k in ("n_products", "n_accounts", "months", "market") if k in body}))


@app.get("/api/book/cuts")
def book_cuts_ep():
    """Aggregate the book by any dimension — the one query surface every screen shares.
    ?dim=category|channel|price_tier|brand|...  ?measure=revenue|cases|pod."""
    import book
    dim = request.args.get("dim", "category")
    measure = request.args.get("measure", "revenue")
    try:
        return jsonify(ok=True, dim=dim, measure=measure, rows=book.cuts(dim, measure))
    except Exception as e:
        return jsonify(ok=True, dim=dim, measure=measure, rows=[],
                       note="no book yet — POST /api/seed/build (%s)" % str(e)[:120])


_OPENAPI_PATH = os.path.join(APP_DIR, "openapi.yaml")
_openapi_cache = None

@app.get("/api/openapi.json")
def openapi_json():
    """Serve the API contract so the app generates its typed client from the running API
    (single source of truth). Gated like the rest of /api* — codegen runs against the
    ungated local engine in dev."""
    global _openapi_cache
    if _openapi_cache is None:
        import yaml
        with open(_OPENAPI_PATH) as f:
            _openapi_cache = yaml.safe_load(f)
    return jsonify(_openapi_cache)

@app.get("/api/openapi.yaml")
def openapi_yaml():
    return send_file(_OPENAPI_PATH, mimetype="application/yaml")

@app.get("/api/territories")
def territories_ep():
    """Named account groups (Territory Builder). Powers target-vs-competition comparisons."""
    import territories
    return jsonify(ok=True, territories=territories.list_territories())


@app.get("/api/territories/<tid>")
def territory_members_ep(tid):
    import territories
    return jsonify(ok=True, **territories.members(tid))


@app.get("/api/book/summary")
def book_summary_ep():
    import book
    try:
        return jsonify(ok=True, **book.summary())
    except Exception as e:
        return jsonify(ok=True, empty=True, note="no book yet — POST /api/seed/build (%s)" % str(e)[:120])


# ── Label Reader — read ONE PDP/label URL into clean MDM fields, interactively ──
# The human-in-the-loop twin of the catalog scrapers: paste a Total Wine / ABC / any retail product
# URL → fetch (mobile-UA where PerimeterX-walled, BD Unlocker fallback) → every describing field the
# page structures, + optional Claude-vision read of the label image. Reads land in `label_reads`.
@app.post("/api/label/read")
def label_read_ep():
    body = request.get_json(force=True, silent=True) or {}
    url = (body.get("url") or "").strip()
    if not url:
        return jsonify(ok=False, error="url required"), 400
    import label_reader
    try:
        result = label_reader.read(url, vision=bool(body.get("vision")),
                                   land=bool(body.get("land", True)), log=app.logger.info)
    except ValueError as e:               # bad/blocked URL — a user-fixable input error
        return jsonify(ok=False, error=str(e)), 400
    except Exception as e:                # fetch/parse failure
        return jsonify(ok=False, error=str(e)[:220]), 502
    return jsonify(result)


@app.get("/api/label/reads")
def label_reads_ep():
    import label_reader
    try:
        limit = min(200, max(1, int(request.args.get("limit", 50))))
    except Exception:
        limit = 50
    return jsonify(ok=True, reads=label_reader.recent(limit))


# ── Wholesale ordering — distributor menus in, one ordering page, orders fanned back out ──
# The team ask: every distributor emails a wholesale menu (xlsx); parse them ALL into one central
# catalog, let a retailer build a single order across distributors, then return each distributor
# exactly their slice as an order sheet. Menus land in the warehouse (distributor_menu_items via
# menu_ingest); orders are small transactional state → orders.json (the load/_save_json store).
ORDERS = load("orders.json", [])


@app.post("/api/menus/upload")
def menus_upload_ep():
    """Upload one distributor menu file. Body: {filename, b64, distributor?}. Parses (header-row
    detection + column synonyms), lands to distributor_menu_items, returns the parse summary so the
    UI can show what mapped before the retailer trusts it."""
    import base64 as _b64
    import menu_ingest
    body = request.get_json(force=True, silent=True) or {}
    fn = (body.get("filename") or "menu.xlsx").strip()
    try:
        data = _b64.b64decode(body.get("b64") or "")
    except Exception:
        return jsonify(ok=False, error="bad b64"), 400
    if not data:
        return jsonify(ok=False, error="filename + b64 file content required"), 400
    if len(data) > 15 * 1024 * 1024:
        return jsonify(ok=False, error="file too large (15MB max)"), 400
    try:
        # parse_smart = deterministic parse + a Claude fallback that only fires when that's
        # low-confidence (odd layouts still land, no LLM cost on menus that already parse cleanly).
        parsed = menu_ingest.parse_smart(data, filename=fn, distributor=(body.get("distributor") or "").strip())
    except Exception as e:
        return jsonify(ok=False, error="parse failed: %s" % str(e)[:200]), 500
    if not parsed.get("ok"):
        return jsonify(ok=False, error=parsed.get("error", "unparseable")), 422
    n = menu_ingest.land(parsed)
    its = parsed["items"]
    return jsonify(ok=True, menu_id=parsed["menu_id"], distributor=parsed["distributor"],
                   menu_date=parsed["menu_date"], license=parsed["license"], lines=n, method=parsed.get("method"),
                   brands=sorted({i["brand"] for i in its if i["brand"]}),
                   warnings=parsed["warnings"])


@app.post("/api/menus/poll")
def menus_poll_ep():
    """Pull distributor menus straight from the mailbox (IMAP) → parse + land, no manual upload.
    Needs MENU_IMAP_* env; returns mailbox-not-configured otherwise so the UI can hide the button."""
    import menu_mailbox
    body = request.get_json(silent=True) or {}
    if not menu_mailbox.configured():
        return jsonify(ok=False, error="mailbox-not-configured",
                       note="Set MENU_IMAP_HOST/USER/PASS to auto-ingest emailed menus."), 200
    try:
        r = menu_mailbox.poll(limit=min(100, int(body.get("limit", 25))),
                              dry_run=bool(body.get("dry_run")), log=app.logger.info)
    except Exception as e:
        return jsonify(ok=False, error="poll failed: %s" % str(e)[:200]), 502
    return jsonify(r)


@app.get("/api/menus/mailbox")
def menus_mailbox_status_ep():
    import menu_mailbox
    return jsonify(ok=True, configured=menu_mailbox.configured())


@app.get("/api/menus")
def menus_list_ep():
    """Every ingested menu, newest first — plus which one is CURRENT per distributor (the ordering
    page shops current menus only; older ones stay for history/price movement)."""
    try:
        rows = warehouse.query("distributor_menu_items",
                               "SELECT menu_id, distributor, menu_date, license, source_file, "
                               "COUNT(*) AS lines, MAX(ts) AS ts FROM t "
                               "GROUP BY 1,2,3,4,5 ORDER BY menu_date DESC, ts DESC")
    except Exception:
        return jsonify(ok=True, menus=[], note="no menus yet — upload a distributor menu")
    seen = set()
    for r in rows:
        r["current"] = r["distributor"] not in seen
        seen.add(r["distributor"])
    return jsonify(ok=True, menus=rows)


@app.get("/api/menus/items")
def menus_items_ep():
    """The combined catalog the ordering page shops: the CURRENT menu's lines per distributor
    (or one menu via ?menu_id=). ?q= filters name/brand/batch."""
    mid = (request.args.get("menu_id") or "").strip()
    q = (request.args.get("q") or "").strip().lower()
    try:
        if mid:
            rows = warehouse.query("distributor_menu_items",
                                   "SELECT * FROM t WHERE menu_id = ? ORDER BY line", [mid])
        else:
            rows = warehouse.query("distributor_menu_items",
                                   "SELECT t.* FROM t JOIN (SELECT distributor, MAX(menu_date) AS d "
                                   "FROM t GROUP BY 1) cur ON t.distributor = cur.distributor AND "
                                   "t.menu_date = cur.d ORDER BY t.distributor, t.line")
    except Exception:
        return jsonify(ok=True, items=[], note="no menus yet")
    if q:
        rows = [r for r in rows if q in " ".join(str(r.get(k) or "") for k in
                                                 ("product_name", "brand", "batch", "category")).lower()]
    return jsonify(ok=True, items=rows, total=len(rows))


@app.post("/api/orders")
def orders_submit_ep():
    """Submit a retailer order. Body: {retailer, license?, notes?, lines:[{menu_id, line, qty}]}.
    Lines are re-resolved against the WAREHOUSE (price/batch come from the menu of record, never
    the client), grouped per distributor, and persisted. Returns the order incl. the per-
    distributor breakdown the sheets are generated from."""
    body = request.get_json(force=True, silent=True) or {}
    retailer = (body.get("retailer") or "").strip()
    lines = body.get("lines") or []
    if not retailer or not lines:
        return jsonify(ok=False, error="retailer + lines[] required"), 400
    want = {}
    for l in lines:
        try:
            qty = int(l.get("qty") or 0)
        except Exception:
            qty = 0
        if qty > 0 and l.get("menu_id") is not None:
            want[(str(l["menu_id"]), int(l.get("line", -1)))] = qty
    if not want:
        return jsonify(ok=False, error="no lines with qty > 0"), 400
    try:
        mids = sorted({k[0] for k in want})
        rows = warehouse.query("distributor_menu_items",
                               "SELECT * FROM t WHERE menu_id IN (%s)" % ",".join("?" * len(mids)), mids)
    except Exception as e:
        return jsonify(ok=False, error="menu lookup failed: %s" % str(e)[:120]), 500
    by_key = {(str(r["menu_id"]), int(r["line"])): r for r in rows}
    resolved, missing = [], 0
    for k, qty in want.items():
        r = by_key.get(k)
        if not r:
            missing += 1
            continue
        unit = r.get("base_price")
        resolved.append({"menu_id": k[0], "line": k[1], "distributor": r.get("distributor"),
                         "brand": r.get("brand"), "product_name": r.get("product_name"),
                         "batch": r.get("batch"), "size": r.get("size"), "thc_pct": r.get("thc_pct"),
                         "bin_size": r.get("bin_size"), "qty": qty, "unit_price": unit,
                         "line_total": round(unit * qty, 2) if unit is not None else None})
    if not resolved:
        return jsonify(ok=False, error="no order lines resolved against the stored menus"), 422
    dists = {}
    for l in resolved:
        d = dists.setdefault(l["distributor"], {"distributor": l["distributor"], "lines": 0,
                                                "units": 0, "total": 0.0})
        d["lines"] += 1
        d["units"] += l["qty"]
        d["total"] = round(d["total"] + (l["line_total"] or 0), 2)
    oid = "ORD-%s-%03d" % (time.strftime("%Y%m%d"), len(ORDERS) % 1000)
    order = {"id": oid, "retailer": retailer[:120], "license": (body.get("license") or "").strip()[:60],
             "notes": (body.get("notes") or "").strip()[:500], "ts": int(time.time() * 1000),
             "status": "submitted", "lines": resolved, "distributors": sorted(dists.values(),
                                                                              key=lambda d: d["distributor"]),
             "total": round(sum(l["line_total"] or 0 for l in resolved), 2),
             "units": sum(l["qty"] for l in resolved), "unresolved": missing}
    ORDERS.append(order)
    _save_json("orders.json", ORDERS)
    return jsonify(ok=True, order=order)


@app.get("/api/orders")
def orders_list_ep():
    slim = [{k: o[k] for k in ("id", "retailer", "ts", "status", "total", "units")} |
            {"distributors": [d["distributor"] for d in o.get("distributors", [])],
             "lines": len(o.get("lines", []))}
            for o in reversed(ORDERS)]
    return jsonify(ok=True, orders=slim)


@app.get("/api/orders/<oid>")
def order_get_ep(oid):
    o = next((o for o in ORDERS if o["id"] == oid), None)
    return (jsonify(ok=True, order=o) if o else (jsonify(ok=False, error="not found"), 404))


def _po_csv(o, dist):
    """Build the per-distributor PO sheet → (filename, csv_text). Shared by the download + the
    auto-send attachment so both are byte-identical. `dist`='' → the whole order."""
    import csv as _csv
    import io as _io
    lines = [l for l in o["lines"] if not dist or l["distributor"] == dist]
    if not lines:
        return None, None
    buf = _io.StringIO()
    w = _csv.writer(buf)
    w.writerow(["Purchase Order", o["id"]])
    w.writerow(["Retailer", o["retailer"], "License", o.get("license", "")])
    w.writerow(["Distributor", dist or "ALL", "Date", time.strftime("%Y-%m-%d", time.localtime(o["ts"] / 1000))])
    w.writerow([])
    w.writerow(["Product", "Brand", "Batch #", "Size", "THC %", "Case Size", "Qty", "Unit $", "Line Total $"])
    for l in lines:
        w.writerow([l["product_name"], l["brand"], l["batch"], l["size"], l["thc_pct"],
                    l["bin_size"], l["qty"], l["unit_price"], l["line_total"]])
    w.writerow([])
    w.writerow(["", "", "", "", "", "TOTAL", sum(l["qty"] for l in lines), "",
                round(sum(l["line_total"] or 0 for l in lines), 2)])
    fn = "%s_%s.csv" % (o["id"], re.sub(r"[^A-Za-z0-9]+", "_", dist or "all"))
    return fn, buf.getvalue()


@app.get("/api/orders/<oid>/sheet")
def order_sheet_ep(oid):
    """The per-distributor order sheet — ?distributor=X → a CSV of exactly that distributor's slice
    (the artifact that goes back to them). Download form of the same file /api/orders/<id>/send emails."""
    o = next((o for o in ORDERS if o["id"] == oid), None)
    if not o:
        return jsonify(ok=False, error="not found"), 404
    fn, csv_text = _po_csv(o, (request.args.get("distributor") or "").strip())
    if csv_text is None:
        return jsonify(ok=False, error="no lines for that distributor"), 404
    return Response(csv_text, mimetype="text/csv; charset=utf-8",
                    headers={"Content-Disposition": 'attachment; filename="%s"' % fn})


# ── Distributor contact book — where each distributor's order goes back to ──
DISTRIBUTOR_CONTACTS = load("distributor_contacts.json", {})    # {distributor: {email, contact, cc}}


@app.get("/api/distributors")
def distributors_ep():
    """Every distributor we have a menu from, merged with its saved order-contact — so the ordering
    page can show who's ready to auto-send and who still needs an email."""
    known = []
    try:
        known = [r["distributor"] for r in warehouse.query(
            "distributor_menu_items", "SELECT DISTINCT distributor FROM t ORDER BY 1")]
    except Exception:
        known = []
    for d in DISTRIBUTOR_CONTACTS:
        if d not in known:
            known.append(d)
    out = []
    for d in known:
        c = DISTRIBUTOR_CONTACTS.get(d, {})
        out.append({"distributor": d, "email": c.get("email", ""), "contact": c.get("contact", ""),
                    "cc": c.get("cc", ""), "has_email": bool(c.get("email"))})
    return jsonify(ok=True, distributors=out, smtp_configured=bool(os.environ.get("SMTP_HOST", "").strip()))


@app.post("/api/distributors")
def distributor_save_ep():
    """Save/update one distributor's order contact. Body: {distributor, email, contact?, cc?}."""
    b = request.get_json(force=True, silent=True) or {}
    d = (b.get("distributor") or "").strip()
    email = (b.get("email") or "").strip()
    if not d:
        return jsonify(ok=False, error="distributor required"), 400
    if email and not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
        return jsonify(ok=False, error="that doesn't look like a valid email"), 400
    DISTRIBUTOR_CONTACTS[d] = {"email": email, "contact": (b.get("contact") or "").strip()[:80],
                              "cc": (b.get("cc") or "").strip()[:200]}
    _save_json("distributor_contacts.json", DISTRIBUTOR_CONTACTS)
    return jsonify(ok=True, distributor=d, **DISTRIBUTOR_CONTACTS[d])


def _smtp_cfg():
    h = os.environ.get("SMTP_HOST", "").strip()
    if not h:
        return None
    return {"host": h, "port": int(os.environ.get("SMTP_PORT", "587") or 587),
            "user": os.environ.get("SMTP_USER", "").strip(), "pass": os.environ.get("SMTP_PASS", ""),
            "from": os.environ.get("SMTP_FROM", "").strip() or os.environ.get("SMTP_USER", "").strip(),
            "starttls": os.environ.get("SMTP_STARTTLS", "1").strip() not in ("0", "false", "no")}


def _send_po_email(cfg, to, cc, subject, body, fn, csv_text):
    import smtplib
    from email.message import EmailMessage
    msg = EmailMessage()
    msg["From"] = cfg["from"]; msg["To"] = to
    if cc:
        msg["Cc"] = cc
    msg["Subject"] = subject
    msg.set_content(body)
    # Attach as str (not bytes) so the part declares charset=utf-8 — otherwise accented chars /
    # em-dashes in retailer or product names arrive mojibake'd. A str routes through
    # set_text_content(), which takes `subtype` but not `maintype` (text/* is implied).
    msg.add_attachment(csv_text, subtype="csv", filename=fn)
    rcpts = [to] + ([c.strip() for c in cc.split(",") if c.strip()] if cc else [])
    with smtplib.SMTP(cfg["host"], cfg["port"], timeout=30) as s:
        if cfg["starttls"]:
            s.starttls()
        if cfg["user"]:
            s.login(cfg["user"], cfg["pass"])
        s.send_message(msg, from_addr=cfg["from"], to_addrs=rcpts)


@app.post("/api/orders/<oid>/send")
def order_send_ep(oid):
    """Email the PO sheet back to each distributor on the order (or just ?distributor=/body distributor).
    Needs SMTP_* env + a saved contact email per distributor. Returns per-distributor results; the UI
    falls back to download + copy-email for any that aren't send-ready. Records sends on the order."""
    o = next((o for o in ORDERS if o["id"] == oid), None)
    if not o:
        return jsonify(ok=False, error="not found"), 404
    b = request.get_json(force=True, silent=True) or {}
    only = (b.get("distributor") or "").strip()
    dists = [d["distributor"] for d in o.get("distributors", []) if not only or d["distributor"] == only]
    if not dists:
        return jsonify(ok=False, error="no matching distributor on this order"), 400
    cfg = _smtp_cfg()
    if not cfg:
        return jsonify(ok=False, error="email-not-configured",
                       note="Set SMTP_HOST/PORT/USER/PASS/FROM to enable auto-send. Until then use the "
                            "PO-sheet download + copy-email.",
                       needs_contact=[d for d in dists if not DISTRIBUTOR_CONTACTS.get(d, {}).get("email")]), 200
    results, sent_log = [], o.get("sent", [])
    for d in dists:
        c = DISTRIBUTOR_CONTACTS.get(d, {})
        to = c.get("email", "")
        if not to:
            results.append({"distributor": d, "ok": False, "reason": "no-contact"})
            continue
        fn, csv_text = _po_csv(o, d)
        dtot = next((x for x in o["distributors"] if x["distributor"] == d), {})
        body = ("Purchase order %s from %s%s.\n\n%s lines · %s units · $%s.\n\nThe PO sheet is attached "
                "as a CSV.%s\n\n— sent via Hoodie Suite Wholesale Ordering" % (
                    o["id"], o["retailer"], (" (%s)" % o["license"]) if o.get("license") else "",
                    dtot.get("lines", "?"), dtot.get("units", "?"), dtot.get("total", "?"),
                    ("\n\nNotes: " + o["notes"]) if o.get("notes") else ""))
        try:
            _send_po_email(cfg, to, c.get("cc", ""), "PO %s — %s" % (o["id"], o["retailer"]), body, fn, csv_text)
            rec = {"distributor": d, "to": to, "at": int(time.time() * 1000)}
            sent_log = [s for s in sent_log if s.get("distributor") != d] + [rec]
            results.append({"distributor": d, "ok": True, "to": to})
        except Exception as e:
            results.append({"distributor": d, "ok": False, "reason": str(e)[:160]})
    o["sent"] = sent_log
    if any(r["ok"] for r in results):
        o["status"] = "sent" if all(r["ok"] for r in results) else "partially-sent"
    _save_json("orders.json", ORDERS)
    return jsonify(ok=any(r["ok"] for r in results), results=results, order_status=o.get("status"))


# Order lifecycle: submitted → sent → confirmed (distributor acknowledged) → delivered; cancellable
# until delivered. A forward-only ladder (no going back a step) keeps the history a clean audit trail.
_ORDER_FLOW = ["submitted", "sent", "confirmed", "delivered"]


@app.post("/api/orders/<oid>/status")
def order_status_ep(oid):
    """Advance/set an order's status. Body: {status}. Allowed: submitted|sent|confirmed|delivered|
    cancelled. Records a timestamped status_history so the lifecycle is auditable."""
    o = next((o for o in ORDERS if o["id"] == oid), None)
    if not o:
        return jsonify(ok=False, error="not found"), 404
    new = (request.get_json(force=True, silent=True) or {}).get("status", "").strip().lower()
    if new not in _ORDER_FLOW and new != "cancelled":
        return jsonify(ok=False, error="status must be one of: %s, cancelled" % ", ".join(_ORDER_FLOW)), 400
    cur = o.get("status", "submitted")
    if cur == "delivered" and new != "delivered":
        return jsonify(ok=False, error="a delivered order is closed"), 409
    if new != "cancelled" and cur in _ORDER_FLOW and new in _ORDER_FLOW and \
            _ORDER_FLOW.index(new) < _ORDER_FLOW.index(cur):
        return jsonify(ok=False, error="can't move an order backward (%s → %s)" % (cur, new)), 409
    o["status"] = new
    o.setdefault("status_history", []).append({"status": new, "at": int(time.time() * 1000)})
    _save_json("orders.json", ORDERS)
    return jsonify(ok=True, id=oid, status=new, status_history=o["status_history"])


# ---- optional: serve the static suite from THIS app (all-in-one image, e.g. Fly.io) ----
# When SUITE_ROOT is set, one gunicorn process serves BOTH /api/* and the public suite from a single
# origin, so the apps' same-origin /api/* fetches work with no separate frontend host and no CORS.
# An ALLOWLIST of public top-level entries is enforced so the engine (unifyd/, *.py, scripts/, docs,
# dotfiles, .env) is NEVER web-served — it mirrors the deploy.sh / deploy.yml exclude lists.
SUITE_ROOT = os.environ.get("SUITE_ROOT", "").strip()
_SUITE_OK_TOP = {"index.html", "hub.html", "apps", "spine", "suite.css", "suite-header.js",
                 "suite-export.js", "fullread.js", "dq.js", "dq_frontier.js", "datagrid.js",
                 "apps.registry.json", "favicon.ico"}

def _suite_send(relpath):
    from flask import abort
    if not SUITE_ROOT:
        abort(404)
    root = os.path.abspath(SUITE_ROOT)
    full = os.path.normpath(os.path.join(root, (relpath or "").lstrip("/")))
    if not (full == root or full.startswith(root + os.sep)):
        abort(403)                                   # no traversal outside the suite root
    # Allowlist the RESOLVED top-level segment (after collapsing ..), so `/apps/../unifyd/x` can't
    # hop from an allowed subtree into the engine. Engine + dotfiles + secrets are never web-served.
    rel = os.path.relpath(full, root)
    top = rel.split(os.sep, 1)[0]
    if top in ("", ".", "..") or top not in _SUITE_OK_TOP:
        abort(404)
    if os.path.isfile(full):
        if full.endswith(".webmanifest"):
            return send_file(full, mimetype="application/manifest+json")
        resp = send_file(full)
        # HTML/JS must REVALIDATE every load (ETag conditional) so a deploy shows immediately instead of
        # being served stale from the browser cache — the "I deployed but it looks unchanged" trap.
        if full.endswith((".html", ".js")):
            resp.headers["Cache-Control"] = "no-cache, must-revalidate"
        return resp
    abort(404)

@app.get("/prism")
@app.get("/prism.html")
def prism_shortcut():
    return redirect("/apps/prism.html")              # friendly short URL for the mobile app

@app.get("/hub")
def hub_shortcut():
    return _suite_send("hub.html")                    # the four-bucket hub (also the homepage below)

@app.get("/home")
def home_console():
    return _suite_send("hub.html")                    # same shell, "console" mode (AI landing + app-launcher + fav dock) — hub.html branches on pathname /home

@app.get("/classic")
@app.get("/launcher")
def classic_launcher():
    return _suite_send("index.html")                 # the previous flat-grid launcher — kept, one-line reversible

@app.get("/")
def index():
    if SUITE_ROOT:
        return _suite_send("hub.html")               # the four-bucket hub is now the homepage (was index.html; still at /classic)
    return send_file(HTML_PATH)

@app.get("/<path:relpath>")
def suite_static(relpath):
    return _suite_send(relpath)                       # /api/* rules are more specific and win over this

def _monitor_warm():
    """Ensure the Data Console manifest is built + kept fresh in the background (build at boot from a persisted
    snapshot if present, else a background build), so /api/monitor is always served from cache — never a cold
    ~130s live sweep of object storage on a user request."""
    import monitor
    monitor.warm()
    monitor.snapshot()                                   # serves cache + kicks a background refresh when stale
    try:
        _mon_con()                                       # pay the ~6s httpfs/S3 setup at boot, not on first click
    except Exception:
        pass
    try:
        monitor.warm_search_pool()                       # warm the 8-connection search pool so first search is fast
    except Exception:
        pass


def _wb_warm():
    """Keep the heavy read caches WARM in the background so no user (or the first after a rebuild) hits a cold
    scan. At the larger master scale the cold cost is real: the 1M-row name maps (~17s), the workbench views, the
    coverage/catalog rollups. Each is rebuild-invalidated (or short-TTL), so we re-warm on a loop — after a rebuild
    the caches repopulate before the next user arrives, and steady-state re-warms return instantly from cache."""
    import time as _t
    _t.sleep(2)                                          # let the app finish importing
    warmers = [
        _wb_maps,                                        # 1M-row product/brand name maps (the big shared cost)
        lambda: [_wb_view(v) for v in ("wb_summary", "wb_master", "wb_queue", "wb_merges")],
        lambda: _ttl("wb_summary_resp", 90, _wb_summary_data),   # the funnel summary (workbench landing)
        lambda: _ttl("wb_review_resp", 90, lambda: _wb_review_data("")),   # analyst queue (default, no filter)
        lambda: _ttl("wb_master_resp", 90, lambda: _wb_master_data("", "")),   # trusted master (default)
        lambda: _ttl("wb_merges_resp", 90, lambda: _wb_merges_data("")),   # merge queue (default)
        _wb_merge_upc_attrs,                              # per-member attrs for cluster-open (was a scan/click)
        _wb_prod_items,                                  # product->items reverse index (candidates modal)
        _wb_detail,                                      # bulk detail for queue/master items (item-detail modal)
        lambda: _ttl("img_matches_resp", 120, _img_matches_data),   # image-similarity candidate queue
        lambda: _ttl("skus_resp_50", 90, lambda: _skus_data("", 0, 50)),        # catalog landing page
        lambda: _ttl("outlets_resp_50", 90, lambda: _outlets_data("", "", 0, 50)),  # outlets landing page
        lambda: _ttl("cov_accounts_300", 90, lambda: _cov_accounts_data("", "", "", "", "", 300)),  # coverage list
        lambda: _ttl("cov_dots", 90, lambda: _cov_dots_data("", "")),           # coverage full-book dot layer
        lambda: _ttl("outlets_geo", 90, lambda: _outlets_geo_data("", "", "")),  # workbench outlet map
        lambda: [_wq_cached(t, "SELECT count(*) c FROM t") for t in           # catalog header counts
                 ("dim_brand", "dim_product", "dim_item", "dim_sku", "dim_supplier",
                  "dim_outlet_resolved", "fact_inventory", "fact_pricing")],
        lambda: _ttl("coverage", 90, _coverage_data),
        lambda: _ttl("catalog", 90, _catalog_map),
        lambda: _ttl("source_spec", 600, _source_spec_data),
        _monitor_warm,                                   # Data Console manifest — build at boot, keep warm (never
                                                         # let a user hit the cold ~130s object-storage sweep)
    ]
    n = 0
    while True:
        for w in warmers:
            # source_spec is a heavy full scan; warm it only occasionally (its own 10-min TTL covers the rest).
            if w is warmers[-1] and n % 6 != 0:
                continue
            try:
                w()
            except Exception:
                pass
        n += 1
        _t.sleep(75)                                     # < the 90s TTL, so the short-TTL rollups stay warm


def _monitor_sample_loop():
    """Continuously (re)build the Data Console's precomputed preview samples in the background — a few datasets per
    pass, only those whose data changed — so the drawer serves from memory (instant) at any dataset size. Runs on
    its own thread so it never blocks the other warmers; incremental + capped so cost stays flat as data grows."""
    import time as _t
    import monitor
    _t.sleep(8)                                          # let the manifest warm first (samples read its source list)
    while True:
        try:
            monitor.refresh_samples()
        except Exception:
            pass
        _t.sleep(25)


try:
    threading.Thread(target=_monitor_sample_loop, daemon=True).start()
except Exception:
    pass

try:
    threading.Thread(target=_wb_warm, daemon=True).start()
except Exception:
    pass


if __name__ == "__main__":
    _port = int(os.environ.get("PORT", "8765"))          # PORT override lets a second dev agent run beside 8765
    print("Unifyd agent on http://127.0.0.1:%d  (Ctrl-C to stop)" % _port)
    app.run(host="127.0.0.1", port=_port, debug=False, threaded=True)
