#!/usr/bin/env python3
"""xsource_gold.py — build the HUMAN-labelled gold set that `xsource_match` needs.

WHY THIS IS NEEDED AT ALL
  The cross-source merge scored precision 0.233 against UPC-derived gold, and that number cannot be
  trusted in either direction: 59,455 of 67,099 rows were unscoreable because the sources that most
  need merging (binnys, abc, total-wine) carry no UPC, and a `resolved_id` can legitimately span
  several UPCs, which makes UPC-disagreement a harsher test than "different item". So the matcher is
  currently unmeasured, not proven bad — and an unmeasured matcher cannot ship
  ([[matching-at-scale]]: humans handle exceptions, machines handle scale).

THE SHAPE, WHICH IS THE PLATFORM'S EXISTING PATTERN
  Precompute candidates → a model pre-adjudicates → a human confirms a prioritized queue. Three
  rules make the resulting gold trustworthy rather than circular:

  1. **STRATIFIED, AND SCORED PER STRATUM.** Sampling only the pairs the rule merges measures
     precision and tells you nothing about recall. So the sheet mixes `merged` (what the rule
     claims), `near_miss` (same brand+size, names differ — what it declined), and `control` (same
     brand, DIFFERENT size — which must always be NO). Blending the strata into one accuracy number
     would hide which half is broken, so `score()` reports them separately.

  2. **THE MODEL'S OPINION NEVER OCCUPIES THE ANSWER COLUMN.** `suggested` sits in its own column
     with its reason; `label` ships EMPTY. A pre-filled answer column produces rubber-stamping, and
     a gold set that agrees with the model by construction measures nothing. The model is there to
     order the queue and explain itself, not to answer.

  3. **THE CONTROL STRATUM IS THE LABELLER'S OWN CHECK.** Controls have a known answer (different
     size = different item). A labelled sheet whose controls come back wrong is a sheet filled in
     carelessly, and `ingest()` reports that rather than folding it into the score.

USAGE
    python xsource_gold.py export --n 300 --out /tmp/gold.xlsx    # build the sheet
    #   ... a human fills the `label` column with y / n / ? ...
    python xsource_gold.py ingest --path /tmp/gold.xlsx           # land it
    python xsource_gold.py score                                  # measure the matcher on it
"""
import argparse
import csv
import json
import os
import random
import sys
import time

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import xsource_match as xm  # noqa: E402

TABLE = "xsource_gold"
FIELDS = ["pair_id", "stratum", "a_id", "a_source", "a_brand", "a_name", "a_size", "a_upc",
          "b_id", "b_source", "b_brand", "b_name", "b_size", "b_upc",
          "rule_merges", "suggested", "suggest_reason", "label", "labelled_by", "labelled_at",
          "sample_seed", "built_at"]

# Columns the human sees, in the order they help. `label` is last and EMPTY.
SHEET_COLS = ["pair_id", "stratum", "a_source", "a_brand", "a_name", "a_size",
              "b_source", "b_brand", "b_name", "b_size", "suggested", "suggest_reason", "label"]

VALID = {"y": True, "yes": True, "1": True, "same": True,
         "n": False, "no": False, "0": False, "diff": False, "different": False}


def _pid(a, b):
    import hashlib
    return "P" + hashlib.sha1(("%s|%s" % tuple(sorted([str(a), str(b)]))).encode()).hexdigest()[:12]


def candidates(rows, n=300, seed=7, log=print):
    """Stratified candidate pairs from master rows. Reproducible for a given seed, so a re-export
    produces the SAME sheet — a labeller who has done half a sheet must not be handed a new one."""
    rnd = random.Random(seed)
    by_sig, by_brand_size = {}, {}
    for r in rows:
        sig = xm.signature(r.get("brand"), r.get("name"), r.get("size"))
        if not sig:
            continue
        by_sig.setdefault(sig, []).append(r)
        by_brand_size.setdefault((sig[0], sig[2]), []).append((sig[1], r))

    merged, near, control = [], [], []
    # MERGED — pairs the rule actually unites. Measures precision.
    for sig, members in by_sig.items():
        ids = {m.get("resolved_id") for m in members if m.get("resolved_id")}
        if len(ids) < 2:
            continue
        ms = sorted(members, key=lambda m: str(m.get("resolved_id")))
        for i in range(len(ms) - 1):
            if ms[i].get("resolved_id") != ms[i + 1].get("resolved_id"):
                merged.append((ms[i], ms[i + 1], "merged"))
    # NEAR MISS — same brand and size, DIFFERENT name signature. Measures recall: these are the
    # merges the rule declined, and some of them are real.
    for (bk, sz), items in by_brand_size.items():
        sigs = {}
        for ns, r in items:
            sigs.setdefault(ns, []).append(r)
        keys = sorted(sigs)
        for i in range(len(keys)):
            for j in range(i + 1, min(i + 3, len(keys))):
                near.append((sigs[keys[i]][0], sigs[keys[j]][0], "near_miss"))
    # CONTROL — the SAME PRODUCT at a DIFFERENT SIZE. The answer is known (no: item grain is
    # product+size), and it is the subtle case: "Absolut Citron 750ml" vs "Absolut Citron 1750ml"
    # is where a careless labeller says "same".
    #
    # The first cut grouped only by brand and took whatever product came first per size, which
    # produced "Shipyard Blood Orange vs Shipyard Smashed Pumpkin" — obviously different, so a
    # labeller gets it right without reading, and the control audits nothing.
    by_prod = {}
    for (bk, sz), items in by_brand_size.items():
        for ns, r in items:
            by_prod.setdefault((bk, ns), {}).setdefault(sz, r)
    for (bk, ns), by_sz in by_prod.items():
        if len(by_sz) >= 2:
            szs = sorted(by_sz)
            control.append((by_sz[szs[0]], by_sz[szs[-1]], "control"))

    rnd.shuffle(merged); rnd.shuffle(near); rnd.shuffle(control)
    take = {"merged": int(n * 0.5), "near_miss": int(n * 0.35), "control": n - int(n * 0.5) - int(n * 0.35)}
    picked = merged[:take["merged"]] + near[:take["near_miss"]] + control[:take["control"]]
    log("candidates: merged=%d near_miss=%d control=%d (pool %d/%d/%d)"
        % (min(len(merged), take["merged"]), min(len(near), take["near_miss"]),
           min(len(control), take["control"]), len(merged), len(near), len(control)))

    now = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
    out = []
    for a, b, stratum in picked:
        rule = xm.merge([a, b])
        out.append({
            "pair_id": _pid(a.get("resolved_id"), b.get("resolved_id")), "stratum": stratum,
            "a_id": a.get("resolved_id"), "a_source": a.get("source"), "a_brand": a.get("brand"),
            "a_name": a.get("name"), "a_size": a.get("size"), "a_upc": a.get("upc"),
            "b_id": b.get("resolved_id"), "b_source": b.get("source"), "b_brand": b.get("brand"),
            "b_name": b.get("name"), "b_size": b.get("size"), "b_upc": b.get("upc"),
            "rule_merges": bool(rule), "suggested": "", "suggest_reason": "",
            "label": "", "labelled_by": "", "labelled_at": "",
            "sample_seed": seed, "built_at": now})
    return out


def adjudicate(pairs, log=print):
    """Optional pre-adjudication. Fills `suggested` + `suggest_reason` ONLY — never `label`.
    Off unless XSOURCE_LLM=1 and a key is present; a missing model leaves the columns blank rather
    than degrading the sheet."""
    if os.environ.get("XSOURCE_LLM", "0") != "1" or not os.environ.get("ANTHROPIC_API_KEY"):
        log("adjudicate: off (set XSOURCE_LLM=1 with a key) — the sheet ships without suggestions")
        return pairs
    try:
        import anthropic
        client = anthropic.Anthropic()
    except Exception as e:
        log("adjudicate: unavailable (%s)" % str(e)[:80])
        return pairs
    TOOL = {"name": "judge", "description": "Are these two retail listings the SAME item?",
            "input_schema": {"type": "object", "properties": {
                "same": {"type": "boolean"}, "reason": {"type": "string"}}, "required": ["same"]}}
    for p in pairs:
        try:
            msg = client.messages.create(
                model="claude-sonnet-5", max_tokens=200, tools=[TOOL],
                tool_choice={"type": "tool", "name": "judge"},
                messages=[{"role": "user", "content":
                           "Same item (same product AND same size)?\nA: %s | %s | %s\nB: %s | %s | %s"
                           % (p["a_brand"], p["a_name"], p["a_size"],
                              p["b_brand"], p["b_name"], p["b_size"])}])
            for b in msg.content:
                if getattr(b, "type", "") == "tool_use":
                    p["suggested"] = "y" if b.input.get("same") else "n"
                    p["suggest_reason"] = (b.input.get("reason") or "")[:180]
        except Exception as e:
            p["suggest_reason"] = "adjudication failed: %s" % str(e)[:60]
    return pairs


def export(pairs, path, log=print):
    """Write the labelling sheet. .xlsx via the repo's own writer, .csv otherwise."""
    if path.endswith(".xlsx"):
        import xlsx_write
        wb = xlsx_write.Workbook()
        sh = wb.sheet("label")
        sh.header(SHEET_COLS, widths={2: 12, 3: 18, 4: 46, 6: 12, 7: 18, 8: 46, 11: 40, 12: 10})
        for p in pairs:
            sh.row([p.get(c, "") for c in SHEET_COLS])
        g = wb.sheet("guide")
        g.header(["how to label"])
        for line in ["Put y / n / ? in the LAST column (label).",
                     "y = the SAME item: same product AND same size.",
                     "n = different item (different product, or different size).",
                     "? = you cannot tell from what is shown. '?' is a real answer — use it.",
                     "",
                     "stratum=control rows are a check on the sheet: they are same-brand,",
                     "DIFFERENT-size pairs, so the answer is always n. If yours come back y,",
                     "ingest will tell you the sheet was filled in too fast.",
                     "",
                     "The `suggested` column is a machine guess. Ignore it if it is unhelpful —",
                     "it is deliberately NOT in the answer column."]:
            g.row([line])
        with open(path, "wb") as f:
            f.write(wb.tobytes())
    else:
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=SHEET_COLS, extrasaction="ignore")
            w.writeheader()
            for p in pairs:
                w.writerow({c: p.get(c, "") for c in SHEET_COLS})
    log("wrote %d pairs -> %s" % (len(pairs), path))
    return path


def read_labels(path):
    """Read a filled sheet back → {pair_id: bool-or-None}. Unrecognised values are None ('?'),
    never coerced to a judgement."""
    out = {}
    if path.endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sh = wb[wb.sheetnames[0]]
        rows = list(sh.iter_rows(values_only=True))
        head = [str(c or "").strip() for c in rows[0]]
        idx = {c: head.index(c) for c in SHEET_COLS if c in head}
        for r in rows[1:]:
            pid = str(r[idx["pair_id"]] or "").strip()
            if pid:
                out[pid] = VALID.get(str(r[idx["label"]] or "").strip().lower())
    else:
        with open(path, encoding="utf-8") as f:
            for r in csv.DictReader(f):
                if r.get("pair_id"):
                    out[r["pair_id"].strip()] = VALID.get((r.get("label") or "").strip().lower())
    return out


def ingest(path, pairs=None, labelled_by="", land=True, log=print):
    """Merge a filled sheet into the gold table, and AUDIT it via the control stratum."""
    labels = read_labels(path)
    if pairs is None:
        try:
            import warehouse
            pairs = warehouse.query(TABLE, "SELECT * FROM t")
        except Exception as e:
            log("ingest: no candidate set to attach labels to (%s)" % str(e)[:80])
            return [], {"status": "degraded"}
    now = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
    n_lab = 0
    for p in pairs:
        v = labels.get(p["pair_id"])
        if v is None and p["pair_id"] not in labels:
            continue
        p["label"] = "" if v is None else ("y" if v else "n")
        p["labelled_by"], p["labelled_at"] = labelled_by, now
        n_lab += 1 if v is not None else 0

    ctrl = [p for p in pairs if p["stratum"] == "control" and p["label"]]
    ctrl_bad = [p for p in ctrl if p["label"] == "y"]
    rep = {"labelled": n_lab, "unlabelled": sum(1 for p in pairs if not p["label"]),
           "controls": len(ctrl), "controls_wrong": len(ctrl_bad),
           "control_accuracy": round(1 - len(ctrl_bad) / len(ctrl), 3) if ctrl else None}
    log("ingest: %d labelled, %d left blank | controls %d/%d correct"
        % (n_lab, rep["unlabelled"], len(ctrl) - len(ctrl_bad), len(ctrl)))
    if ctrl_bad:
        log("  WARNING: %d control pairs (same brand, DIFFERENT size) were labelled SAME. Those "
            "have a known answer, so this sheet was filled in too quickly to trust." % len(ctrl_bad))
    if land and pairs:
        try:
            import warehouse
            warehouse.write_accumulate(TABLE, pairs, key="pair_id", fields=FIELDS, coverage=False)
            log("landed %s: %d rows" % (TABLE, len(pairs)))
        except Exception as e:
            log("%s land skipped: %s" % (TABLE, str(e)[:90]))
    return pairs, rep


def score(pairs, log=print):
    """Score the matcher against the human labels — PER STRATUM, never blended."""
    lab = [p for p in pairs if p.get("label") in ("y", "n")]
    if not lab:
        log("score: nothing labelled yet")
        return {"status": "no_labels"}
    out = {}
    for st in ("merged", "near_miss", "control", "ALL"):
        sel = lab if st == "ALL" else [p for p in lab if p["stratum"] == st]
        if not sel:
            continue
        tp = sum(1 for p in sel if p["rule_merges"] and p["label"] == "y")
        fp = sum(1 for p in sel if p["rule_merges"] and p["label"] == "n")
        fn = sum(1 for p in sel if not p["rule_merges"] and p["label"] == "y")
        tn = sum(1 for p in sel if not p["rule_merges"] and p["label"] == "n")
        out[st] = {"n": len(sel), "tp": tp, "fp": fp, "fn": fn, "tn": tn,
                   "precision": round(tp / (tp + fp), 4) if (tp + fp) else None,
                   "recall": round(tp / (tp + fn), 4) if (tp + fn) else None}
    for st, m in out.items():
        log("  %-10s n=%-4d precision=%-7s recall=%-7s (tp %d fp %d fn %d tn %d)"
            % (st, m["n"], m["precision"], m["recall"], m["tp"], m["fp"], m["fn"], m["tn"]))
    return out


def main(argv=None):
    ap = argparse.ArgumentParser(description="Build/ingest the human gold set for xsource_match.")
    ap.add_argument("cmd", choices=["export", "ingest", "score"])
    ap.add_argument("--n", type=int, default=300)
    ap.add_argument("--seed", type=int, default=7)
    ap.add_argument("--out", default="/tmp/xsource_gold.xlsx")
    ap.add_argument("--path")
    ap.add_argument("--by", default="")
    a = ap.parse_args(argv)
    import warehouse
    if a.cmd == "export":
        rows = warehouse.query("_stage_product",
                               "SELECT item_key AS resolved_id, brand, product_name AS name, "
                               "size_raw AS size, NULL AS upc, source FROM t")
        pairs = adjudicate(candidates(rows, n=a.n, seed=a.seed))
        try:
            warehouse.write_accumulate(TABLE, pairs, key="pair_id", fields=FIELDS, coverage=False)
        except Exception as e:
            print("(candidate set not landed: %s)" % str(e)[:70])
        export(pairs, a.out)
    elif a.cmd == "ingest":
        ingest(a.path or a.out, labelled_by=a.by)
    else:
        print(json.dumps(score(warehouse.query(TABLE, "SELECT * FROM t")), indent=2))


if __name__ == "__main__":
    main()
