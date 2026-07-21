"""analyze.py — the data-reader brain behind "Overlay your data".

Given an unfamiliar dataset (any of 5 verticals: bev-alc, hemp, cannabis, broader CPG,
supplemental/census-panel), it produces a context-aware first pass the way the user reads
data: (1) what is this & can I trust it, (2) headline KPIs, (3) what's anomalous, (4)
findings in the house style (caution/neutral/opportunity), each with a JUSTIFICATION for
the measures + dimensions chosen — and, when the data maps onto the bev-alc Report Builder
model, ready-to-materialize report specs so reports get built through the builder for
consistency.

The heavy lifting (column profiling + aggregation) happens deterministically in Python;
the LLM does the interpretation + layout. OFF unless ANTHROPIC_API_KEY is set (and
`anthropic` is imported lazily), so the agent has no new hard dependency.
"""
import json, os, re, statistics

MODEL = os.environ.get("AGENT_LLM_MODEL", "claude-opus-4-8")


def enabled():
    return bool(os.environ.get("ANTHROPIC_API_KEY"))


# ---------------- deterministic profiling (no LLM) ----------------
_NUM_RE  = re.compile(r"^-?\$?\d[\d,]*\.?\d*%?$")
_DATE_RE = re.compile(r"^\d{1,4}[-/]\d{1,2}([-/]\d{1,4})?$|^\d{4}$")


def _num(v):
    try:
        return float(str(v).replace(",", "").replace("$", "").replace("%", ""))
    except (ValueError, AttributeError):
        return None


def parse_upload(filename, raw, scan=15):
    """Parse an uploaded CSV/TSV or XLSX into (header, rows), auto-detecting the header row so files with a
    title/preamble (common in exported menus & inventory — e.g. a dispensary menu whose header is on row 3)
    still read correctly. Returns {header, rows, header_row, sheet}. stdlib for CSV; openpyxl (lazy) for XLSX."""
    import io
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        grid = [[("" if c is None else c) for c in row] for row in ws.iter_rows(values_only=True)]
        sheet = wb.sheetnames[0]
    else:
        import csv
        text = raw.decode("utf-8", "replace") if isinstance(raw, (bytes, bytearray)) else raw
        head = text[:4000]
        delim = "\t" if (name.endswith(".tsv") or head.count("\t") > head.count(",")) else ","
        grid = list(csv.reader(io.StringIO(text), delimiter=delim))
        sheet = None
    grid = [r for r in grid if r is not None]
    if not grid:
        return {"header": [], "rows": [], "header_row": 0, "sheet": sheet}

    def _score(row):                                     # label-ish header score: non-empty, short, non-numeric
        n = 0
        for c in row:
            s = str(c).strip()
            if not s:
                continue
            try:
                float(s.replace(",", "").replace("$", "").replace("%", "")); continue
            except Exception:
                pass
            if len(s) <= 40:
                n += 1
        return n
    hi = max(range(min(scan, len(grid))), key=lambda i: (_score(grid[i]), -i))
    header = [str(c).strip() for c in grid[hi]]
    keep = [j for j, h in enumerate(header) if h]        # drop columns with an empty header (trims blank margins)
    header = [header[j] for j in keep]
    rows = []
    row_src = []
    for off, r in enumerate(grid[hi + 1:]):
        vals = [str(r[j]).strip() if j < len(r) else "" for j in keep]
        if any(vals):
            rows.append(vals)
            row_src.append(hi + 1 + off)          # physical 0-based grid row — for faithful write-back/refill
    return {"header": header, "rows": rows, "row_src": row_src, "col_src": keep,  # keep = parsed col → physical 0-based col
            "header_row": hi, "sheet": sheet}


def profile_columns(header, rows, top_k=8):
    """Per-column type/cardinality/null-rate/stats/top-values — the structured summary the
    LLM reads. Works on any tabular data, cheap, stdlib-only."""
    n = len(rows)
    cols = []
    for ci, name in enumerate(header):
        vals = [(r[ci] if ci < len(r) else "") for r in rows]
        nonblank = [v for v in vals if str(v).strip() != ""]
        nums = [x for x in (_num(v) for v in nonblank) if x is not None]
        is_num = len(nonblank) > 0 and len(nums) / len(nonblank) >= 0.8
        is_date = (not is_num) and len(nonblank) > 0 and \
            sum(1 for v in nonblank if _DATE_RE.match(str(v).strip())) / len(nonblank) >= 0.8
        freq = {}
        for v in nonblank:
            key = str(v).strip()
            freq[key] = freq.get(key, 0) + 1
        top = sorted(freq.items(), key=lambda kv: -kv[1])[:top_k]
        col = {"name": name, "type": "number" if is_num else "date" if is_date else "string",
               "null_rate": round(1 - len(nonblank) / n, 3) if n else 1.0,
               "distinct": len(freq),
               "top_values": [{"v": v, "count": c} for v, c in top]}
        if is_num and nums:
            col["stats"] = {"min": round(min(nums), 4), "max": round(max(nums), 4),
                            "mean": round(statistics.fmean(nums), 4),
                            "median": round(statistics.median(nums), 4)}
        cols.append(col)
    return {"row_count": n, "column_count": len(header), "columns": cols}


# ---------------- the reading style + schema (system prompt) ----------------
SYSTEM = """You are Hoodie's data reader. An analyst hands you a dataset they've never seen \
and you give a sharp first pass — in the exact order they read data:

1. WHAT IS THIS & CAN I TRUST IT — say what the dataset is, which vertical it belongs to \
(bev-alc, hemp, cannabis, broader CPG, or supplemental like census/panel), what one row \
represents (the grain), the time span, and data-quality flags (nulls, duplicates, \
outliers, suspicious or mislabeled columns, too-few rows). Be honest about confidence.
2. HEADLINE KPIs — the 3-5 numbers that matter for THIS vertical, computed from the \
profile. Each with a one-line "why it matters".
3. WHAT'S ANOMALOUS / WHAT CHANGED — what's surprising: concentration, outliers, gaps vs \
expectation. Lead with what deserves a second look.
4. FINDINGS in the house style — short, each tagged priority high|positive|caution|neutral, \
written as {strong headline}{supporting detail}. Mirror an experienced analyst's voice.

ALWAYS justify the measures and dimensions you pick — say WHY each one earns a place. \
Analyze the data at its ACTUAL grain — including transaction/row-level — and surface the \
insight; do NOT refuse or withhold analysis because the data is granular. Where data is \
account- or person-level, add an ADVISORY note that any externally SHARED output should \
roll up to groups of ≥10 — but this is a flag the user can override, not a stop.

REPORT BUILDER: if (and only if) the data clearly maps onto the bev-alc cell model — \
category × sub-channel × census-division with volume/POD/share-style measures — set \
report_builder.fits=true and propose report specs using ONLY the provided dimension keys, \
measure keys, and viz types. Otherwise set fits=false with a one-line reason; the universal \
read above stands on its own for other verticals. Never invent dimension/measure keys."""

# JSON shape the model must return (kept permissive so it validates reliably).
SCHEMA = {
    "type": "object", "additionalProperties": False,
    "required": ["identity", "trust", "kpis", "dimensions", "measures", "anomalies",
                 "findings", "suggested_questions", "report_builder"],
    "properties": {
        "identity": {"type": "object", "additionalProperties": True, "properties": {
            "summary": {"type": "string"},
            "vertical": {"type": "string",
                         "enum": ["bev-alc", "hemp", "cannabis", "cpg", "supplemental", "unknown"]},
            "grain": {"type": "string"}, "time_span": {"type": "string"},
            "confidence": {"type": "string", "enum": ["high", "medium", "low"]}}},
        "trust": {"type": "object", "additionalProperties": True, "properties": {
            "assessment": {"type": "string"},
            "flags": {"type": "array", "items": {"type": "object", "additionalProperties": True}}}},
        "kpis": {"type": "array", "items": {"type": "object", "additionalProperties": True,
            "properties": {"label": {"type": "string"}, "value": {"type": "string"},
                           "unit": {"type": "string"}, "why": {"type": "string"}}}},
        "dimensions": {"type": "array", "items": {"type": "object", "additionalProperties": True,
            "properties": {"column": {"type": "string"}, "role": {"type": "string"},
                           "why": {"type": "string"}}}},
        "measures": {"type": "array", "items": {"type": "object", "additionalProperties": True,
            "properties": {"column": {"type": "string"}, "agg": {"type": "string"},
                           "why": {"type": "string"}}}},
        "anomalies": {"type": "array", "items": {"type": "object", "additionalProperties": True}},
        "findings": {"type": "array", "items": {"type": "object", "additionalProperties": True,
            "properties": {"strong": {"type": "string"}, "rest": {"type": "string"},
                           "priority": {"type": "string",
                                        "enum": ["high", "positive", "caution", "neutral"]}}}},
        "suggested_questions": {"type": "array", "items": {"type": "string"}},
        "report_builder": {"type": "object", "additionalProperties": True, "properties": {
            "fits": {"type": "boolean"}, "reason": {"type": "string"},
            "reports": {"type": "array", "items": {"type": "object", "additionalProperties": True,
                "properties": {"name": {"type": "string"}, "why": {"type": "string"},
                               "spec": {"type": "object", "additionalProperties": True}}}}}},
    },
}


def _extract_json(text):
    """Tolerant: return the first JSON object in the text (handles stray prose/fences)."""
    try:
        return json.loads(text)
    except (ValueError, TypeError):
        m = re.search(r"\{.*\}", text or "", re.S)
        return json.loads(m.group(0)) if m else None


def build_payload(profile, sample_rows, header, filename, registries, full_rows=None):
    """The user message: profile + sample (or full) data + the Report Builder vocabulary."""
    body = {
        "filename": filename,
        "profile": profile,
        "sample_rows": sample_rows,
        "report_builder_vocabulary": registries,
    }
    if full_rows is not None:
        body["all_rows"] = {"header": header, "rows": full_rows}
    return ("Read this dataset and return the JSON object described by the schema.\n\n"
            + json.dumps(body, ensure_ascii=False, default=str))


def analyze(header, rows, filename="dataset.csv", registries=None, full=True, _client=None):
    """Profile the data, ask Claude to read it, return the validated dashboard/analysis dict.
    Returns {"error": ...} on any failure so the caller can fall back to the deterministic path.
    `_client` lets tests inject a stub; in production we lazily construct the Anthropic client."""
    if not enabled():
        return {"error": "llm-disabled", "detail": "ANTHROPIC_API_KEY not set"}
    profile = profile_columns(header, rows)
    sample = [dict(zip(header, r)) for r in rows[:25]]
    registries = registries or {}
    full_rows = rows if full else None
    payload = build_payload(profile, sample, header, filename, registries, full_rows)
    try:
        client = _client
        if client is None:
            import anthropic  # lazy
            client = anthropic.Anthropic()
        with client.messages.stream(
            model=MODEL, max_tokens=12000,
            system=[{"type": "text", "text": SYSTEM, "cache_control": {"type": "ephemeral"}}],
            thinking={"type": "adaptive"},
            output_config={"effort": "high",
                           "format": {"type": "json_schema", "schema": SCHEMA}},
            messages=[{"role": "user", "content": payload}],
        ) as s:
            msg = s.get_final_message()
        text = "".join(b.text for b in msg.content if getattr(b, "type", None) == "text")
        data = _extract_json(text)
        if not data:
            return {"error": "parse-failed", "raw": (text or "")[:400]}
        data["_profile"] = profile   # hand the deterministic profile back for rendering
        return data
    except Exception as e:
        return {"error": "analyze-failed", "detail": str(e)[:200]}


# ---------------- AI read (file-as-root, causation-vs-correlation) ----------------
# For data opened on its OWN terms (not the bev-alc model). Hoodie's pulls are PUBLIC scraped
# data, so the actual rows go to the model (the N≥10 / rows-never-leave rule is a SipSource
# constraint, not a Hoodie one). EXACT computed aggregates are sent alongside so precise
# figures are cited rather than eyeballed.
SYSTEM_AIREAD = (
    "You are a rigorous analyst reading a table on its OWN terms — it is the root; assume no "
    "fixed schema. This is public data, so you get the actual ROWS (or a sample of them), a "
    "column PROFILE, and a set of EXACT COMPUTED AGGREGATES. "
    "For any precise figure, cite the computed aggregates (they are exact); read the rows for "
    "qualitative patterns, examples, and data-quality issues. Do not fabricate figures. "
    "SURFACE WHAT'S WORTH SEEING UP FRONT — be selective, not exhaustive (no 10-page dump). "
    "Besides findings, propose 2-4 TABLES that highlight something notable: each is a single "
    "group-by — a `dimension` column to group on (prefer lower-cardinality columns from the "
    "profile), a `measure` column, and an `agg` (sum/mean/min/max/count). Use ONLY exact column "
    "names present in the profile. You choose the cut and say WHY it matters; the app computes "
    "the real numbers — never put numbers in the table spec yourself. "
    "CAUSATION vs CORRELATION (strict): report associations by default; NEVER assert causation "
    "without an explicit identification strategy (experiment / quasi-experiment, or clear "
    "temporal order + mechanism). Use hedged language for observational patterns ('is associated "
    "with', not 'drives'/'causes'). Explicitly flag confounders, reverse causality, and "
    "selection / survivorship bias where relevant. "
    "Be concise and concrete. Return ONLY the JSON object described by the schema."
)
SCHEMA_AIREAD = {
    "type": "object", "additionalProperties": False,
    "required": ["headline", "findings"],
    "properties": {
        "headline": {"type": "string"},
        "what_is_this": {"type": "string"},
        "findings": {"type": "array", "items": {"type": "object", "additionalProperties": False,
            "properties": {"text": {"type": "string"},
                           "kind": {"type": "string",
                                    "enum": ["association", "opportunity", "caveat", "data_quality"]}}}},
        "tables": {"type": "array", "items": {"type": "object", "additionalProperties": False,
            "properties": {"title": {"type": "string"}, "why": {"type": "string"},
                           "dimension": {"type": "string"}, "measure": {"type": "string"},
                           "agg": {"type": "string", "enum": ["sum", "mean", "min", "max", "count", "distinct"]},
                           "sort": {"type": "string", "enum": ["desc", "asc"]},
                           "limit": {"type": "integer"}}}},
        "caveats": {"type": "array", "items": {"type": "string"}},
        "questions": {"type": "array", "items": {"type": "string"}},
    },
}


def ai_read(profile, summary=None, filename="dataset.csv", header=None, rows=None, _client=None):
    """Interpret a dataset from its actual rows (public data) + profile + exact computed
    aggregates. Returns the validated analysis dict, or {"error": ...} for a graceful state."""
    if not enabled():
        return {"error": "llm-disabled", "detail": "ANTHROPIC_API_KEY not set"}
    body = {"filename": filename, "profile": profile, "computed_aggregates": summary or {}}
    if header and rows:
        cap = 3000
        body["data"] = {"header": header, "rows": rows[:cap],
                        "rows_shown": min(len(rows), cap), "rows_total": len(rows)}
    payload = ("Read this dataset on its own terms and return the JSON object in the schema.\n\n"
               + json.dumps(body, ensure_ascii=False, default=str))
    try:
        client = _client
        if client is None:
            import anthropic  # lazy
            client = anthropic.Anthropic()
        with client.messages.stream(
            model=MODEL, max_tokens=6000,
            system=[{"type": "text", "text": SYSTEM_AIREAD, "cache_control": {"type": "ephemeral"}}],
            thinking={"type": "adaptive"},
            output_config={"effort": "high",
                           "format": {"type": "json_schema", "schema": SCHEMA_AIREAD}},
            messages=[{"role": "user", "content": payload}],
        ) as s:
            msg = s.get_final_message()
        text = "".join(b.text for b in msg.content if getattr(b, "type", None) == "text")
        data = _extract_json(text)
        return data or {"error": "parse-failed", "raw": (text or "")[:400]}
    except Exception as e:
        return {"error": "ai-read-failed", "detail": str(e)[:200]}
