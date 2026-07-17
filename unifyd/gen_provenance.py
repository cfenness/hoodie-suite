"""Rebuild bev_alc_data_sources_provenance.xlsx (updated + new sources + Inventory signal split) AND emit
provenance.json for the MDM page. Single source of truth for both."""
import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Columns: category, source, origin, method, provides, inventory (Counts|Availability|—), status, size, notes
# "inventory" separates numeric COUNTS (units on hand) from AVAILABILITY (in/out, no number).
S = [
 # ── Master / Reference ──
 ("Master / Reference","TTB COLA (label registry)","TTB federal label-approval registry","Custom scraper (warmed-cookie session, no BD) + label-image OCR","Items, brand/bottle identity, label images, ABV/UPC (OCR)","—","Live","~1M bottle filings (7-day incremental)","National product identity backbone; detail has no ABV — labels OCR'd for ABV/UPC"),
 ("Master / Reference","US Census (ACS + CBP)","US Census Bureau APIs","Official public API","Trade-area demographics + establishment counts by county","—","Live","All US counties","Joins to outlets by county FIPS for market context"),
 # ── Outlet spine (licenses / rosters) ──
 ("Outlet spine","AB InBev retailer locator","api.beertech.com (no-auth GraphQL)","Direct GraphQL, national zip sweep, dedup by vpid","Outlets (name/address/geo) + brand carriage","Availability","Live","~180k outlets (national)","Universal-outlet spine — Bud is in ~every outlet. vpid is a corroborating attribute, NOT the master spine"),
 ("Outlet spine","VTInfo (Tito's locator)","VTInfo where-to-buy","Direct locator","Outlets (geo) + carriage","Availability","Live","~340 outlets","Shares the vpid identity with AB — cross-reference"),
 ("Outlet spine","California (ABC)","CA Dept of Alcoholic Beverage Control","Roster download","Outlets (license type → sell-flags, address)","—","Live","~96k licenses","Biggest license roster; license type drives beer/wine/spirits flags"),
 ("Outlet spine","Florida (DBPR/ABT)","FL Dept of Business & Professional Reg","Government CSV","Outlets, items","—","Live","FL licensees","Home-state outlet master"),
 ("Outlet spine","Texas (TABC)","TX open data (Socrata)","Official API","Outlets, companies","—","Live","TX licensees","Socrata"),
 ("Outlet spine","New York (SLA)","NY open data (Socrata)","Official API","Outlets (pre-geocoded)","—","Live","NY licensees","Socrata"),
 ("Outlet spine","Chicago / IL","Chicago open data","Official API","Outlets (geocoded)","—","Live","Chicago licensees","Socrata"),
 ("Outlet spine","Colorado","CO open data (Socrata)","Official API","Outlets (geocoded)","—","Live","CO licensees","Socrata"),
 ("Outlet spine","Missouri","MO open data (Socrata)","Official API","Outlets","—","Live","MO licensees","Socrata"),
 ("Outlet spine","Connecticut (DCP)","CT Dept of Consumer Protection","Government data","Outlets, brand registrations","—","Live","CT licensees","Brand registry too"),
 ("Outlet spine","City business licenses (LA/SD/SF/Sac/SJ)","CA city open-data portals","Official APIs","Business licenses (NAICS + geo)","—","Available","Per-city","Supplemental CA outlet geo"),
 ("Outlet spine","Outback / on-premise (DoorDash geo)","DoorDash Maps + merchant pages","BD Maps + GraphQL (occasional)","Outlets (name/address/geo/hours), on-premise menus","Availability","Built","Per-metro (Orlando ~320)","Market-intel sweep, not daily"),
 # ── Retail chains (owned scrapers) ──
 ("Retail chain","Walmart","walmart.com","DIRECT — mobile UA past PerimeterX + __NEXT_DATA__ (search + /ip/ detail)","Items, price, rich attrs (varietal/region/vintage/ABV/container), aisle","Availability","Live","bev-alc catalog","OFF Bright Data as of this session; detail page = RNDC-style vector + aisle inventory"),
 ("Retail chain","Total Wine","totalwine.com","DIRECT — mobile UA crawl + getProduct API (warmed PX cookie)","Items, price, geo attrs, exact per-store unit counts","Counts","Live","~100k SKUs","getProduct = exact per-store units + ABV + shelf location"),
 ("Retail chain","Kroger (+ banners)","Kroger Developer API + internal atlas endpoint","Official OAuth2 API (status) + atlas /product/v2 (counts, x-laf-object header)","Items, outlets, price, stock status; EXACT per-store units via atlas (availableToSell)","Counts","Live","bev-alc by store","Official API = HIGH/LOW/OOS status only; the internal atlas endpoint (kroger_atlas.py) carries inventorySummaries[].details[].availableToSell = a real number + dims + planogram"),
 ("Retail chain","Target","redsky.target.com","RedSky JSON API (via BD Unlocker — Akamai)","Items, outlets, price, numeric per-store inventory","Counts","Live","bev-alc by store","REAL numeric per-store count. On light BD (Akamai captcha, no mobile gap)"),
 ("Retail chain","ABC Fine Wine & Spirits (FL)","abcfws.com (BigCommerce)","Direct (sitemap + product pages, robots-polite) + Storefront GraphQL","Items, outlets, price, EXACT per-store units (availableToSell)","Counts","Live","~13.9k products","Direct by default (BD optional). BigCommerce Storefront GraphQL inventory.aggregated.availableToSell = real bottle count (ABC_QTY=1); HTML in/out is the fallback"),
 ("Retail chain","Spec's (TX)","specsonline.com","Direct (BD optional fallback) + inventory-count API","Items, outlets, price, EXACT per-store units (available)","Counts","Live","~41k products","Direct by default. Per-store count WIRED: GET /api/products/stock/{storeCode}-{upc}/ → {status, available:N, tracked} (85 @ store0, 128 @ store35). SPECS_QTY=1 default; SPECS_COUNT_STORES focuses the store set (in-stock fan-out is ~174 stores/product). Variants block still gives in/out for all ~190 stores in one fetch"),
 ("Retail chain","Binny's (IL)","binnys.com","Direct — free public Algolia search key","Items, outlets, price, units","Counts","Live","~22k products","Algolia public API — no anti-bot, no BD"),
 ("Retail chain","7-Eleven (7NOW)","7now.com (first-party delivery BFF)","Direct — warmed Incapsula cookie (reese84 JS-readable) → conv/inventory/subcategories","Items, price, EXACT per-store on-hand (store_quantity), UPC, deals, age flag","Counts","Live","~9k stores; beer/wine + hemp where carried","store_quantity is uncapped (availableQuantity=order cap ≤100). Deals captured. Non-alc: Drinks dept included"),
 ("Retail chain","Haskell's (MN)","haskells.com (BigCommerce)","Direct — sitemap + Storefront GraphQL (public JWT)","Items, price, EXACT on-hand (availableToSell), UPC","Counts","Live","~10.5k products","Same recipe as ABC FWS. Sitemap = full catalog (no pagination-100 cap). Live /thc/ category → hemp flagged"),
 ("Retail chain","Stop & Shop (+ Ahold banners)","stopandshop.com (Ahold/Peapod API)","Warmed session (robots disallows /api/, anti-bot)","Items, UPC, price, planogram (aisle/section/pick location), nutrition, bottle-deposit","Availability","Built (parser)","Northeast per-store","Peapod platform → Giant/Hannaford/Food Lion too. outOfStock bool, no count. ToS-sensitive (/api/ disallowed)"),
 ("Retail chain","H-E-B (TX)","heb.com (React + GraphQL)","TBD — GraphQL (robots disallows /graphql), sitemap open","Items, price, per-store (TX/MX)","—","Target","TX grocery","Recon needed: GraphQL is robots-blocked; sitemap allowed. Anti-bot. Beer/wine strong in TX"),
 ("Retail chain","Aldi (US)","aldi.us (API)","TBD (robots disallows /api/)","Items, price, mostly private-label wine","—","Target","Select states","Recon needed. Alcohol = private-label (Winking Owl etc.) in wine-permitted states. ToS-sensitive"),
 # ── Distributor / importer tier (the TRANSFERABLE supplier signal — see ttb-not-spine) ──
 ("Distributor / importer","Winebow","winebow.com /our-brands (Drupal)","Direct — server-rendered portfolio, polite, stdlib","Importer→brand portfolio (~1,394 brands) + producer website per brand","—","Built","~1.4k brands","winebow.py → winebow_brands. The supplier-tier grouping the TTB plant/vendor codes couldn't give; feeds producer_site enrichment. FIRST of the importer-portfolio class (same shape generalizes)"),
 ("Distributor / importer","Breakthru Beverage","breakthrubev.com (brand portfolio)","TBD — portfolio scrape (winebow.py pattern: base + card regex)","Distributor→brand portfolio (national wine/spirits/beer)","—","Target","National (a top-3 US distributor)","Supplier-tier signal for the transferable spine. Recon: find the /brands portfolio surface"),
 ("Distributor / importer","Opici Family Distributing","opici.com (brand portfolio)","TBD — portfolio scrape (winebow.py pattern)","Importer/distributor→brand portfolio","—","Target","Northeast US importer/distributor","Same portfolio-scrape shape as Winebow"),
 ("Distributor / importer","Southern Glazer's (SGWS)","southernglazers.com / Proof","TBD — portfolio (largest US distributor; may be login-gated Proof platform)","Distributor→brand portfolio; possibly per-market catalog","—","Target","National (the largest US distributor)","Biggest of the 'etc.' — portfolio if public; ordering catalog (Proof) is account-gated"),
 ("Distributor / importer","RNDC (Republic National)","rndc-usa.com (brand portfolio)","TBD — portfolio scrape","Distributor→brand portfolio","—","Target","National (a top-2 US distributor)","Second major of the 'etc.'"),
 # ── Off-premise independents (platform recipes) ──
 ("Off-premise independents","Shopify / WooCommerce / Wix / Squarespace","Independent retailer sites","Direct platform recipes (Google-Maps census → per-store catalog)","Items, per-store catalog, price, hemp flag","Availability","Live","~180 stores / ~293k products","Free direct-fetch sweep; each store = an account"),
 ("Off-premise independents","City Hive","City Hive platform (~2000 retailers)","Direct SEO surface (sitemap + server-rendered JSON-LD/OpenGraph product pages, via BD Unlocker)","Items, price, size, per-store catalog","—","Live","~2000 retailers; Top Ten Liquors (MN, 6.4k) + Liquor Barn (KY) wired","Full catalog + price from the SEO surface, no browser/session. Per-store COUNT is session-walled (widget API). Named chains → cityhive_chain_products"),
 ("Off-premise independents","Bottlecapps","Bottlecapps platform","Direct platform recipe","Items, price","Availability","Built","Per-store","Off-premise platform recipe"),
 # ── Control-state price/inventory ──
 ("Control-state","Montgomery County MD","Montgomery County ABS","Government data","Items, price, on-hand inventory","Counts","Live","County catalog","On-hand counts"),
 ("Control-state","Utah (DABS)","UT DABS","Government data","Items, price, per-store availability","Availability","Live","State catalog","Per-store availability"),
 ("Control-state","Oregon (OLCC)","OR OLCC","Government data","Items, price (bottle/case)","—","Live","State price list",""),
 ("Control-state","British Columbia","BC Liquor","Government data","Items, price, UPC, ABV","—","Live","Province catalog","Rich item fields (UPC + ABV)"),
 ("Control-state","Idaho","ID","Government data","Items, price, availability","Availability","Live","State catalog",""),
 ("Control-state","North Carolina","NC ABC","Government data","Items, price","—","Live","State catalog",""),
 ("Control-state","Maine","ME","Government data","Items, price, UPC","—","Live","State catalog","UPC"),
 ("Control-state","Montana","MT","Government data","Items, price","—","Live","State catalog",""),
 ("Control-state","Alabama","AL ABC","Government data","Items, price","—","Live (select)","State catalog",""),
 ("Control-state","Iowa","IA (BigQuery)","Public BigQuery","Items, transaction sales (movement)","Counts","Planned","Statewide sales","Depletion/movement, not on-hand"),
 ("Control-state","Vermont","VT","Gov (Unlocker-gated)","Items, price","—","Planned","State catalog","Needs Unlocker"),
 ("Control-state","PA / Virginia / Ohio / New Hampshire","State control boards","Gov (Unlocker-gated)","Items, price, per-store inventory","Counts","Planned","Per-state","Some give per-store counts"),
 # ── Delivery aggregators (occasional market-intel, not daily) ──
 ("Delivery aggregator","Publix","Publix weekly-ad API","BD Browser (Akamai)","Items, price, promo/BOGO","Availability","Built","Weekly ad","Weekly-ad/savings feed"),
 ("Delivery aggregator","Instacart","instacart.com","BD Browser (Forter/reCAPTCHA)","Items, outlets, price, availability","Availability","In progress","Zone-gated catalog","Anti-bot solved; last-mile landing"),
 ("Delivery aggregator","DoorDash / UberEats","aggregator GraphQL","BD Browser + Unlocker","Items, outlets, price, availability","Availability","In progress","Zone-gated",""),
 ("Delivery aggregator","Circle K / CVS / Walgreens","chain + aggregator","BD (planned)","Items, outlets, price","Availability","Planned","National c-store/pharmacy","7-Eleven is NOT here — it's LIVE with exact counts via 7NOW (see the 7-Eleven (7NOW) Retail-chain row)"),
 # ── Vendor feed ──
 ("Vendor feed","Grepsr hemp-bev feed","Daily S3 delivery","Managed feed","Availability, price (hemp-bev, 10+ sources)","Availability","Not ingested","Daily by 05:00 UTC","1 bespoke scraper/source model; superseded by our direct off-premise for most"),
]

LEGEND = [("Live","Pulling now, in our data"),("Live (select)","Live but a partial subset"),
          ("Built","Connector built; run on demand"),("In progress","Being built / last-mile"),
          ("Available","Data source identified, not yet wired"),("Planned","On the roadmap"),
          ("Not ingested","Feed exists, not landed")]
INV_LEGEND=[("Counts","Numeric units on hand / sales (a real number)"),
            ("Availability","In / out of stock (no count)"),("—","No inventory signal (identity/price/outlet only)")]

HEAD = ["Category","Source","Where it comes from (origin)","How we get it (method)","What it provides",
        "Inventory signal","Status","Expected size of source","Notes"]

if __name__ == "__main__":
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "data_sources"
    hf = Font(bold=True, color="FFFFFF"); hfill = PatternFill("solid", fgColor="1E5F55")
    ws.append(HEAD)
    for c in ws[1]:
        c.font = hf; c.fill = hfill; c.alignment = Alignment(vertical="top", wrap_text=True)
    for row in S:
        ws.append(list(row))
    widths = [18, 30, 30, 34, 40, 13, 13, 22, 44]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for r in ws.iter_rows(min_row=2):
        for c in r:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    lg = wb.create_sheet("legend")
    lg.append(["Status", "Meaning"])
    for r in LEGEND:
        lg.append(list(r))
    lg.append([]); lg.append(["Inventory signal", "Meaning"])
    for r in INV_LEGEND:
        lg.append(list(r))
    dst = "/Users/chrisfennessey/Downloads/bev_alc_data_sources_provenance.xlsx"
    wb.save(dst)
    payload = {"head": HEAD, "rows": [list(r) for r in S], "legend": LEGEND, "inv_legend": INV_LEGEND}
    json.dump(payload, open("/private/tmp/prov.json", "w"))
    # Sync the MDM Provenance surface (embeds the table inline) so it can never drift from this generator.
    import os, re
    html_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "apps", "mdm-provenance.html")
    if os.path.exists(html_path):
        html = open(html_path).read()
        line = "const PROV=" + json.dumps(payload, ensure_ascii=True, separators=(",", ":")) + ";"
        html2 = re.sub(r"const PROV=\{.*?\};", lambda _m: line, html, count=1, flags=re.S)
        if html2 != html:
            open(html_path, "w").write(html2)
            print("synced apps/mdm-provenance.html")
    print("wrote %s (%d sources) + /private/tmp/prov.json" % (dst, len(S)))
    # quick tallies
    from collections import Counter
    print("by category:", dict(Counter(r[0] for r in S)))
    print("by inventory:", dict(Counter(r[5] for r in S)))
